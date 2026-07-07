#!/usr/bin/env python3
"""
EQ Map Editor v0.4

This version starts from the v0.3 feature set and adds:

- Edit Mode safety toggle:
  - Select Only
  - Move Points
  - Move Lines
  - Move Line Endpoints
- Click a line and move the whole line segment.
- Show draggable endpoint handles on selected lines.
- Move endpoint 1 or endpoint 2 separately.
- Save moved line records back to the source text file.
- Undo / Redo for:
  - Move point
  - Edit point label
  - Edit point colour
  - Move line
  - Move line endpoint
  - Edit line colour
- Dirty-file indicator.
- Prompt before closing with unsaved edits.
- Reload from disk.
- Restore from timestamped backup.
- Light/dark background.
- Display-only Y flip.
"""

from __future__ import annotations

import argparse
import csv
import difflib
import json
import math
import os
import re
import traceback
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional, Any

from PySide6.QtCore import Qt, QPointF, QRectF, QRect, QTimer, QSize
from PySide6.QtGui import QAction, QActionGroup, QColor, QPainter, QPainterPath, QPen, QBrush, QFont, QPixmap, QIcon, QPolygonF
from PySide6.QtWidgets import (
    QApplication,
    QColorDialog,
    QComboBox,
    QFileDialog,
    QFormLayout,
    QGraphicsEllipseItem,
    QGraphicsItem,
    QGraphicsLineItem,
    QGraphicsScene,
    QGraphicsTextItem,
    QGraphicsView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QDoubleSpinBox,
    QSpinBox,
    QAbstractSpinBox,
    QSplitter,
    QStatusBar,
    QToolBar,
    QToolButton,
    QVBoxLayout,
    QWidget,
    QGroupBox,
    QCheckBox,
    QInputDialog,
    QListWidget,
    QListWidgetItem,
    QDialog,
    QDialogButtonBox,
    QTabWidget,
    QStackedWidget,
    QMenu,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QAbstractItemView,
    QScrollArea,
    QSizePolicy,
    QStyledItemDelegate,
    QStyle,
)


VERSION = "v1.1.27-npc-table-filters-delete"
APP_ROOT = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
BUNDLE_ROOT = Path(getattr(sys, "_MEIPASS", APP_ROOT))

# In PyInstaller one-folder builds, bundled data usually lives in dist/EQMapEditor/_internal/.
# Keep user-writable folders beside the EXE/source, but look for bundled resources in both places.
BUNDLED_ROOT_CANDIDATES = [
    APP_ROOT,
    APP_ROOT / "_internal",
    BUNDLE_ROOT,
    Path(__file__).resolve().parent,
]

LOGS_DIR = APP_ROOT / "logs"
SETTINGS_DIR = APP_ROOT / "settings"
BACKUPS_DIR = APP_ROOT / "backups"
RESOURCES_DIR = APP_ROOT / "resources"
BUNDLED_RESOURCES_DIRS = [root / "resources" for root in BUNDLED_ROOT_CANDIDATES]
PALETTES_DIR = APP_ROOT / "palettes"
BUNDLED_PALETTES_DIRS = [root / "palettes" for root in BUNDLED_ROOT_CANDIDATES]
APP_ICON_PATH = RESOURCES_DIR / "eq_maps_icon.png"
APP_ICON_ICO_PATH = RESOURCES_DIR / "eq_maps_icon.ico"
SETTINGS_PATH = SETTINGS_DIR / "eq_map_editor_settings.json"
LOG_PATH = LOGS_DIR / "eq_map_editor.log"


EXPANSION_OPTIONS = [
    ("Classic", 0),
    ("The Ruins of Kunark", 1),
    ("The Scars of Velious", 2),
    ("The Shadows of Luclin", 3),
    ("The Planes of Power", 4),
    ("Legacy of Ykesha", 5),
    ("Lost Dungeons of Norrath", 6),
    ("Gates of Discord", 7),
    ("Omens of War", 8),
    ("Dragons of Norrath", 9),
    ("Depths of Darkhollow", 10),
    ("Prophecy of Ro", 11),
    ("The Serpent's Spine", 12),
    ("The Buried Sea", 13),
    ("Secrets of Faydwer", 14),
    ("Seeds of Destruction", 15),
    ("Underfoot", 16),
    ("House of Thule", 17),
    ("Veil of Alaris", 18),
    ("Rain of Fear", 19),
    ("Call of the Forsaken", 20),
    ("The Darkened Sea", 21),
    ("The Broken Mirror", 22),
    ("Empires of Kunark", 23),
    ("Ring of Scale", 24),
    ("The Burning Lands", 25),
    ("Torment of Velious", 26),
    ("Claws of Veeshan", 27),
    ("Terror of Luclin", 28),
    ("Night of Shadows", 29),
    ("Laurion's Song", 30),
    ("The Outer Brood", 31),
]

def expansion_label(number: Optional[int]) -> str:
    if number is None or number < 0:
        return "Unknown"
    for name, value in EXPANSION_OPTIONS:
        if value == number:
            return f"{name} ({value})"
    return str(number)

def npc_is_valid_for_expansion(npc: "NpcDataRow", selected_expansion: int) -> bool:
    """Return True when an NPC should exist in the selected expansion era.

    Unknown/negative min values are treated as no known lower bound.
    Unknown/negative max values are treated as no known upper bound.
    """
    min_num = npc.min_expansion_number
    max_num = npc.max_expansion_number
    if min_num is not None and min_num >= 0 and min_num > selected_expansion:
        return False
    if max_num is not None and max_num >= 0 and max_num < selected_expansion:
        return False
    return True

for _folder in (LOGS_DIR, SETTINGS_DIR, BACKUPS_DIR, RESOURCES_DIR, PALETTES_DIR):
    _folder.mkdir(parents=True, exist_ok=True)


def first_existing_path(relative_path: str, roots: list[Path]) -> Optional[Path]:
    for root in roots:
        candidate = root / relative_path
        if candidate.exists():
            return candidate
    return None


def set_windows_app_user_model_id() -> None:
    """Set a stable Windows AppUserModelID so the taskbar uses this app's icon."""
    if os.name != "nt":
        return
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("EQMaps.Editor")
    except Exception:
        pass


def app_icon() -> QIcon:
    icon_roots = BUNDLED_RESOURCES_DIRS + [RESOURCES_DIR]
    ico_path = first_existing_path("eq_maps_icon.ico", icon_roots)
    png_path = first_existing_path("eq_maps_icon.png", icon_roots)

    # Windows title bars/taskbars behave best with .ico.
    if os.name == "nt" and ico_path:
        return QIcon(str(ico_path))
    if png_path:
        return QIcon(str(png_path))
    if ico_path:
        return QIcon(str(ico_path))
    return QIcon()


def resource_icon(relative_path: str, fallback_kind: str = "") -> QIcon:
    """Load an SVG/PNG/ICO resource from bundled or source resource folders."""
    icon_path = first_existing_path(relative_path, BUNDLED_RESOURCES_DIRS + [RESOURCES_DIR])
    if icon_path:
        icon = QIcon(str(icon_path))
        if not icon.isNull():
            return icon
    return make_canvas_icon(fallback_kind) if fallback_kind else QIcon()


# Built-in zone display names. Based on the RedGuides / ReadGuides zone short-name reference.
# The folder scanner also includes unknown shortnames it finds in the selected map folder.
ZONE_SHORTNAME_TO_FULLNAME = {
    "qeynos": "South Qeynos", "qeynos2": "North Qeynos", "qrg": "Surefall Glade",
    "qeytoqrg": "Qeynos Hills", "highkeep": "HighKeep", "freportn": "North Freeport",
    "freportw": "West Freeport", "freporte": "East Freeport", "runnyeye": "Clan RunnyEye",
    "qey2hh1": "West Karana", "northkarana": "North Karana", "southkarana": "South Karana",
    "eastkarana": "East Karana", "beholder": "Gorge of King Xorbb", "blackburrow": "BlackBurrow",
    "paw": "Infected Paw", "rivervale": "Rivervale", "kithicor": "Kithicor Forest (A)",
    "commons": "West Commonlands", "ecommons": "East Commonlands", "erudnint": "Erudin Palace",
    "erudnext": "Erudin", "nektulos": "Nektulos Forest", "cshome": "Sunset Home",
    "lavastorm": "Lavastorm Mountains", "nektropos": "Nektropos", "halas": "Halas",
    "everfrost": "Everfrost Peaks", "soldunga": "Solusek's Eye", "soldungb": "Nagafen's Lair",
    "misty": "Misty Thicket (A)", "nro": "North Ro (A)", "sro": "South Ro (A)",
    "befallen": "Befallen (A)", "oasis": "Oasis of Marr", "tox": "Toxxulia Forest",
    "hole": "The Ruins of Old Paineel", "neriaka": "Neriak Foreign Quarter", "neriakb": "Neriak Commons",
    "neriakc": "Neriak Third Gate", "neriakd": "Neriak Palace", "najena": "Najena",
    "qcat": "Qeynos Catacombs", "innothule": "Innothule Swamp (A)", "feerrott": "The Feerrott (A)",
    "cazicthule": "Cazic-Thule", "oggok": "Oggok", "rathemtn": "Mountains of Rathe",
    "lakerathe": "Lake Rathetear", "grobb": "Grobb", "aviak": "Aviak Village",
    "gfaydark": "The Greater Faydark", "akanon": "Ak'Anon", "steamfont": "Steamfont Mountains",
    "lfaydark": "The Lesser Faydark", "crushbone": "Clan Crushbone", "mistmoore": "Castle Mistmoore",
    "kaladima": "Kaladim (A)", "felwithea": "Felwithe (A)", "felwitheb": "Felwithe (B)",
    "unrest": "Estate of Unrest", "kedge": "Kedge Keep", "guktop": "Upper Guk",
    "gukbottom": "Lower Guk", "kaladimb": "Kaladim (B)", "butcher": "Butcherblock Mountains",
    "oot": "Ocean of Tears", "cauldron": "Dagnor's Cauldron", "airplane": "Plane of Sky",
    "fearplane": "Plane of Fear", "permafrost": "Permafrost Keep", "kerraridge": "Kerra Isle",
    "paineel": "Paineel", "hateplane": "The Plane of Hate", "arena": "The Arena (A)",
    "soltemple": "Temple of Solusek Ro", "erudsxing": "Erud's Crossing", "stonebrunt": "Stonebrunt Mountains",
    "warrens": "The Warrens", "bazaar": "The Bazaar", "bazaar2": "The Bazaar (2)",
    "arena2": "The Arena (B)", "jaggedpine": "The Jaggedpine Forest", "nedaria": "Nedaria's Landing",
    "tutorial": "Tutorial Zone", "load": "Loading (A)", "load2": "Loading (B)", "hateplaneb": "The Plane of Hate",
    "shadowrest": "Shadowrest", "tutoriala": "The Mines of Gloomingdeep (A)",
    "tutorialb": "The Mines of Gloomingdeep (B)", "clz": "Loading (C)", "poknowledge": "Plane of Knowledge",
    "soldungc": "The Caverns of Exile", "guildlobby": "The Guild Lobby", "barter": "The Barter Hall",
    "takishruins": "Ruins of Takish-Hiz", "freeporteast": "East Freeport", "freeportwest": "West Freeport",
    "freeportsewers": "Freeport Sewers", "northro": "North Ro (B)", "southro": "South Ro (B)",
    "highpasshold": "Highpass Hold", "commonlands": "Commonlands", "oceanoftears": "Ocean Of Tears",
    "kithforest": "Kithicor Forest (B)", "befallenb": "Befallen (B)", "highpasskeep": "Highpass Keep",
    "innothuleb": "Innothule Swamp (B)", "toxxulia": "Toxxulia Forest", "mistythicket": "Misty Thicket (B)",
    "steamfontmts": "Steamfont Mountains", "dragonscalea": "Tinmizer's Wunderwerks",
    "crafthalls": "Ngreth's Den", "weddingchapel": "Wedding Chapel", "weddingchapeldark": "Wedding Chapel",
    "dragoncrypt": "Lair of the Fallen", "arttest": "Art Testing Domain", "fhalls": "The Forgotten Halls",
    "fieldofbone": "The Field of Bone", "warslikswood": "Warsliks Wood", "droga": "Temple of Droga",
    "cabwest": "West Cabilis", "swampofnohope": "Swamp of No Hope", "firiona": "Firiona Vie",
    "lakeofillomen": "Lake of Ill Omen", "dreadlands": "Dreadlands", "burningwood": "Burning Woods",
    "kaesora": "Kaesora", "sebilis": "Old Sebilis", "citymist": "City of Mist",
    "skyfire": "Skyfire Mountains", "frontiermtns": "Frontier Mountains", "overthere": "The Overthere",
    "emeraldjungle": "The Emerald Jungle", "trakanon": "Trakanon's Teeth", "timorous": "Timorous Deep",
    "kurn": "Kurn's Tower", "karnor": "Karnor's Castle", "chardok": "Chardok",
    "dalnir": "Dalnir", "charasis": "Howling Stones", "cabeast": "East Cabilis",
    "nurga": "Mines of Nurga", "veeshan": "Veeshan's Peak", "veksar": "Veksar",
    "chardokb": "The Halls of Betrayal", "iceclad": "Iceclad Ocean", "frozenshadow": "Tower of Frozen Shadow",
    "velketor": "Velketor's Labyrinth", "kael": "Kael Drakkal", "skyshrine": "Skyshrine",
    "thurgadina": "Thurgadin", "eastwastes": "Eastern Wastes", "cobaltscar": "Cobalt Scar",
    "greatdivide": "Great Divide", "wakening": "The Wakening Land", "westwastes": "Western Wastes",
    "crystal": "Crystal Caverns", "necropolis": "Dragon Necropolis", "templeveeshan": "Temple of Veeshan",
    "sirens": "Siren's Grotto", "mischiefplane": "Plane of Mischief", "growthplane": "Plane of Growth",
    "sleeper": "Sleeper's Tomb", "thurgadinb": "Icewell Keep", "shadowhaven": "Shadow Haven",
    "nexus": "The Nexus", "echo": "Echo Caverns", "acrylia": "Acrylia Caverns",
    "sharvahl": "Shar Vahl", "paludal": "Paludal Caverns", "fungusgrove": "Fungus Grove",
    "vexthal": "Vex Thal", "sseru": "Sanctus Seru", "katta": "Katta Castellum",
    "netherbian": "Netherbian Lair", "ssratemple": "Ssraeshza Temple", "griegsend": "Grieg's End",
    "thedeep": "The Deep", "shadeweaver": "Shadeweaver's Thicket", "hollowshade": "Hollowshade Moor",
    "grimling": "Grimling Forest", "mseru": "Marus Seru", "letalis": "Mons Letalis",
    "twilight": "The Twilight Sea", "thegrey": "The Grey", "tenebrous": "The Tenebrous Mountains",
    "maiden": "The Maiden's Eye", "dawnshroud": "Dawnshroud Peaks", "scarlet": "The Scarlet Desert",
    "umbral": "The Umbral Plains", "akheva": "Akheva Ruins", "codecay": "Ruins of Lxanvom",
    "pojustice": "Plane of Justice", "potranquility": "Plane of Tranquility",
    "ponightmare": "Plane of Nightmare", "podisease": "Plane of Disease", "poinnovation": "Plane of Innovation",
    "potorment": "Plane of Torment", "povalor": "Plane of Valor", "bothunder": "Torden, The Bastion of Thunder",
    "postorms": "Plane of Storms", "hohonora": "Halls of Honor", "solrotower": "Solusek Ro's Tower",
    "powar": "Plane of War", "potactics": "Drunder, Fortress of Zek", "poair": "Eryslai, the Kingdom of Wind",
    "powater": "Reef of Coirnav", "pofire": "Doomfire, The Burning Lands",
    "poeartha": "Vegarlson, The Earthen Badlands", "potimea": "Plane of Time (A)",
    "hohonorb": "Temple of Marr (A)", "nightmareb": "Lair of Terris Thule",
    "poearthb": "Stronghold of the Twelve", "potimeb": "Plane of Time (B)",
    "gunthak": "Gulf of Gunthak", "dulak": "Dulak's Harbor", "torgiran": "Torgiran Mines",
    "nadox": "Crypt of Nadox", "hatesfury": "Hate's Fury, The Scorned Maiden",
    "abysmal": "Abysmal Sea", "natimbi": "Natimbi, The Broken Shores", "qinimi": "Qinimi, Court of Nihilia",
    "riwwi": "Riwwi, Coliseum of Games", "barindu": "Barindu, Hanging Gardens",
    "ferubi": "Ferubi, Forgotten Temple of Taelosia", "tipt": "Tipt, Treacherous Crags",
    "vxed": "Vxed, The Crumbling Caverns", "yxtta": "Yxtta, Pulpit of Exiles",
    "uqua": "Uqua, The Ocean God Chantry", "kodtaz": "Kod'Taz, Broken Trial Grounds",
    "qvic": "Qvic, Prayer Grounds of Calling", "inktuta": "Inktu`Ta, The Unmasked Chapel",
    "txevu": "Txevu, Lair of the Elite", "tacvi": "Tacvi, Seat of the Slaver",
    "wallofslaughter": "Wall of Slaughter", "bloodfields": "The Bloodfields",
    "draniksscar": "Dranik's Scar", "causeway": "Nobles' Causeway",
    "provinggrounds": "Muramite Proving Grounds", "anguish": "Asylum of Anguish",
    "crescent": "Crescent Reach", "moors": "Blightfire Moors", "stonehive": "Stone Hive",
    "mesa": "Goru`kar Mesa", "roost": "Blackfeather Roost", "steppes": "The Steppes",
    "icefall": "Icefall Glacier", "valdeholm": "Valdeholm", "frostcrypt": "Frostcrypt, Throne of the Shade King",
    "sunderock": "Sunderock Springs", "vergalid": "Vergalid Mines", "direwind": "Direwind Cliffs",
    "ashengate": "Ashengate, Reliquary of the Scale",
}
MAX_BULK_HIGHLIGHTS = 0  # 0 means highlight all matching records


@dataclass
class MapLineRecord:
    file_path: Path
    line_index: int
    raw_text: str
    x1: float
    y1: float
    z1: float
    x2: float
    y2: float
    z2: float
    r: int
    g: int
    b: int
    dirty: bool = False
    deleted: bool = False

    @property
    def color(self) -> QColor:
        return QColor(self.r, self.g, self.b)

    def to_map_line(self) -> str:
        return (
            f"L {self.x1:.4f}, {self.y1:.4f}, {self.z1:.4f}, "
            f"{self.x2:.4f}, {self.y2:.4f}, {self.z2:.4f}, "
            f"{self.r}, {self.g}, {self.b}"
        )


@dataclass
class MapPointRecord:
    file_path: Path
    line_index: int
    raw_text: str
    x: float
    y: float
    z: float
    r: int
    g: int
    b: int
    size: float
    label: str
    dirty: bool = False
    deleted: bool = False

    @property
    def color(self) -> QColor:
        return QColor(self.r, self.g, self.b)

    def to_map_line(self) -> str:
        return (
            f"P {self.x:.4f}, {self.y:.4f}, {self.z:.4f}, "
            f"{self.r}, {self.g}, {self.b}, {self.size:g}, {self.label}"
        )


class ExistingLabelStyleDelegate(QStyledItemDelegate):
    """Draws existing map-label style choices with an inline RGB swatch.

    The combo box still stores the MapPointRecord object as item data, but the
    popup list is easier to scan because each row visually shows the colour that
    would be copied by the Match button.
    """

    def paint(self, painter: QPainter, option, index) -> None:  # type: ignore[override]
        painter.save()
        try:
            point = index.data(Qt.UserRole)
            display_text = str(index.data(Qt.DisplayRole) or "")
            selected = bool(option.state & QStyle.State_Selected)
            if selected:
                painter.fillRect(option.rect, option.palette.highlight())
                text_colour = option.palette.highlightedText().color()
            else:
                text_colour = option.palette.text().color()

            if point is None:
                painter.setPen(text_colour)
                painter.drawText(option.rect.adjusted(8, 0, -4, 0), Qt.AlignVCenter | Qt.AlignLeft, display_text)
                return

            r = max(0, min(255, int(getattr(point, "r", 0))))
            g = max(0, min(255, int(getattr(point, "g", 0))))
            b = max(0, min(255, int(getattr(point, "b", 0))))
            size_value = float(getattr(point, "size", 0.0))
            label = str(getattr(point, "label", display_text)).strip() or display_text

            swatch_rect = QRect(option.rect.left() + 6, option.rect.top() + 4, 22, max(12, option.rect.height() - 8))
            border = QColor(30, 30, 30) if (r + g + b) > 360 else QColor(220, 220, 220)
            painter.setPen(QPen(border, 1))
            painter.setBrush(QBrush(QColor(r, g, b)))
            painter.drawRoundedRect(swatch_rect, 3, 3)

            painter.setPen(text_colour)
            text_rect = option.rect.adjusted(36, 0, -4, 0)
            painter.drawText(
                text_rect,
                Qt.AlignVCenter | Qt.AlignLeft,
                f"{label}  [{r},{g},{b} / {size_value:g}]",
            )
        finally:
            painter.restore()

    def sizeHint(self, option, index) -> QSize:  # type: ignore[override]
        size = super().sizeHint(option, index)
        return QSize(max(size.width(), 220), max(size.height(), 26))


class ExistingLabelStyleCombo(QComboBox):
    """Combo that refreshes its map-label style list right before opening."""

    def showPopup(self) -> None:  # type: ignore[override]
        owner = self.property("owner")
        if owner is not None and hasattr(owner, "rebuild_missing_style_match_combo"):
            owner.rebuild_missing_style_match_combo()
        if self.view() is not None:
            self.view().setMinimumWidth(max(self.width(), 360))
        super().showPopup()


@dataclass
class LoadedMap:
    lines: list[MapLineRecord]
    points: list[MapPointRecord]


@dataclass
class NpcDataRow:
    source_index: int
    zone_shortname: str
    npc_name: str
    npc_role: str
    npc_label: str
    x: float
    y: float
    z: float
    min_expansion_number: Optional[int]
    max_expansion_number: Optional[int]
    min_expansion_name: str
    max_expansion_name: str
    raw: dict[str, Any]


@dataclass
class NpcMatchResult:
    map_record: Optional[MapPointRecord]
    npc_row: Optional[NpcDataRow]
    match_type: str
    confidence: str
    score: float
    distance: Optional[float]
    selected: bool = False


@dataclass
class EraCleanupResult:
    map_record: MapPointRecord
    npc_row: NpcDataRow
    era_status: str
    match_type: str
    confidence: str
    score: float
    distance: Optional[float]
    selected: bool = False


@dataclass
class MissingNpcResult:
    npc_row: NpcDataRow
    npc_label: str
    selected: bool = False


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text in {"\\N", "None", "nan", "NaN"}:
        return ""
    return text


def numeric_cell(value: Any) -> Optional[float]:
    text = clean_cell(value)
    if not text:
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def normalize_npc_match_text(value: str) -> str:
    """Match the batch script's normalize_text(): lowercase and remove non-alphanumerics."""
    if value is None:
        return ""
    value = str(value).strip().lower()
    return re.sub(r"[^a-z0-9]", "", value)


def clean_map_label_for_npc_match(label: str) -> str:
    """Match the batch script: remove trailing role notes like _(Banker)."""
    if label is None:
        return ""
    label = str(label).strip()
    label = re.sub(r"(?:_\([^)]*\))+$", "", label)
    return label.strip()


def get_name_aliases_for_npc_match(value: str) -> set[str]:
    """Build the same alias forms used by extract_eq_map_points.py."""
    if value is None:
        return set()

    raw = str(value).strip()
    cleaned = clean_map_label_for_npc_match(raw)
    aliases: set[str] = set()

    common_prefix_roles = {
        "banker", "merchant", "general", "parcels", "parcel", "guildmaster",
        "gm", "trainer", "vendor", "keeper", "priest", "high", "grandmaster",
        "master", "mistress", "lord", "lady", "warlord", "grave", "savage",
        "oracle", "phantasmist", "virtuoso", "warder", "assassin", "warlock",
        "shaman", "druid", "wizard", "enchanter", "bard", "necromancer",
        "shadow", "knight", "beastlord", "ranger", "cleric", "warrior",
        "rogue", "monk", "paladin", "magician", "berserker",
    }

    for candidate in [raw, cleaned]:
        candidate = candidate.strip()
        if not candidate:
            continue
        aliases.add(candidate)
        parts = [part for part in candidate.split("_") if part]
        if len(parts) >= 2:
            first = parts[0].lower()
            if first in common_prefix_roles:
                aliases.add("_".join(parts[1:]))
            if len(parts[-1]) >= 4:
                aliases.add(parts[-1])

    role_match = re.search(r"_\(([^)]*)\)$", raw)
    if role_match and cleaned:
        role_text = role_match.group(1).strip().split(",")[0].strip()
        if role_text:
            aliases.add(f"{role_text}_{cleaned}")

    return {alias for alias in aliases if alias}


def true_exact_name_match(npc_name: str, map_match_label: str) -> bool:
    return normalize_npc_match_text(npc_name) == normalize_npc_match_text(map_match_label)


def possible_name_match(npc_name: str, map_match_label: str) -> bool:
    npc_aliases = get_name_aliases_for_npc_match(npc_name)
    map_aliases = get_name_aliases_for_npc_match(map_match_label)
    for npc_alias in npc_aliases:
        for map_alias in map_aliases:
            npc_norm = normalize_npc_match_text(npc_alias)
            map_norm = normalize_npc_match_text(map_alias)
            if not npc_norm or not map_norm:
                continue
            if npc_norm == map_norm:
                return True
            if npc_norm.startswith(map_norm) and len(map_norm) >= 4:
                return True
            if map_norm.startswith(npc_norm) and len(npc_norm) >= 4:
                return True
            ratio = difflib.SequenceMatcher(None, npc_norm, map_norm).ratio()
            if ratio >= 0.88:
                return True
    return False


def normalized_key(text: str) -> str:
    # Kept for older UI sorting/search helpers. NPC matching now uses the script-compatible helpers above.
    text = clean_cell(text).lower()
    text = text.replace("`", "").replace("'", "")
    text = re.sub(r"[()\[\]{}]", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(part for part in text.split() if part)


def normalized_tokens(text: str) -> set[str]:
    return set(normalized_key(text).split())


def map_safe_label_part(text: str) -> str:
    text = clean_cell(text)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def generated_npc_label(npc_name: str, npc_role: str) -> str:
    name = map_safe_label_part(npc_name)
    role = map_safe_label_part(npc_role)
    if role:
        return f"{name}_({role})"
    return name


def row_get(row: dict[str, Any], *names: str) -> str:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for name in names:
        if name.lower() in lowered:
            return clean_cell(lowered[name.lower()])
    return ""


def row_get_number(row: dict[str, Any], *names: str) -> Optional[float]:
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for name in names:
        if name.lower() in lowered:
            value = numeric_cell(lowered[name.lower()])
            if value is not None:
                return value
    return None


def row_get_int(row: dict[str, Any], *names: str) -> Optional[int]:
    value = row_get_number(row, *names)
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def point_distance(point: MapPointRecord, npc: NpcDataRow) -> float:
    return math.sqrt((point.x - npc.x) ** 2 + (point.y - npc.y) ** 2 + (point.z - npc.z) ** 2)


def score_point_to_npc(point: MapPointRecord, npc: NpcDataRow) -> tuple[float, str, Optional[float]]:
    """Score a single point/NPC pair using the same status rules as extract_eq_map_points.py.

    Priority:
      1. exact NPC-name-to-map-match-label => Yes
      2. alias/fuzzy possible name match => Possible
      3. if that Possible match is within 20 units => Coordinate Match
      4. no name-based possible match => No

    This intentionally does not coordinate-match unrelated nearby labels.
    """
    map_match_label = clean_map_label_for_npc_match(point.label)
    dist = point_distance(point, npc)
    if true_exact_name_match(npc.npc_name, map_match_label):
        return 100.0, "Yes", dist
    if possible_name_match(npc.npc_name, map_match_label):
        if dist != math.inf and -20 < dist < 20:
            return 96.0, "Coordinate Match", dist
        return 82.0, "Possible", dist
    return 0.0, "No", dist


def confidence_from_score(score: float) -> str:
    if score >= 88:
        return "High"
    if score >= 72:
        return "Medium"
    if score >= 60:
        return "Low"
    return ""


def current_zone_shortname_from_files(file_paths: list[Path]) -> str:
    if not file_paths:
        return ""
    stem = Path(file_paths[0]).stem
    return stem.split("_", 1)[0].lower()


def clamp_rgb(value: str | int) -> int:
    number = int(float(str(value).strip()))
    return max(0, min(255, number))



DEFAULT_MAP_PALETTES: dict[str, dict[str, Any]] = {
    "EQ Map Standard": {
        "name": "EQ Map Standard",
        "description": "Balanced default palette for switching map files between light and dark backgrounds.",
        "entries": [
            {"name": "Wall / Structure", "light": [25, 25, 25], "dark": [210, 210, 210]},
            {"name": "Secondary Wall", "light": [95, 95, 95], "dark": [150, 150, 150]},
            {"name": "Background Detail", "light": [150, 150, 150], "dark": [95, 95, 95]},
            {"name": "Label White", "light": [35, 35, 35], "dark": [255, 255, 255]},
            {"name": "Label Yellow", "light": [150, 120, 0], "dark": [255, 255, 0]},
            {"name": "Important Red", "light": [180, 0, 0], "dark": [255, 0, 0]},
            {"name": "Safe Green", "light": [0, 130, 0], "dark": [0, 240, 0]},
            {"name": "Water / Blue", "light": [0, 90, 200], "dark": [80, 170, 255]},
            {"name": "Magic / Purple", "light": [110, 40, 190], "dark": [190, 110, 255]},
            {"name": "Orange POI", "light": [170, 85, 0], "dark": [255, 145, 0]},
            {"name": "Cyan POI", "light": [0, 130, 150], "dark": [80, 240, 255]},
            {"name": "Soft Green", "light": [70, 160, 70], "dark": [140, 255, 140]},
            {"name": "Vendor", "light": [145, 100, 0], "dark": [255, 190, 60]},
            {"name": "Banker", "light": [80, 80, 80], "dark": [255, 255, 255]},
            {"name": "Monster / Hostile", "light": [170, 0, 0], "dark": [255, 70, 70]},
            {"name": "Guard / Friendly", "light": [0, 110, 0], "dark": [90, 255, 90]},
            {"name": "Zone Connection", "light": [120, 95, 0], "dark": [255, 230, 0]},
            {"name": "Portal / Travel", "light": [95, 40, 180], "dark": [190, 130, 255]},
            {"name": "Water", "light": [0, 75, 180], "dark": [70, 165, 255]},
        ],
    },
    "High Contrast": {
        "name": "High Contrast",
        "description": "Higher contrast lines and labels for very dark or very light backgrounds.",
        "entries": [
            {"name": "Wall / Structure", "light": [0, 0, 0], "dark": [255, 255, 255]},
            {"name": "Secondary Wall", "light": [80, 80, 80], "dark": [185, 185, 185]},
            {"name": "Label White", "light": [0, 0, 0], "dark": [255, 255, 255]},
            {"name": "Label Yellow", "light": [120, 95, 0], "dark": [255, 235, 0]},
            {"name": "Important Red", "light": [190, 0, 0], "dark": [255, 65, 65]},
            {"name": "Safe Green", "light": [0, 120, 0], "dark": [80, 255, 80]},
            {"name": "Water / Blue", "light": [0, 80, 220], "dark": [95, 190, 255]},
            {"name": "Magic / Purple", "light": [120, 0, 220], "dark": [210, 125, 255]},
            {"name": "Orange POI", "light": [190, 90, 0], "dark": [255, 165, 20]},
        ],
    },
}


def normalise_rgb(value: Any) -> tuple[int, int, int]:
    if isinstance(value, (list, tuple)) and len(value) >= 3:
        return (clamp_rgb(value[0]), clamp_rgb(value[1]), clamp_rgb(value[2]))
    raise ValueError(f"Invalid RGB value: {value}")


def palette_entries(palette: dict[str, Any]) -> list[dict[str, Any]]:
    entries = []
    for entry in palette.get("entries", []):
        try:
            entries.append({
                "name": str(entry.get("name", "Colour")),
                "light": normalise_rgb(entry.get("light")),
                "dark": normalise_rgb(entry.get("dark")),
            })
        except Exception:
            continue
    return entries


def load_user_palettes() -> dict[str, dict[str, Any]]:
    palettes: dict[str, dict[str, Any]] = {}
    PALETTES_DIR.mkdir(parents=True, exist_ok=True)

    seen_paths: set[Path] = set()
    for folder in BUNDLED_PALETTES_DIRS + [PALETTES_DIR]:
        if not folder.exists():
            continue
        for path in sorted(folder.glob("*.json")):
            try:
                resolved = path.resolve()
                if resolved in seen_paths:
                    continue
                seen_paths.add(resolved)
                data = json.loads(path.read_text(encoding="utf-8"))
                if "name" in data and "entries" in data:
                    palettes[str(data["name"])] = data
            except Exception:
                continue
    return palettes


def available_palettes() -> dict[str, dict[str, Any]]:
    palettes = dict(DEFAULT_MAP_PALETTES)
    palettes.update(load_user_palettes())
    return palettes


def save_user_palette(palette: dict[str, Any]) -> Path:
    name = str(palette.get("name", "Custom Palette")).strip() or "Custom Palette"
    safe_name = re.sub(r"[^A-Za-z0-9_. -]+", "_", name).strip().replace(" ", "_")
    path = PALETTES_DIR / f"{safe_name}.json"
    path.write_text(json.dumps(palette, indent=2), encoding="utf-8")
    return path


def rgb_distance_sq(a: tuple[int, int, int], b: tuple[int, int, int]) -> int:
    return sum((int(a[i]) - int(b[i])) ** 2 for i in range(3))


def map_rgb_to_palette(rgb: tuple[int, int, int], palette: dict[str, Any], target_mode: str) -> tuple[int, int, int]:
    entries = palette_entries(palette)
    if not entries:
        return rgb
    target_key = "dark" if target_mode.lower().startswith("dark") else "light"
    best_entry = min(
        entries,
        key=lambda entry: min(
            rgb_distance_sq(rgb, entry["light"]),
            rgb_distance_sq(rgb, entry["dark"]),
        ),
    )
    return best_entry[target_key]



def palette_role_names(palette: dict[str, Any]) -> list[str]:
    return [entry["name"] for entry in palette_entries(palette)]


def palette_entry_by_name(palette: dict[str, Any], name: str) -> Optional[dict[str, Any]]:
    for entry in palette_entries(palette):
        if entry["name"] == name:
            return entry
    return None


def infer_point_role_from_labels(labels: list[str], palette: dict[str, Any]) -> Optional[str]:
    combined = " ".join(label.lower() for label in labels if label).strip()
    if not combined:
        return None

    candidates = palette_role_names(palette)
    def choose(*names: str) -> Optional[str]:
        for preferred in names:
            for candidate in candidates:
                if preferred.lower() in candidate.lower():
                    return candidate
        return None

    if any(word in combined for word in ["bank", "banker"]):
        return choose("Banker", "Label White")
    if any(word in combined for word in ["merchant", "vendor", "shop", "sell", "buy"]):
        return choose("Vendor", "Orange", "Label Yellow")
    if any(word in combined for word in ["guard"]):
        return choose("Guard", "Friendly", "Safe Green")
    if any(word in combined for word in ["portal", "translocator", "stone", "book"]):
        return choose("Portal", "Travel", "Magic")
    if any(word in combined for word in ["to ", " zone", "zoneline", "zone line", "harbor", "north", "south", "east", "west"]):
        return choose("Zone Connection", "Label Yellow")
    if any(word in combined for word in ["raid", "named", "monster", "beast", "dragon", "orc", "goblin", "kobold"]):
        return choose("Monster", "Hostile", "Important Red")
    return None


def parse_l_record(row: str, file_path: Path, line_index: int) -> Optional[MapLineRecord]:
    stripped = row.strip()
    if not stripped.startswith("L "):
        return None
    parts = [part.strip() for part in stripped[2:].split(",")]
    if len(parts) < 9:
        return None
    try:
        x1, y1, z1, x2, y2, z2 = [float(value) for value in parts[:6]]
        r, g, b = [clamp_rgb(value) for value in parts[6:9]]
    except ValueError:
        return None
    return MapLineRecord(file_path, line_index, row.rstrip("\n"), x1, y1, z1, x2, y2, z2, r, g, b)


def parse_p_record(row: str, file_path: Path, line_index: int) -> Optional[MapPointRecord]:
    stripped = row.rstrip("\n")
    if not stripped.startswith("P "):
        return None
    parts = [part.strip() for part in stripped[2:].split(",", 7)]
    if len(parts) < 8:
        return None
    try:
        x, y, z = [float(value) for value in parts[:3]]
        r, g, b = [clamp_rgb(value) for value in parts[3:6]]
        size = float(parts[6])
        label = parts[7].strip()
    except ValueError:
        return None
    return MapPointRecord(file_path, line_index, row.rstrip("\n"), x, y, z, r, g, b, size, label)


def load_map_files(file_paths: Iterable[Path]) -> LoadedMap:
    lines: list[MapLineRecord] = []
    points: list[MapPointRecord] = []

    for file_path in file_paths:
        file_path = Path(file_path)
        with file_path.open("r", encoding="utf-8", errors="replace") as handle:
            for line_index, row in enumerate(handle):
                line_record = parse_l_record(row, file_path, line_index)
                if line_record is not None:
                    lines.append(line_record)
                    continue

                point_record = parse_p_record(row, file_path, line_index)
                if point_record is not None:
                    points.append(point_record)

    return LoadedMap(lines=lines, points=points)


class CoordinateMapper:
    def __init__(self, flip_display_y: bool = False) -> None:
        self.flip_display_y = flip_display_y

    def map_to_scene(self, x: float, y: float) -> QPointF:
        return QPointF(x, -y if self.flip_display_y else y)

    def scene_to_map(self, point: QPointF) -> tuple[float, float]:
        return point.x(), -point.y() if self.flip_display_y else point.y()


class UndoCommand:
    def __init__(self, label: str, record: Any, before: dict[str, Any], after: dict[str, Any]) -> None:
        self.label = label
        self.record = record
        self.before = before
        self.after = after

    def undo(self) -> None:
        for key, value in self.before.items():
            setattr(self.record, key, value)
        self.record.dirty = True

    def redo(self) -> None:
        for key, value in self.after.items():
            setattr(self.record, key, value)
        self.record.dirty = True



class AddRecordCommand:
    def __init__(self, label: str, main_window: "EqMapMainWindow", record: Any, collection_name: str) -> None:
        self.label = label
        self.main_window = main_window
        self.record = record
        self.collection_name = collection_name

    @property
    def collection(self):
        return getattr(self.main_window.loaded_map, self.collection_name)

    def undo(self) -> None:
        if self.record in self.collection:
            self.collection.remove(self.record)
        self.main_window.render_map(keep_view=True)
        self.main_window.side_panel.rebuild_layers()
        self.main_window.update_dirty_indicator()

    def redo(self) -> None:
        if self.record not in self.collection:
            self.collection.append(self.record)
        self.record.dirty = True
        self.main_window.render_map(keep_view=True)
        self.main_window.side_panel.rebuild_layers()
        self.main_window.update_dirty_indicator()



class AddRecordsCommand:
    def __init__(self, label: str, main_window: "EqMapMainWindow", records: list[Any], collection_name: str) -> None:
        self.label = label
        self.main_window = main_window
        self.records = records
        self.collection_name = collection_name

    @property
    def collection(self):
        return getattr(self.main_window.loaded_map, self.collection_name)

    def undo(self) -> None:
        for record in self.records:
            if record in self.collection:
                self.collection.remove(record)
        self.main_window.render_map(keep_view=True)
        self.main_window.side_panel.rebuild_layers()
        self.main_window.update_dirty_indicator()

    def redo(self) -> None:
        for record in self.records:
            if record not in self.collection:
                self.collection.append(record)
            record.dirty = True
        self.main_window.render_map(keep_view=True)
        self.main_window.side_panel.rebuild_layers()
        self.main_window.update_dirty_indicator()



class BulkEditCommand:
    def __init__(self, label: str, records: list[Any], before: list[dict[str, Any]], after: list[dict[str, Any]]) -> None:
        self.label = label
        self.records = records
        self.before = before
        self.after = after

    def undo(self) -> None:
        for record, values in zip(self.records, self.before):
            for key, value in values.items():
                setattr(record, key, value)
            record.dirty = True

    def redo(self) -> None:
        for record, values in zip(self.records, self.after):
            for key, value in values.items():
                setattr(record, key, value)
            record.dirty = True


class PointDetailsDialog(QDialog):
    def __init__(self, parent=None, title="Add Point", label="", r=255, g=255, b=255, size=2.0, files: Optional[list[Path]] = None, active_file: Optional[Path] = None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.files = files or []
        layout = QVBoxLayout(self)
        form = QFormLayout()
        layout.addLayout(form)

        self.file_combo = QComboBox()
        for file_path in self.files:
            self.file_combo.addItem(file_path.name, str(file_path))
        if active_file is not None:
            for index in range(self.file_combo.count()):
                if self.file_combo.itemData(index) == str(active_file):
                    self.file_combo.setCurrentIndex(index)
                    break
        form.addRow("Target file", self.file_combo)

        self.label_edit = QLineEdit(label)
        form.addRow("Label", self.label_edit)

        self.r_spin = QSpinBox(); self.r_spin.setRange(0, 255); self.r_spin.setValue(r)
        self.g_spin = QSpinBox(); self.g_spin.setRange(0, 255); self.g_spin.setValue(g)
        self.b_spin = QSpinBox(); self.b_spin.setRange(0, 255); self.b_spin.setValue(b)
        form.addRow("R", self.r_spin)
        form.addRow("G", self.g_spin)
        form.addRow("B", self.b_spin)

        self.size_spin = QDoubleSpinBox()
        self.size_spin.setRange(0.1, 99.0)
        self.size_spin.setDecimals(2)
        self.size_spin.setValue(size)
        form.addRow("Size", self.size_spin)

        self.pick_button = QPushButton("Pick Colour")
        self.pick_button.clicked.connect(self.pick_colour)
        layout.addWidget(self.pick_button)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def pick_colour(self):
        c = choose_colour_dialog(QColor(self.r_spin.value(), self.g_spin.value(), self.b_spin.value()), self, "Choose point colour")
        if c.isValid():
            self.r_spin.setValue(c.red())
            self.g_spin.setValue(c.green())
            self.b_spin.setValue(c.blue())

    def values(self):
        return {
            "file_path": Path(self.file_combo.currentData()) if self.file_combo.currentData() else None,
            "label": self.label_edit.text().strip() or "New_Point",
            "r": self.r_spin.value(),
            "g": self.g_spin.value(),
            "b": self.b_spin.value(),
            "size": self.size_spin.value(),
        }


class LineColorDialog(QDialog):
    def __init__(self, parent=None, r=255, g=255, b=255, files: Optional[list[Path]] = None, active_file: Optional[Path] = None):
        super().__init__(parent)
        self.setWindowTitle("Line Colour")
        self.files = files or []
        layout = QVBoxLayout(self)
        form = QFormLayout()
        layout.addLayout(form)

        self.file_combo = QComboBox()
        for file_path in self.files:
            self.file_combo.addItem(file_path.name, str(file_path))
        if active_file is not None:
            for index in range(self.file_combo.count()):
                if self.file_combo.itemData(index) == str(active_file):
                    self.file_combo.setCurrentIndex(index)
                    break
        form.addRow("Target file", self.file_combo)

        self.r_spin = QSpinBox(); self.r_spin.setRange(0, 255); self.r_spin.setValue(r)
        self.g_spin = QSpinBox(); self.g_spin.setRange(0, 255); self.g_spin.setValue(g)
        self.b_spin = QSpinBox(); self.b_spin.setRange(0, 255); self.b_spin.setValue(b)
        form.addRow("R", self.r_spin)
        form.addRow("G", self.g_spin)
        form.addRow("B", self.b_spin)

        self.pick_button = QPushButton("Pick Colour")
        self.pick_button.clicked.connect(self.pick_colour)
        layout.addWidget(self.pick_button)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def pick_colour(self):
        c = choose_colour_dialog(QColor(self.r_spin.value(), self.g_spin.value(), self.b_spin.value()), self, "Choose line colour")
        if c.isValid():
            self.r_spin.setValue(c.red())
            self.g_spin.setValue(c.green())
            self.b_spin.setValue(c.blue())

    def values(self):
        return {
            "file_path": Path(self.file_combo.currentData()) if self.file_combo.currentData() else None,
            "r": self.r_spin.value(),
            "g": self.g_spin.value(),
            "b": self.b_spin.value(),
        }


def snapshot_point(point: MapPointRecord) -> dict[str, Any]:
    return {
        "x": point.x,
        "y": point.y,
        "z": point.z,
        "r": point.r,
        "g": point.g,
        "b": point.b,
        "size": point.size,
        "label": point.label,
    }


def snapshot_line(line: MapLineRecord) -> dict[str, Any]:
    return {
        "x1": line.x1,
        "y1": line.y1,
        "z1": line.z1,
        "x2": line.x2,
        "y2": line.y2,
        "z2": line.z2,
        "r": line.r,
        "g": line.g,
        "b": line.b,
    }



def make_canvas_icon(kind: str) -> QIcon:
    pixmap = QPixmap(28, 28)
    pixmap.fill(Qt.transparent)
    painter = QPainter(pixmap)
    painter.setRenderHint(QPainter.Antialiasing, True)
    pen = QPen(QColor(238, 242, 248), 2)
    pen.setCapStyle(Qt.RoundCap)
    pen.setJoinStyle(Qt.RoundJoin)
    painter.setPen(pen)
    painter.setBrush(Qt.NoBrush)

    if kind == "select":
        # Cursor arrow similar to the mockup.
        arrow = QPolygonF([
            QPointF(7, 5), QPointF(7, 22), QPointF(11, 18),
            QPointF(14, 24), QPointF(17, 22), QPointF(14, 16),
            QPointF(20, 16)
        ])
        painter.setBrush(QBrush(QColor(238, 242, 248)))
        painter.drawPolygon(arrow)
        painter.setBrush(Qt.NoBrush)
    elif kind == "pan":
        # Simple outlined hand/pan icon.
        painter.drawLine(QPointF(10, 15), QPointF(10, 8))
        painter.drawLine(QPointF(14, 15), QPointF(14, 6))
        painter.drawLine(QPointF(18, 16), QPointF(18, 8))
        painter.drawLine(QPointF(22, 18), QPointF(22, 11))
        painter.drawArc(7, 13, 18, 11, 180 * 16, 175 * 16)
        painter.drawLine(QPointF(7, 16), QPointF(11, 22))
        painter.drawLine(QPointF(11, 22), QPointF(21, 22))
    elif kind in ("zoom", "zoom_in", "zoom_out"):
        painter.drawEllipse(QPointF(12, 12), 6, 6)
        painter.drawLine(QPointF(17, 17), QPointF(23, 23))
        if kind != "zoom_out":
            painter.drawLine(QPointF(12, 8), QPointF(12, 16))
        painter.drawLine(QPointF(8, 12), QPointF(16, 12))
    elif kind == "move_point":
        painter.setBrush(QBrush(QColor(238, 242, 248)))
        painter.drawEllipse(QPointF(14, 14), 5, 5)
        painter.setBrush(Qt.NoBrush)
        painter.drawLine(QPointF(14, 3), QPointF(14, 8))
        painter.drawLine(QPointF(14, 20), QPointF(14, 25))
        painter.drawLine(QPointF(3, 14), QPointF(8, 14))
        painter.drawLine(QPointF(20, 14), QPointF(25, 14))
    elif kind == "move_line":
        painter.drawLine(QPointF(6, 20), QPointF(22, 8))
        painter.setBrush(QBrush(QColor(238, 242, 248)))
        painter.drawEllipse(QPointF(6, 20), 3, 3)
        painter.drawEllipse(QPointF(22, 8), 3, 3)
        painter.setBrush(Qt.NoBrush)
    elif kind == "move_endpoint":
        painter.drawLine(QPointF(7, 21), QPointF(21, 7))
        painter.setBrush(QBrush(QColor(238, 242, 248)))
        painter.drawEllipse(QPointF(21, 7), 4, 4)
        painter.setBrush(Qt.NoBrush)
        painter.drawEllipse(QPointF(7, 21), 3, 3)
    elif kind == "undo":
        # Bold standard undo arrow, with the arrow head clearly at the end.
        pen = QPen(QColor(238, 242, 248), 3.2, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin)
        painter.setPen(pen)
        path = QPainterPath(QPointF(21, 18))
        path.lineTo(QPointF(13, 18))
        path.cubicTo(QPointF(8.5, 18), QPointF(8.5, 10.5), QPointF(13, 10.5))
        path.lineTo(QPointF(17, 10.5))
        painter.drawPath(path)
        painter.setBrush(QBrush(QColor(238, 242, 248)))
        painter.setPen(Qt.NoPen)
        painter.drawPolygon(QPolygonF([
            QPointF(7.0, 10.5),
            QPointF(13.0, 6.7),
            QPointF(13.0, 14.3),
        ]))
        painter.setBrush(Qt.NoBrush)
        painter.setPen(pen)
    elif kind == "redo":
        # Bold standard redo arrow, mirrored from undo.
        pen = QPen(QColor(238, 242, 248), 3.2, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin)
        painter.setPen(pen)
        path = QPainterPath(QPointF(5, 18))
        path.lineTo(QPointF(13, 18))
        path.cubicTo(QPointF(17.5, 18), QPointF(17.5, 10.5), QPointF(13, 10.5))
        path.lineTo(QPointF(9, 10.5))
        painter.drawPath(path)
        painter.setBrush(QBrush(QColor(238, 242, 248)))
        painter.setPen(Qt.NoPen)
        painter.drawPolygon(QPolygonF([
            QPointF(21.0, 10.5),
            QPointF(15.0, 6.7),
            QPointF(15.0, 14.3),
        ]))
        painter.setBrush(Qt.NoBrush)
        painter.setPen(pen)
    elif kind == "save":
        # Simpler, chunkier floppy disk icon.
        pen = QPen(QColor(238, 242, 248), 2.6, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin)
        painter.setPen(pen)
        painter.drawRoundedRect(6.5, 5.5, 15, 17, 1.5, 1.5)
        painter.drawRect(9, 7, 7, 4)
        painter.drawLine(QPointF(17.5, 7), QPointF(17.5, 11))
        painter.drawRoundedRect(9, 15, 9, 5.5, 1.2, 1.2)
    elif kind == "revert":
        # Page with a heavy return arrow, inspired by the reference icon.
        pen = QPen(QColor(238, 242, 248), 2.8, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin)
        painter.setPen(pen)
        page = QPainterPath(QPointF(8, 4.5))
        page.lineTo(QPointF(17, 4.5))
        page.lineTo(QPointF(22, 9.5))
        page.lineTo(QPointF(22, 22))
        page.lineTo(QPointF(8, 22))
        page.closeSubpath()
        painter.drawPath(page)
        painter.drawLine(QPointF(17, 4.5), QPointF(17, 10))
        painter.drawLine(QPointF(17, 10), QPointF(22, 10))
        # bold return arrow
        painter.drawLine(QPointF(18.5, 18), QPointF(12.5, 18))
        painter.drawArc(8.3, 14.2, 8.4, 7.6, 80 * 16, 190 * 16)
        painter.setBrush(QBrush(QColor(238, 242, 248)))
        painter.setPen(Qt.NoPen)
        painter.drawPolygon(QPolygonF([
            QPointF(10.0, 16.0),
            QPointF(14.4, 12.8),
            QPointF(14.4, 19.2),
        ]))
    elif kind == "fit":
        painter.drawLine(QPointF(6, 11), QPointF(6, 6))
        painter.drawLine(QPointF(6, 6), QPointF(11, 6))
        painter.drawLine(QPointF(17, 6), QPointF(22, 6))
        painter.drawLine(QPointF(22, 6), QPointF(22, 11))
        painter.drawLine(QPointF(22, 17), QPointF(22, 22))
        painter.drawLine(QPointF(22, 22), QPointF(17, 22))
        painter.drawLine(QPointF(11, 22), QPointF(6, 22))
        painter.drawLine(QPointF(6, 22), QPointF(6, 17))

    painter.end()
    return QIcon(pixmap)


class EqMapView(QGraphicsView):
    def __init__(self, main_window: "EqMapMainWindow") -> None:
        super().__init__(main_window)
        self.main_window = main_window
        self.setRenderHint(QPainter.Antialiasing, True)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.AnchorViewCenter)
        self.setDragMode(QGraphicsView.RubberBandDrag)
        self.setRubberBandSelectionMode(Qt.IntersectsItemShape)
        self.setMouseTracking(True)
        self._is_panning = False
        self._last_pan_point = QPointF()

    def wheelEvent(self, event):
        factor = 1.2 if event.angleDelta().y() > 0 else 1 / 1.2
        self.scale(factor, factor)
        self.main_window.update_canvas_overlays()

    def mousePressEvent(self, event):
        if event.button() in (Qt.MiddleButton, Qt.RightButton):
            self._is_panning = True
            self._last_pan_point = event.position()
            self.setCursor(Qt.ClosedHandCursor)
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseDoubleClickEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.main_window.show_map_context_menu(event.globalPosition().toPoint(), self.mapToScene(event.position().toPoint()))
            event.accept()
            return
        super().mouseDoubleClickEvent(event)

    def mouseMoveEvent(self, event):
        self.main_window.update_canvas_cursor_position(self.mapToScene(event.position().toPoint()))
        if self._is_panning:
            delta = event.position() - self._last_pan_point
            self._last_pan_point = event.position()
            self.horizontalScrollBar().setValue(self.horizontalScrollBar().value() - int(delta.x()))
            self.verticalScrollBar().setValue(self.verticalScrollBar().value() - int(delta.y()))
            self.main_window.update_canvas_overlays()
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if event.button() in (Qt.MiddleButton, Qt.RightButton) and self._is_panning:
            self._is_panning = False
            self.setCursor(Qt.ArrowCursor)
            self.main_window.update_canvas_overlays()
            event.accept()
            return
        super().mouseReleaseEvent(event)


class MapCanvasContainer(QWidget):
    def __init__(self, main_window: "EqMapMainWindow") -> None:
        super().__init__(main_window)
        self.main_window = main_window

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        self.main_window.update_scene_pan_padding()
        self.main_window.update_canvas_overlays()


class CanvasControlsOverlay(QWidget):
    def __init__(self, main_window: "EqMapMainWindow") -> None:
        super().__init__(main_window)
        self.main_window = main_window
        self.setObjectName("canvasControlsOverlay")
        self.setAttribute(Qt.WA_StyledBackground, True)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(12, 8, 12, 8)
        layout.setSpacing(8)

        self.mode_buttons: dict[str, QToolButton] = {}

        mode_specs = [
            ("Select Only", "select", "Select / inspect records"),
            ("Move Points", "move_point", "Move point records"),
            ("Move Lines", "move_line", "Move whole line records"),
            ("Move Line Endpoints", "move_endpoint", "Move selected line endpoints"),
        ]
        for mode_name, icon_name, tooltip in mode_specs:
            button = QToolButton()
            button.setObjectName("canvasControlButton")
            button.setIcon(make_canvas_icon(icon_name))
            button.setToolTip(tooltip)
            button.setCheckable(True)
            button.setIconSize(QSize(26, 26))
            button.setFixedSize(42, 38)
            button.clicked.connect(lambda checked=False, mode=mode_name: self.main_window.set_edit_mode_from_overlay(mode))
            layout.addWidget(button)
            self.mode_buttons[mode_name] = button

        layout.addSpacing(8)

        self.zoom_out_button = QToolButton()
        self.zoom_out_button.setObjectName("canvasControlButton")
        self.zoom_out_button.setIcon(make_canvas_icon("zoom_out"))
        self.zoom_out_button.setToolTip("Zoom out")
        self.zoom_in_button = QToolButton()
        self.zoom_in_button.setObjectName("canvasControlButton")
        self.zoom_in_button.setIcon(make_canvas_icon("zoom_in"))
        self.zoom_in_button.setToolTip("Zoom in")
        self.fit_button = QToolButton()
        self.fit_button.setObjectName("canvasControlButton")
        self.fit_button.setIcon(make_canvas_icon("fit"))
        self.fit_button.setToolTip("Fit map")

        for button in [self.zoom_out_button, self.zoom_in_button, self.fit_button]:
            button.setIconSize(QSize(26, 26))
            button.setFixedSize(42, 38)
            layout.addWidget(button)

        layout.addSpacing(8)

        self.undo_button = QToolButton()
        self.undo_button.setObjectName("canvasControlButton")
        self.undo_button.setIcon(resource_icon("icons/undo.svg", "undo"))
        self.undo_button.setToolTip("Undo")
        self.redo_button = QToolButton()
        self.redo_button.setObjectName("canvasControlButton")
        self.redo_button.setIcon(resource_icon("icons/redo.svg", "redo"))
        self.redo_button.setToolTip("Redo")
        self.save_button = QToolButton()
        self.save_button.setObjectName("canvasControlButton")
        self.save_button.setIcon(resource_icon("icons/save.svg", "save"))
        self.save_button.setToolTip("Save edits")
        self.revert_button = QToolButton()
        self.revert_button.setObjectName("canvasControlButton")
        self.revert_button.setIcon(resource_icon("icons/revert.svg", "revert"))
        self.revert_button.setToolTip("Revert unsaved edits from disk")

        for button in [self.undo_button, self.redo_button, self.save_button, self.revert_button]:
            button.setIconSize(QSize(26, 26))
            button.setFixedSize(42, 38)
            layout.addWidget(button)

        self.coord_label = QLabel("X: 0.00   Y: 0.00   Z: 0.00")
        self.coord_label.setObjectName("canvasOverlayText")
        layout.addSpacing(18)
        layout.addWidget(self.coord_label)
        layout.addStretch(1)

        self.zoom_label = QLabel("Zoom: 100%")
        self.zoom_label.setObjectName("canvasOverlayText")
        layout.addWidget(self.zoom_label)

        self.zoom_out_button.clicked.connect(self.main_window.zoom_out)
        self.zoom_in_button.clicked.connect(self.main_window.zoom_in)
        self.fit_button.clicked.connect(self.main_window.fit_map)
        self.undo_button.clicked.connect(self.main_window.undo)
        self.redo_button.clicked.connect(self.main_window.redo)
        self.save_button.clicked.connect(self.main_window.save_edits)
        self.revert_button.clicked.connect(self.main_window.reload_from_disk)
        self.set_active_mode("Select Only")

    def update_status(self, text: str, zoom_text: str) -> None:
        self.coord_label.setText(text)
        self.zoom_label.setText(zoom_text)

    def set_active_mode(self, mode_name: str) -> None:
        for name, button in self.mode_buttons.items():
            button.setChecked(name == mode_name)


class MiniMapOverlay(QWidget):
    def __init__(self, main_window: "EqMapMainWindow") -> None:
        super().__init__(main_window)
        self.main_window = main_window
        self.setObjectName("miniMapOverlay")
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.setMinimumSize(150, 120)

    def paintEvent(self, event) -> None:
        super().paintEvent(event)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, True)

        outer = self.rect().adjusted(8, 8, -8, -8)
        painter.setPen(QPen(QColor(170, 170, 170), 1))
        painter.setBrush(QBrush(QColor(20, 20, 20, 180)))
        painter.drawRect(outer)

        scene_rect = self.main_window.scene.itemsBoundingRect()
        if scene_rect.isNull() or scene_rect.width() <= 0 or scene_rect.height() <= 0:
            painter.setPen(QColor(150, 150, 150))
            painter.drawText(outer, Qt.AlignCenter, "No map")
            painter.end()
            return

        inner = outer.adjusted(10, 10, -10, -10)
        scale = min(inner.width() / scene_rect.width(), inner.height() / scene_rect.height())
        draw_w = scene_rect.width() * scale
        draw_h = scene_rect.height() * scale
        off_x = inner.x() + (inner.width() - draw_w) / 2
        off_y = inner.y() + (inner.height() - draw_h) / 2

        def map_point(point: QPointF) -> QPointF:
            return QPointF(
                off_x + (point.x() - scene_rect.x()) * scale,
                off_y + (point.y() - scene_rect.y()) * scale,
            )

        painter.setPen(QPen(QColor(135, 135, 135), 1))
        for record in self.main_window.loaded_map.lines:
            if getattr(record, "deleted", False) or not self.main_window.layer_visible.get(record.file_path, True):
                continue
            p1 = self.main_window.mapper.map_to_scene(record.x1, record.y1)
            p2 = self.main_window.mapper.map_to_scene(record.x2, record.y2)
            painter.drawLine(map_point(p1), map_point(p2))

        painter.setPen(QPen(QColor(235, 235, 235), 2))
        viewport_poly = self.main_window.view.mapToScene(self.main_window.view.viewport().rect())
        viewport_rect = viewport_poly.boundingRect()
        vr = QRectF(
            off_x + (viewport_rect.x() - scene_rect.x()) * scale,
            off_y + (viewport_rect.y() - scene_rect.y()) * scale,
            viewport_rect.width() * scale,
            viewport_rect.height() * scale,
        )
        painter.drawRect(vr)
        painter.end()


class MovablePointMarker(QGraphicsEllipseItem):
    def __init__(self, main_window: "EqMapMainWindow", record: MapPointRecord):
        super().__init__()
        self.main_window = main_window
        self.record = record
        self._drag_before: Optional[dict[str, Any]] = None
        self.setData(0, record)
        self.setFlag(QGraphicsItem.ItemIsSelectable, True)
        self.setFlag(QGraphicsItem.ItemSendsGeometryChanges, True)

    def mousePressEvent(self, event):
        if self.main_window.edit_mode() == "Move Points":
            self._drag_before = snapshot_point(self.record)
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if self._drag_before is not None:
            after = snapshot_point(self.record)
            if self._drag_before != after:
                self.main_window.add_undo("Move point", self.record, self._drag_before, after)
            self._drag_before = None
        super().mouseReleaseEvent(event)

    def itemChange(self, change, value):
        if change == QGraphicsItem.ItemPositionHasChanged and self.scene() is not None:
            self.main_window.on_marker_dragged(self.record, self.pos())
        return super().itemChange(change, value)


class MovableLineItem(QGraphicsLineItem):
    def __init__(self, main_window: "EqMapMainWindow", record: MapLineRecord):
        super().__init__()
        self.main_window = main_window
        self.record = record
        self._drag_before: Optional[dict[str, Any]] = None
        self.setData(0, record)
        self.setFlag(QGraphicsItem.ItemIsSelectable, True)
        self.setFlag(QGraphicsItem.ItemSendsGeometryChanges, True)

    def mousePressEvent(self, event):
        if self.main_window.edit_mode() == "Move Lines":
            self._drag_before = snapshot_line(self.record)
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if self._drag_before is not None:
            after = snapshot_line(self.record)
            if self._drag_before != after:
                self.main_window.add_undo("Move line", self.record, self._drag_before, after)
            self._drag_before = None
        super().mouseReleaseEvent(event)

    def itemChange(self, change, value):
        if change == QGraphicsItem.ItemPositionHasChanged and self.scene() is not None:
            self.main_window.on_line_dragged(self.record, self.pos())
        return super().itemChange(change, value)


class EndpointHandle(QGraphicsEllipseItem):
    def __init__(self, main_window: "EqMapMainWindow", record: MapLineRecord, endpoint: int):
        super().__init__(-5, -5, 10, 10)
        self.main_window = main_window
        self.record = record
        self.endpoint = endpoint
        self._drag_before: Optional[dict[str, Any]] = None
        self.setBrush(QBrush(QColor(255, 255, 0)))
        self.setPen(QPen(QColor(0, 0, 0), 1))
        self.setZValue(100000)
        self.setFlag(QGraphicsItem.ItemIsMovable, True)
        self.setFlag(QGraphicsItem.ItemSendsGeometryChanges, True)

    def mousePressEvent(self, event):
        self._drag_before = snapshot_line(self.record)
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if self._drag_before is not None:
            after = snapshot_line(self.record)
            if self._drag_before != after:
                self.main_window.add_undo("Move line endpoint", self.record, self._drag_before, after)
            self._drag_before = None
        super().mouseReleaseEvent(event)

    def itemChange(self, change, value):
        if change == QGraphicsItem.ItemPositionHasChanged and self.scene() is not None:
            self.main_window.on_endpoint_dragged(self.record, self.endpoint, self.pos())
        return super().itemChange(change, value)




def ensure_spinbox_arrow_images() -> tuple[str, str]:
    """Create tiny light-coloured arrow PNGs for dark-mode spin boxes.

    Qt stylesheet triangle borders can render as squares on some Windows/PyInstaller builds.
    Real image arrows are much more reliable.
    """
    up_path = RESOURCES_DIR / "spinbox_arrow_up_light.png"
    down_path = RESOURCES_DIR / "spinbox_arrow_down_light.png"

    if not up_path.exists() or not down_path.exists():
        up = QPixmap(9, 9)
        up.fill(Qt.transparent)
        painter = QPainter(up)
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setBrush(QBrush(QColor(245, 245, 245)))
        painter.setPen(Qt.NoPen)
        painter.drawPolygon([
            QPointF(4.5, 1.5),
            QPointF(8.0, 6.5),
            QPointF(1.0, 6.5),
        ])
        painter.end()
        up.save(str(up_path), "PNG")

        down = QPixmap(9, 9)
        down.fill(Qt.transparent)
        painter = QPainter(down)
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setBrush(QBrush(QColor(245, 245, 245)))
        painter.setPen(Qt.NoPen)
        painter.drawPolygon([
            QPointF(1.0, 2.5),
            QPointF(8.0, 2.5),
            QPointF(4.5, 7.5),
        ])
        painter.end()
        down.save(str(down_path), "PNG")

    # Qt stylesheets prefer forward slashes, even on Windows.
    return up_path.as_posix(), down_path.as_posix()


def choose_colour_dialog(initial: QColor, parent=None, title: str = "Choose colour") -> QColor:
    dialog = QColorDialog(initial, parent)
    dialog.setWindowTitle(title)
    dialog.setOption(QColorDialog.DontUseNativeDialog, True)
    dialog.setStyle(QApplication.style())

    up_arrow, down_arrow = ensure_spinbox_arrow_images()

    # Dark-mode colour dialog with explicit image arrows.
    # This keeps the dialog visually consistent with the app while avoiding the missing-arrow
    # behavior that can happen when Qt has to draw stylesheet-modified spinbox arrows itself.
    dialog.setStyleSheet(f"""
        QColorDialog, QWidget {{
            background-color: #242424;
            color: #f0f0f0;
        }}
        QLabel {{
            color: #f0f0f0;
        }}
        QLineEdit, QSpinBox, QDoubleSpinBox {{
            background-color: #1b1b1b;
            color: #f0f0f0;
            border: 1px solid #666;
            min-height: 22px;
            padding-right: 22px;
        }}
        QSpinBox::up-button, QDoubleSpinBox::up-button {{
            subcontrol-origin: border;
            subcontrol-position: top right;
            width: 20px;
            border-left: 1px solid #777;
            border-bottom: 1px solid #555;
            background-color: #3a3a3a;
        }}
        QSpinBox::down-button, QDoubleSpinBox::down-button {{
            subcontrol-origin: border;
            subcontrol-position: bottom right;
            width: 20px;
            border-left: 1px solid #777;
            border-top: 1px solid #555;
            background-color: #3a3a3a;
        }}
        QSpinBox::up-button:hover, QDoubleSpinBox::up-button:hover,
        QSpinBox::down-button:hover, QDoubleSpinBox::down-button:hover {{
            background-color: #505050;
        }}
        QSpinBox::up-arrow, QDoubleSpinBox::up-arrow {{
            image: url("{up_arrow}");
            width: 9px;
            height: 9px;
        }}
        QSpinBox::down-arrow, QDoubleSpinBox::down-arrow {{
            image: url("{down_arrow}");
            width: 9px;
            height: 9px;
        }}
        QPushButton {{
            background-color: #333;
            color: #f0f0f0;
            border: 1px solid #777;
            padding: 4px 8px;
        }}
        QPushButton:hover {{
            background-color: #444;
        }}
    """)
    if dialog.exec() == QDialog.Accepted:
        return dialog.selectedColor()
    return QColor()


class ColorChoiceDialog(QDialog):
    def __init__(self, colours: list[tuple[tuple[int, int, int], int, int]], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Match Colour from Map")
        self.selected_rgb: Optional[tuple[int, int, int]] = None

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Choose a colour currently used on the map:"))

        self.list_widget = QListWidget()
        layout.addWidget(self.list_widget)

        for rgb, point_count, line_count in colours:
            total = point_count + line_count
            item = QListWidgetItem(f"■ RGB {rgb}    Points: {point_count}    Lines: {line_count}    Total: {total}")
            item.setData(Qt.UserRole, rgb)
            item.setForeground(QColor(*rgb))
            self.list_widget.addItem(item)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def accept(self) -> None:
        item = self.list_widget.currentItem()
        if item is not None:
            self.selected_rgb = item.data(Qt.UserRole)
        super().accept()



class PaletteEditorDialog(QDialog):
    def __init__(self, parent=None, palette: Optional[dict[str, Any]] = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Save Colour Palette")
        self.resize(720, 520)
        self.saved_palette: Optional[dict[str, Any]] = None

        palette = palette or DEFAULT_MAP_PALETTES["EQ Map Standard"]

        layout = QVBoxLayout(self)
        form = QFormLayout()
        layout.addLayout(form)

        self.name_edit = QLineEdit(str(palette.get("name", "Custom Palette")))
        form.addRow("Palette name", self.name_edit)

        self.description_edit = QLineEdit(str(palette.get("description", "")))
        form.addRow("Description", self.description_edit)

        layout.addWidget(QLabel("Each entry has a light-mode colour and a dark-mode colour. Palette switching maps each record to the nearest entry, then applies the chosen light/dark version."))

        self.entries_list = QListWidget()
        layout.addWidget(self.entries_list, 1)

        button_row = QHBoxLayout()
        self.add_button = QPushButton("Add Entry")
        self.remove_button = QPushButton("Remove Selected")
        self.light_button = QPushButton("Pick Light Colour")
        self.dark_button = QPushButton("Pick Dark Colour")
        button_row.addWidget(self.add_button)
        button_row.addWidget(self.remove_button)
        button_row.addWidget(self.light_button)
        button_row.addWidget(self.dark_button)
        layout.addLayout(button_row)

        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.add_button.clicked.connect(self.add_entry)
        self.remove_button.clicked.connect(self.remove_selected)
        self.light_button.clicked.connect(lambda: self.pick_entry_colour("light"))
        self.dark_button.clicked.connect(lambda: self.pick_entry_colour("dark"))

        for entry in palette_entries(palette):
            self.add_entry(entry)

    def add_entry(self, entry: Optional[dict[str, Any]] = None) -> None:
        if entry is None or isinstance(entry, bool):
            entry = {"name": "Custom Colour", "light": (0, 0, 0), "dark": (255, 255, 255)}
        item = QListWidgetItem()
        item.setData(Qt.UserRole, {
            "name": entry["name"],
            "light": tuple(entry["light"]),
            "dark": tuple(entry["dark"]),
        })
        self.entries_list.addItem(item)
        self.refresh_item(item)
        self.entries_list.setCurrentItem(item)

    def refresh_item(self, item: QListWidgetItem) -> None:
        data = item.data(Qt.UserRole)
        light = tuple(data["light"])
        dark = tuple(data["dark"])
        item.setText(f'{data["name"]}    Light RGB {light}  →  Dark RGB {dark}')
        icon = QPixmap(48, 16)
        icon.fill(Qt.transparent)
        painter = QPainter(icon)
        painter.fillRect(0, 0, 24, 16, QColor(*light))
        painter.fillRect(24, 0, 24, 16, QColor(*dark))
        painter.setPen(QPen(QColor(120, 120, 120), 1))
        painter.drawRect(0, 0, 47, 15)
        painter.end()
        item.setIcon(QIcon(icon))

    def remove_selected(self) -> None:
        for item in self.entries_list.selectedItems():
            row = self.entries_list.row(item)
            self.entries_list.takeItem(row)

    def pick_entry_colour(self, key: str) -> None:
        item = self.entries_list.currentItem()
        if item is None:
            return
        data = item.data(Qt.UserRole)
        current = QColor(*data[key])
        colour = choose_colour_dialog(current, self, f"Choose {key} colour")
        if not colour.isValid():
            return
        data[key] = (colour.red(), colour.green(), colour.blue())
        item.setData(Qt.UserRole, data)
        self.refresh_item(item)

    def accept(self) -> None:
        name = self.name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "Palette Name Required", "Please enter a palette name.")
            return
        entries = []
        for row in range(self.entries_list.count()):
            data = self.entries_list.item(row).data(Qt.UserRole)
            entries.append({
                "name": data["name"],
                "light": list(data["light"]),
                "dark": list(data["dark"]),
            })
        if not entries:
            QMessageBox.warning(self, "Palette Empty", "Please add at least one palette entry.")
            return
        self.saved_palette = {
            "name": name,
            "description": self.description_edit.text().strip(),
            "entries": entries,
        }
        super().accept()


class PreferencesDialog(QDialog):
    def __init__(self, main_window: "EqMapMainWindow") -> None:
        super().__init__(main_window)
        self.main_window = main_window
        self.setWindowTitle("Preferences")
        layout = QVBoxLayout(self)
        form = QFormLayout()
        layout.addLayout(form)

        self.remember_last_checkbox = QCheckBox("Remember last loaded map")
        self.remember_last_checkbox.setChecked(main_window.remember_last_loaded_map)
        form.addRow(self.remember_last_checkbox)

        self.auto_fit_checkbox = QCheckBox("Auto-fit restored map on startup")
        self.auto_fit_checkbox.setChecked(main_window.auto_fit_restored_map)
        form.addRow(self.auto_fit_checkbox)

        self.autosave_settings_checkbox = QCheckBox("Autosave settings on exit")
        self.autosave_settings_checkbox.setChecked(main_window.autosave_settings_on_exit)
        form.addRow(self.autosave_settings_checkbox)

        self.show_warning_checkbox = QCheckBox("Show beta safety warning on startup")
        self.show_warning_checkbox.setChecked(main_window.show_beta_warning)
        form.addRow(self.show_warning_checkbox)

        self.flip_y_checkbox = QCheckBox("Flip display Y")
        self.flip_y_checkbox.setChecked(main_window.mapper.flip_display_y)
        form.addRow(self.flip_y_checkbox)

        self.default_background_combo = QComboBox()
        self.default_background_combo.addItems(["Remember Last", "Light", "Dark"])
        bg_value = main_window.default_background_mode
        if bg_value == "light":
            self.default_background_combo.setCurrentText("Light")
        elif bg_value == "dark":
            self.default_background_combo.setCurrentText("Dark")
        else:
            self.default_background_combo.setCurrentText("Remember Last")
        form.addRow("Default background", self.default_background_combo)

        folder_row = QHBoxLayout()
        self.map_folder_edit = QLineEdit(main_window.map_folder)
        self.map_folder_button = QPushButton("Choose...")
        self.map_folder_button.clicked.connect(self.choose_folder)
        folder_row.addWidget(self.map_folder_edit)
        folder_row.addWidget(self.map_folder_button)
        form.addRow("Default map folder", folder_row)

        self.max_highlights_spin = QSpinBox()
        self.max_highlights_spin.setRange(0, 1000000)
        self.max_highlights_spin.setValue(main_window.max_highlights_before_warning)
        self.max_highlights_spin.setToolTip("0 = never warn before highlighting")
        form.addRow("Warn before highlighting more than", self.max_highlights_spin)

        self.confirm_edit_over_spin = QSpinBox()
        self.confirm_edit_over_spin.setRange(0, 1000000)
        self.confirm_edit_over_spin.setValue(main_window.confirm_bulk_edit_over)
        form.addRow("Confirm bulk edits over", self.confirm_edit_over_spin)

        self.confirm_delete_over_spin = QSpinBox()
        self.confirm_delete_over_spin.setRange(0, 1000000)
        self.confirm_delete_over_spin.setValue(main_window.confirm_bulk_delete_over)
        form.addRow("Confirm deletes over", self.confirm_delete_over_spin)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def choose_folder(self) -> None:
        folder = QFileDialog.getExistingDirectory(self, "Choose Default Map Folder", self.map_folder_edit.text())
        if folder:
            self.map_folder_edit.setText(folder)

    def accept(self) -> None:
        bg_text = self.default_background_combo.currentText()
        bg_value = "remember"
        if bg_text == "Light":
            bg_value = "light"
        elif bg_text == "Dark":
            bg_value = "dark"

        self.main_window.remember_last_loaded_map = self.remember_last_checkbox.isChecked()
        self.main_window.auto_fit_restored_map = self.auto_fit_checkbox.isChecked()
        self.main_window.autosave_settings_on_exit = self.autosave_settings_checkbox.isChecked()
        self.main_window.show_beta_warning = self.show_warning_checkbox.isChecked()
        self.main_window.default_background_mode = bg_value
        self.main_window.map_folder = self.map_folder_edit.text().strip()
        self.main_window.max_highlights_before_warning = self.max_highlights_spin.value()
        self.main_window.confirm_bulk_edit_over = self.confirm_edit_over_spin.value()
        self.main_window.confirm_bulk_delete_over = self.confirm_delete_over_spin.value()

        self.main_window.mapper.flip_display_y = self.flip_y_checkbox.isChecked()
        if hasattr(self.main_window, "flip_y_action"):
            self.main_window.flip_y_action.setChecked(self.main_window.mapper.flip_display_y)
        self.main_window.render_map(keep_view=True)

        if self.main_window.default_background_mode == "light":
            self.main_window.set_background("light")
        elif self.main_window.default_background_mode == "dark":
            self.main_window.set_background("dark")

        if self.main_window.map_folder and hasattr(self.main_window, "side_panel"):
            self.main_window.side_panel.map_folder_edit.setText(self.main_window.map_folder)
            self.main_window.rebuild_zone_list()

        self.main_window.save_settings()
        super().accept()


class SidePanel(QWidget):
    def __init__(self, main_window: "EqMapMainWindow") -> None:
        super().__init__(main_window)
        self.main_window = main_window
        self.current_record: Optional[Any] = None
        self._loading = False
        self.layer_checkboxes: dict[Path, QCheckBox] = {}
        self._build_ui()
        self.set_record(None)


    def _make_collapsible_group(self, title: str, expanded: bool = True) -> tuple[QGroupBox, QVBoxLayout]:
        """Create a checkable workflow box whose contents collapse to a one-line header."""
        group = QGroupBox(title)
        group.setCheckable(True)
        group.setChecked(expanded)
        group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        outer_layout = QVBoxLayout(group)
        outer_layout.setContentsMargins(8, 8, 8, 8)
        content = QWidget(group)
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(0, 2, 0, 0)
        outer_layout.addWidget(content)

        def apply_collapsed_state(checked: bool) -> None:
            content.setVisible(checked)
            if checked:
                group.setMinimumHeight(0)
                group.setMaximumHeight(16777215)
            else:
                # A checked QGroupBox keeps its title/checkbox visible at roughly this height.
                group.setMinimumHeight(34)
                group.setMaximumHeight(38)
            if hasattr(self, "rebalance_npc_workflow_boxes"):
                QTimer.singleShot(0, self.rebalance_npc_workflow_boxes)

        apply_collapsed_state(expanded)
        group.toggled.connect(apply_collapsed_state)
        group._collapsible_content = content  # keep a reference for PySide
        return group, content_layout

    def rebalance_npc_workflow_boxes(self) -> None:
        """Give collapsed NPC workflow boxes one line and split remaining height across open boxes."""
        if not hasattr(self, "npc_workflow_splitter") or not hasattr(self, "npc_workflow_groups"):
            return
        groups = list(self.npc_workflow_groups)
        if not groups:
            return
        splitter_height = max(self.npc_workflow_splitter.height(), 420)
        collapsed_size = 38
        open_groups = [group for group in groups if group.isChecked()]
        collapsed_groups = [group for group in groups if not group.isChecked()]
        remaining = max(120, splitter_height - collapsed_size * len(collapsed_groups))
        open_size = max(180, remaining // max(1, len(open_groups)))
        sizes = [open_size if group.isChecked() else collapsed_size for group in groups]
        self.npc_workflow_splitter.setSizes(sizes)

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        heading = QLabel("Inspector")
        heading.setStyleSheet("font-size: 15px; font-weight: bold;")
        layout.addWidget(heading)

        controls_group = QGroupBox("Map Controls")
        controls_layout = QVBoxLayout(controls_group)

        controls_layout.addWidget(QLabel("Edit Mode"))
        self.edit_mode_combo = QComboBox()
        self.edit_mode_combo.addItems([
            "Select Only",
            "Move Points",
            "Move Lines",
            "Move Line Endpoints",
        ])
        controls_layout.addWidget(self.edit_mode_combo)

        controls_layout.addWidget(QLabel("Active File for New Records"))
        self.active_file_combo = QComboBox()
        controls_layout.addWidget(self.active_file_combo)

        undo_row = QHBoxLayout()
        self.undo_button = QPushButton("Undo")
        self.redo_button = QPushButton("Redo")
        undo_row.addWidget(self.undo_button)
        undo_row.addWidget(self.redo_button)
        controls_layout.addLayout(undo_row)

        restore_row = QHBoxLayout()
        self.reload_button = QPushButton("Reload from Disk")
        self.restore_button = QPushButton("Restore Backup")
        restore_row.addWidget(self.reload_button)
        restore_row.addWidget(self.restore_button)
        controls_layout.addLayout(restore_row)

        layout.addWidget(controls_group)

        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)


        selection_tab = QWidget()
        selection_layout = QVBoxLayout(selection_tab)
        selection_layout.setContentsMargins(8, 8, 8, 8)
        selection_layout.setSpacing(10)

        inspector_header_row = QHBoxLayout()
        self.inspector_header_label = QLabel("Inspector")
        self.inspector_header_label.setObjectName("inspectorHeaderLabel")
        inspector_header_row.addWidget(self.inspector_header_label)
        inspector_header_row.addStretch(1)
        self.inspector_close_button = QToolButton()
        self.inspector_close_button.setObjectName("inspectorUtilityButton")
        self.inspector_close_button.setText("✕")
        self.inspector_close_button.setToolTip("Hide inspector")
        inspector_header_row.addWidget(self.inspector_close_button)
        selection_layout.addLayout(inspector_header_row)

        summary_group = QGroupBox("Selection Summary")
        summary_layout = QVBoxLayout(summary_group)
        summary_layout.setSpacing(6)
        self.title_label = QLabel("No selection")
        self.title_label.setObjectName("inspectorSummaryTitle")
        summary_layout.addWidget(self.title_label)

        self.source_label = QLabel("")
        self.source_label.setWordWrap(True)
        self.source_label.setObjectName("inspectorSummaryMeta")
        summary_layout.addWidget(self.source_label)

        summary_action_row = QHBoxLayout()
        self.multi_select_label = QLabel("Multi-select: none")
        self.multi_select_label.setWordWrap(True)
        self.multi_select_label.setObjectName("inspectorMultiSummary")
        summary_action_row.addWidget(self.multi_select_label, 1)
        self.delete_selected_button = QPushButton("Delete Selected")
        self.delete_selected_button.setObjectName("dangerButton")
        self.delete_selected_button.setEnabled(False)
        summary_action_row.addWidget(self.delete_selected_button)
        summary_layout.addLayout(summary_action_row)
        selection_layout.addWidget(summary_group)

        label_group = QGroupBox("Label")
        label_layout = QVBoxLayout(label_group)
        self.label_edit = QLineEdit()
        self.label_edit.setPlaceholderText("Point label")
        label_layout.addWidget(self.label_edit)
        selection_layout.addWidget(label_group)

        point_group = QGroupBox("Point Coordinates")
        point_layout = QHBoxLayout(point_group)
        point_layout.setSpacing(8)
        self.x_spin = self.coord_spin()
        self.y_spin = self.coord_spin()
        self.z_spin = self.coord_spin()
        for _label, _spin in (("X", self.x_spin), ("Y", self.y_spin), ("Z", self.z_spin)):
            col = QVBoxLayout()
            lbl = QLabel(_label)
            lbl.setObjectName("compactFieldLabel")
            col.addWidget(lbl)
            col.addWidget(_spin)
            point_layout.addLayout(col)
        selection_layout.addWidget(point_group)
        self.point_coords_label = point_group

        line_row = QHBoxLayout()
        line_row.setSpacing(8)

        endpoint1_group = QGroupBox("Line Endpoint 1")
        endpoint1_layout = QVBoxLayout(endpoint1_group)
        self.x1_spin = self.coord_spin()
        self.y1_spin = self.coord_spin()
        self.z1_spin = self.coord_spin()
        for _label, _spin in (("X1", self.x1_spin), ("Y1", self.y1_spin), ("Z1", self.z1_spin)):
            col = QVBoxLayout()
            lbl = QLabel(_label)
            lbl.setObjectName("compactFieldLabel")
            col.addWidget(lbl)
            col.addWidget(_spin)
            endpoint1_layout.addLayout(col)
        line_row.addWidget(endpoint1_group)
        self.endpoint1_label = endpoint1_group

        endpoint2_group = QGroupBox("Line Endpoint 2")
        endpoint2_layout = QVBoxLayout(endpoint2_group)
        self.x2_spin = self.coord_spin()
        self.y2_spin = self.coord_spin()
        self.z2_spin = self.coord_spin()
        for _label, _spin in (("X2", self.x2_spin), ("Y2", self.y2_spin), ("Z2", self.z2_spin)):
            col = QVBoxLayout()
            lbl = QLabel(_label)
            lbl.setObjectName("compactFieldLabel")
            col.addWidget(lbl)
            col.addWidget(_spin)
            endpoint2_layout.addLayout(col)
        line_row.addWidget(endpoint2_group)
        self.endpoint2_label = endpoint2_group

        selection_layout.addLayout(line_row)

        colour_group = QGroupBox("Color (RGB)")
        colour_layout = QHBoxLayout(colour_group)
        colour_form_wrap = QVBoxLayout()
        self.color_label = colour_group
        self.r_spin = self.rgb_spin()
        self.g_spin = self.rgb_spin()
        self.b_spin = self.rgb_spin()
        for _label, _spin in (("R", self.r_spin), ("G", self.g_spin), ("B", self.b_spin)):
            row = QHBoxLayout()
            lbl = QLabel(_label)
            lbl.setObjectName("compactFieldLabel")
            lbl.setMinimumWidth(12)
            row.addWidget(lbl)
            row.addWidget(_spin, 1)
            colour_form_wrap.addLayout(row)
        colour_layout.addLayout(colour_form_wrap, 1)

        colour_side = QVBoxLayout()
        self.color_preview = QLabel("")
        self.color_preview.setObjectName("colourPreviewSwatch")
        self.color_preview.setMinimumSize(54, 54)
        self.color_preview.setMaximumWidth(64)
        colour_side.addWidget(self.color_preview, 0, Qt.AlignLeft)
        self.pick_color_btn = QPushButton("Pick Colour")
        colour_side.addWidget(self.pick_color_btn)
        colour_side.addStretch(1)
        colour_layout.addLayout(colour_side)
        selection_layout.addWidget(colour_group)

        point_size_group = QGroupBox("Point Size")
        point_size_layout = QVBoxLayout(point_size_group)
        self.size_spin = QDoubleSpinBox()
        self.size_spin.setRange(0.1, 99.0)
        self.size_spin.setSingleStep(0.5)
        self.size_spin.setDecimals(2)
        point_size_layout.addWidget(self.size_spin)
        selection_layout.addWidget(point_size_group)

        self.apply_button = QPushButton("Apply Changes")
        self.apply_button.setObjectName("primaryButton")
        selection_layout.addWidget(self.apply_button)

        self.raw_label = QLabel("")
        self.raw_label.setWordWrap(True)
        self.raw_label.setObjectName("inspectorHelperText")
        selection_layout.addWidget(self.raw_label)

        self.footer_summary_label = QLabel("Multi-select: 0 item(s)")
        self.footer_summary_label.setObjectName("inspectorFooterSummary")
        self.footer_summary_label.setWordWrap(True)
        selection_layout.addWidget(self.footer_summary_label)

        selection_layout.addStretch(1)
        self.tabs.addTab(selection_tab, "Inspector")

        layers_tab = QWidget()
        self.layers_layout = QVBoxLayout(layers_tab)
        self.layers_hint = QLabel("Open files to show layer toggles.")
        self.layers_hint.setWordWrap(True)
        self.layers_layout.addWidget(self.layers_hint)
        self.tabs.addTab(layers_tab, "Layers")

        bulk_tab = QWidget()
        bulk_layout = QVBoxLayout(bulk_tab)
        bulk_layout.addWidget(QLabel("Colours used by visible map records"))

        self.point_colour_list = QListWidget()
        self.point_colour_list.setSelectionMode(QListWidget.ExtendedSelection)
        self.line_colour_list = QListWidget()
        self.line_colour_list.setSelectionMode(QListWidget.ExtendedSelection)

        bulk_layout.addWidget(QLabel("Point Colours"))
        bulk_layout.addWidget(self.point_colour_list)
        bulk_layout.addWidget(QLabel("Line Colours"))
        bulk_layout.addWidget(self.line_colour_list)

        recolour_group = QGroupBox("Change Selected Matching Colour")
        recolour_layout = QFormLayout(recolour_group)
        self.bulk_old_rgb_label = QLabel("No colour selected")
        recolour_layout.addRow("Selected colour", self.bulk_old_rgb_label)
        self.bulk_new_r = self.rgb_spin()
        self.bulk_new_g = self.rgb_spin()
        self.bulk_new_b = self.rgb_spin()
        recolour_layout.addRow("New R", self.bulk_new_r)
        recolour_layout.addRow("New G", self.bulk_new_g)
        recolour_layout.addRow("New B", self.bulk_new_b)
        self.bulk_pick_colour_button = QPushButton("Pick New Colour")
        self.bulk_match_colour_button = QPushButton("Match Colour from Other Point/Line")
        self.bulk_apply_colour_button = QPushButton("Apply to Selected Matching Records")
        recolour_layout.addRow(self.bulk_pick_colour_button)
        recolour_layout.addRow(self.bulk_match_colour_button)
        recolour_layout.addRow(self.bulk_apply_colour_button)
        bulk_layout.addWidget(recolour_group)

        palette_group = QGroupBox("Palette Conversion")
        palette_layout = QVBoxLayout(palette_group)

        palette_form = QFormLayout()
        self.palette_combo = QComboBox()
        self.palette_target_combo = QComboBox()
        self.palette_target_combo.addItems(["Dark", "Light"])
        self.palette_preview_label = QLabel("Choose a palette to convert map colours between light and dark versions.")
        self.palette_preview_label.setWordWrap(True)
        self.palette_preview_label.setObjectName("inspectorHelperText")
        palette_form.addRow("Palette", self.palette_combo)
        palette_form.addRow("Target", self.palette_target_combo)
        palette_form.addRow(self.palette_preview_label)
        palette_layout.addLayout(palette_form)

        palette_layout.addWidget(QLabel("Mapping Preview"))
        self.palette_mapping_list = QListWidget()
        self.palette_mapping_list.setMinimumHeight(190)
        self.palette_mapping_widgets: dict[int, QComboBox] = {}
        palette_layout.addWidget(self.palette_mapping_list)

        palette_buttons = QHBoxLayout()
        self.palette_rebuild_mapping_button = QPushButton("Build Mapping Preview")
        self.palette_auto_mapping_button = QPushButton("Auto-map")
        self.palette_apply_mapping_button = QPushButton("Apply Mapping")
        palette_buttons.addWidget(self.palette_rebuild_mapping_button)
        palette_buttons.addWidget(self.palette_auto_mapping_button)
        palette_buttons.addWidget(self.palette_apply_mapping_button)
        palette_layout.addLayout(palette_buttons)

        palette_manage_buttons = QHBoxLayout()
        self.palette_apply_visible_button = QPushButton("Quick Apply Nearest")
        self.palette_edit_save_button = QPushButton("Edit / Save Palette")
        self.palette_open_folder_button = QPushButton("Open Palettes Folder")
        palette_manage_buttons.addWidget(self.palette_apply_visible_button)
        palette_manage_buttons.addWidget(self.palette_edit_save_button)
        palette_manage_buttons.addWidget(self.palette_open_folder_button)
        palette_layout.addLayout(palette_manage_buttons)

        bulk_layout.addWidget(palette_group)
        self.tabs.addTab(bulk_tab, "Bulk Colours")

        points_tab = QWidget()
        points_layout = QVBoxLayout(points_tab)
        self.point_search_edit = QLineEdit()
        self.point_search_edit.setPlaceholderText("Search point labels...")
        points_layout.addWidget(self.point_search_edit)

        point_search_buttons = QHBoxLayout()
        self.point_search_button = QPushButton("Search")
        self.point_select_matches_button = QPushButton("Select All Matching")
        self.point_reset_search_button = QPushButton("Reset Search")
        point_search_buttons.addWidget(self.point_search_button)
        point_search_buttons.addWidget(self.point_select_matches_button)
        point_search_buttons.addWidget(self.point_reset_search_button)
        points_layout.addLayout(point_search_buttons)

        self.points_list = QListWidget()
        points_layout.addWidget(self.points_list)

        point_bulk_group = QGroupBox("Bulk Edit Checked / Matching Points")
        point_bulk_layout = QFormLayout(point_bulk_group)
        self.point_bulk_r = self.rgb_spin()
        self.point_bulk_g = self.rgb_spin()
        self.point_bulk_b = self.rgb_spin()
        point_bulk_layout.addRow("New R", self.point_bulk_r)
        point_bulk_layout.addRow("New G", self.point_bulk_g)
        point_bulk_layout.addRow("New B", self.point_bulk_b)
        self.point_bulk_pick_colour_button = QPushButton("Pick New Colour")
        self.point_bulk_apply_colour_button = QPushButton("Change Colour of Checked Points")
        self.point_bulk_delete_button = QPushButton("Delete Checked Points")
        point_bulk_layout.addRow(self.point_bulk_pick_colour_button)
        point_bulk_layout.addRow(self.point_bulk_apply_colour_button)
        point_bulk_layout.addRow(self.point_bulk_delete_button)
        points_layout.addWidget(point_bulk_group)
        self.tabs.addTab(points_tab, "Points")

        npc_tab = QWidget()
        npc_outer_layout = QVBoxLayout(npc_tab)
        npc_outer_layout.setContentsMargins(0, 0, 0, 0)

        npc_scroll = QScrollArea()
        npc_scroll.setWidgetResizable(True)
        npc_scroll.setFrameShape(QScrollArea.NoFrame)
        npc_scroll_widget = QWidget()
        npc_layout = QVBoxLayout(npc_scroll_widget)
        npc_layout.setContentsMargins(8, 8, 8, 8)
        npc_layout.setSpacing(8)

        self.npc_workflow_splitter = QSplitter(Qt.Vertical)
        self.npc_workflow_splitter.setChildrenCollapsible(False)
        npc_layout.addWidget(self.npc_workflow_splitter, 1)

        npc_source_group, npc_source_layout = self._make_collapsible_group("NPC Match & Swap", expanded=True)
        self.npc_source_label = QLabel("No NPC data source selected.")
        self.npc_source_label.setWordWrap(True)
        npc_source_layout.addWidget(self.npc_source_label)

        npc_source_buttons = QHBoxLayout()
        self.npc_choose_source_button = QPushButton("Choose NPC Data")
        self.npc_reload_source_button = QPushButton("Reload")
        npc_source_buttons.addWidget(self.npc_choose_source_button)
        npc_source_buttons.addWidget(self.npc_reload_source_button)
        npc_source_layout.addLayout(npc_source_buttons)

        npc_action_buttons = QHBoxLayout()
        self.npc_compare_button = QPushButton("Compare Current Zone")
        self.npc_select_all_button = QPushButton("Select All")
        self.npc_select_none_button = QPushButton("Select None")
        npc_action_buttons.addWidget(self.npc_compare_button)
        npc_action_buttons.addWidget(self.npc_select_all_button)
        npc_action_buttons.addWidget(self.npc_select_none_button)
        npc_source_layout.addLayout(npc_action_buttons)

        npc_filter_row = QHBoxLayout()
        self.npc_match_type_filter_combo = QComboBox()
        self.npc_match_type_filter_combo.addItems([
            "All match types",
            "Yes",
            "Coordinate Match",
            "Possible",
            "NPC only",
            "Map only",
        ])
        self.npc_presence_filter_combo = QComboBox()
        self.npc_presence_filter_combo.addItems([
            "All presence",
            "Present on map",
            "Missing from map",
        ])
        self.npc_search_edit = QLineEdit()
        self.npc_search_edit.setPlaceholderText("Search current/NPC label, NPC name, role, or coordinates...")
        npc_filter_row.addWidget(QLabel("Filter:"))
        npc_filter_row.addWidget(self.npc_match_type_filter_combo)
        npc_filter_row.addWidget(self.npc_presence_filter_combo)
        npc_filter_row.addWidget(self.npc_search_edit, 1)
        npc_source_layout.addLayout(npc_filter_row)

        npc_preview_buttons = QHBoxLayout()
        self.npc_preview_button = QPushButton("Preview Selected")
        self.npc_clear_preview_button = QPushButton("Clear Preview")
        self.npc_apply_button = QPushButton("Apply Selected")
        self.npc_delete_from_map_button = QPushButton("Delete Selected From Map")
        self.npc_apply_button.setObjectName("primaryButton")
        self.npc_delete_from_map_button.setObjectName("dangerButton")
        self.npc_delete_from_map_button.setToolTip("Marks selected NPC-match rows that are already present on the loaded map for deletion. Save Edits removes them from the text file.")
        npc_preview_buttons.addWidget(self.npc_preview_button)
        npc_preview_buttons.addWidget(self.npc_clear_preview_button)
        npc_preview_buttons.addWidget(self.npc_apply_button)
        npc_preview_buttons.addWidget(self.npc_delete_from_map_button)
        npc_source_layout.addLayout(npc_preview_buttons)

        self.npc_summary_label = QLabel("Load a zone and click Compare Current Zone.")
        self.npc_summary_label.setWordWrap(True)
        self.npc_summary_label.setObjectName("inspectorHelperText")
        npc_source_layout.addWidget(self.npc_summary_label)

        self.npc_match_table = QTableWidget(0, 9)
        self.npc_match_table.setHorizontalHeaderLabels([
            "Use", "Match", "Conf", "Current Label", "NPC Label", "Map XYZ", "NPC XYZ", "Dist", "Role"
        ])
        self.npc_match_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.npc_match_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.npc_match_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.npc_match_table.horizontalHeader().setSectionsMovable(False)
        self.npc_match_table.setMinimumHeight(180)
        npc_source_layout.addWidget(self.npc_match_table, 1)
        self.npc_workflow_splitter.addWidget(npc_source_group)

        era_group, era_layout = self._make_collapsible_group("Expansion / Era Cleanup", expanded=False)
        era_form = QFormLayout()
        self.era_expansion_combo = QComboBox()
        for expansion_name, expansion_number in EXPANSION_OPTIONS:
            self.era_expansion_combo.addItem(f"{expansion_number}: {expansion_name}", expansion_number)
        era_form.addRow("Current expansion", self.era_expansion_combo)
        era_layout.addLayout(era_form)

        era_buttons = QHBoxLayout()
        self.era_save_default_button = QPushButton("Save Expansion Setting")
        self.era_scan_button = QPushButton("Scan Era Labels")
        self.era_select_all_button = QPushButton("Select All Flagged")
        self.era_select_none_button = QPushButton("Select None")
        era_buttons.addWidget(self.era_save_default_button)
        era_buttons.addWidget(self.era_scan_button)
        era_buttons.addWidget(self.era_select_all_button)
        era_buttons.addWidget(self.era_select_none_button)
        era_layout.addLayout(era_buttons)

        era_apply_buttons = QHBoxLayout()
        self.era_preview_button = QPushButton("Preview Removals")
        self.era_clear_preview_button = QPushButton("Clear Preview")
        self.era_remove_button = QPushButton("Remove Selected Labels")
        self.era_remove_button.setObjectName("primaryButton")
        era_apply_buttons.addWidget(self.era_preview_button)
        era_apply_buttons.addWidget(self.era_clear_preview_button)
        era_apply_buttons.addWidget(self.era_remove_button)
        era_layout.addLayout(era_apply_buttons)

        self.era_summary_label = QLabel("Select an expansion, then scan the loaded zone for labels outside that era.")
        self.era_summary_label.setWordWrap(True)
        self.era_summary_label.setObjectName("inspectorHelperText")
        era_layout.addWidget(self.era_summary_label)

        self.era_cleanup_table = QTableWidget(0, 10)
        self.era_cleanup_table.setHorizontalHeaderLabels([
            "Remove", "Status", "Current Label", "Matched NPC", "Min Exp", "Max Exp", "Map XYZ", "NPC XYZ", "Match", "Conf"
        ])
        self.era_cleanup_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.era_cleanup_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.era_cleanup_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.era_cleanup_table.horizontalHeader().setSectionsMovable(False)
        self.era_cleanup_table.setMinimumHeight(160)
        era_layout.addWidget(self.era_cleanup_table, 1)
        self.npc_workflow_splitter.addWidget(era_group)

        missing_group, missing_layout = self._make_collapsible_group("Add Missing NPCs for Current Era", expanded=True)
        missing_buttons = QHBoxLayout()
        self.missing_scan_button = QPushButton("Find Missing NPCs")
        self.missing_select_all_button = QPushButton("Select All")
        self.missing_select_none_button = QPushButton("Select None")
        missing_buttons.addWidget(self.missing_scan_button)
        missing_buttons.addWidget(self.missing_select_all_button)
        missing_buttons.addWidget(self.missing_select_none_button)
        missing_layout.addLayout(missing_buttons)

        missing_style_row = QHBoxLayout()
        missing_style_row.setSpacing(4)
        missing_style_row.setContentsMargins(0, 0, 0, 0)
        self.missing_colour_preview_label = QLabel("")
        self.missing_colour_preview_label.setFixedSize(24, 22)
        self.missing_colour_preview_label.setToolTip("Preview of the colour used for newly added NPC labels.")
        self.missing_r_spin = self.rgb_spin(); self.missing_r_spin.setValue(255)
        self.missing_g_spin = self.rgb_spin(); self.missing_g_spin.setValue(255)
        self.missing_b_spin = self.rgb_spin(); self.missing_b_spin.setValue(0)
        for spin in (self.missing_r_spin, self.missing_g_spin, self.missing_b_spin):
            spin.setMinimumWidth(54)
            spin.setMaximumWidth(64)
        self.missing_size_spin = QDoubleSpinBox()
        self.missing_size_spin.setRange(0.1, 99.0)
        self.missing_size_spin.setDecimals(2)
        self.missing_size_spin.setValue(2.0)
        self.missing_size_spin.setMinimumWidth(58)
        self.missing_size_spin.setMaximumWidth(68)
        self.missing_match_label_combo = ExistingLabelStyleCombo()
        self.missing_match_label_combo.setProperty("owner", self)
        self.missing_match_label_combo.setMinimumWidth(120)
        self.missing_match_label_combo.setMaximumWidth(190)
        self.missing_match_label_combo.setMaxVisibleItems(24)
        self.missing_match_label_combo.setSizeAdjustPolicy(QComboBox.AdjustToMinimumContentsLengthWithIcon)
        self.missing_match_label_combo.setMinimumContentsLength(24)
        self.missing_match_label_combo.setItemDelegate(ExistingLabelStyleDelegate(self.missing_match_label_combo))
        self.missing_match_label_combo.setToolTip("Choose an existing map label; the dropdown shows each label's colour before you copy its RGB/size.")
        self.missing_match_label_combo.addItem("Match existing label style...", None)
        self.missing_match_source_preview_label = QLabel("")
        self.missing_match_source_preview_label.setFixedSize(24, 22)
        self.missing_match_source_preview_label.setToolTip("Preview of the selected existing label colour before copying its style.")
        self.missing_match_label_button = QPushButton("Match")
        self.missing_match_label_button.setMaximumWidth(64)
        missing_style_row.addWidget(QLabel("Style:"))
        missing_style_row.addWidget(self.missing_colour_preview_label)
        missing_style_row.addWidget(QLabel("R"))
        missing_style_row.addWidget(self.missing_r_spin)
        missing_style_row.addWidget(QLabel("G"))
        missing_style_row.addWidget(self.missing_g_spin)
        missing_style_row.addWidget(QLabel("B"))
        missing_style_row.addWidget(self.missing_b_spin)
        missing_style_row.addWidget(QLabel("Size"))
        missing_style_row.addWidget(self.missing_size_spin)
        missing_style_row.addWidget(self.missing_match_label_combo, 1)
        missing_style_row.addWidget(self.missing_match_source_preview_label)
        missing_style_row.addWidget(self.missing_match_label_button)
        missing_layout.addLayout(missing_style_row)

        missing_apply_buttons = QHBoxLayout()
        self.missing_preview_button = QPushButton("Preview Adds")
        self.missing_clear_preview_button = QPushButton("Clear Preview")
        self.missing_add_button = QPushButton("Add Selected NPCs")
        self.missing_add_button.setObjectName("primaryButton")
        missing_apply_buttons.addWidget(self.missing_preview_button)
        missing_apply_buttons.addWidget(self.missing_clear_preview_button)
        missing_apply_buttons.addWidget(self.missing_add_button)
        missing_layout.addLayout(missing_apply_buttons)

        self.missing_summary_label = QLabel("Find NPCs valid for the selected expansion that are not already on the loaded map.")
        self.missing_summary_label.setWordWrap(True)
        self.missing_summary_label.setObjectName("inspectorHelperText")
        missing_layout.addWidget(self.missing_summary_label)

        self.missing_npc_table = QTableWidget(0, 7)
        self.missing_npc_table.setHorizontalHeaderLabels([
            "Add", "NPC Label", "NPC XYZ", "Min Exp", "Max Exp", "NPC Name", "Role"
        ])
        self.missing_npc_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.missing_npc_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.missing_npc_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.missing_npc_table.horizontalHeader().setSectionsMovable(False)
        self.missing_npc_table.setMinimumHeight(180)
        missing_layout.addWidget(self.missing_npc_table, 1)
        self.npc_workflow_splitter.addWidget(missing_group)
        self.npc_workflow_groups = [npc_source_group, era_group, missing_group]
        self.npc_workflow_splitter.setSizes([360, 38, 360])
        QTimer.singleShot(0, self.rebalance_npc_workflow_boxes)

        npc_scroll.setWidget(npc_scroll_widget)
        npc_outer_layout.addWidget(npc_scroll)
        self.tabs.addTab(npc_tab, "NPC Match")

        zones_tab = QWidget()
        zones_layout = QVBoxLayout(zones_tab)

        folder_row = QHBoxLayout()
        self.map_folder_edit = QLineEdit()
        self.map_folder_edit.setPlaceholderText("Choose your EQ maps folder...")
        self.choose_map_folder_button = QPushButton("Choose Map Folder")
        folder_row.addWidget(self.map_folder_edit)
        folder_row.addWidget(self.choose_map_folder_button)
        zones_layout.addLayout(folder_row)

        zone_search_row = QHBoxLayout()
        self.zone_search_edit = QLineEdit()
        self.zone_search_edit.setPlaceholderText("Search zones by full name or shortname...")
        self.zone_search_button = QPushButton("Search")
        self.zone_reset_button = QPushButton("Reset")
        zone_search_row.addWidget(self.zone_search_edit)
        zone_search_row.addWidget(self.zone_search_button)
        zone_search_row.addWidget(self.zone_reset_button)
        zones_layout.addLayout(zone_search_row)

        self.zones_list = QListWidget()
        zones_layout.addWidget(self.zones_list)
        self.open_selected_zone_button = QPushButton("Open Selected Zone")
        zones_layout.addWidget(self.open_selected_zone_button)
        self.tabs.addTab(zones_tab, "Zones")

        pending_tab = QWidget()
        pending_layout = QVBoxLayout(pending_tab)
        pending_layout.addWidget(QLabel("Unsaved changes"))
        self.pending_changes_list = QListWidget()
        pending_layout.addWidget(self.pending_changes_list)
        pending_buttons = QHBoxLayout()
        self.pending_refresh_button = QPushButton("Refresh")
        self.pending_revert_all_button = QPushButton("Revert All Unsaved")
        self.pending_save_button = QPushButton("Save Edits")
        pending_buttons.addWidget(self.pending_refresh_button)
        pending_buttons.addWidget(self.pending_revert_all_button)
        pending_buttons.addWidget(self.pending_save_button)
        pending_layout.addLayout(pending_buttons)
        self.tabs.addTab(pending_tab, "Pending Changes")


        layout.addStretch(1)

        self.edit_mode_combo.currentTextChanged.connect(self.main_window.on_edit_mode_changed)
        self.undo_button.clicked.connect(self.main_window.undo)
        self.redo_button.clicked.connect(self.main_window.redo)
        self.reload_button.clicked.connect(self.main_window.reload_from_disk)
        self.restore_button.clicked.connect(self.main_window.restore_from_backup)
        self.inspector_close_button.clicked.connect(self.main_window.toggle_sidebar)
        self.pick_color_btn.clicked.connect(self.pick_color)
        self.apply_button.clicked.connect(self.apply_changes)
        self.delete_selected_button.clicked.connect(self.main_window.delete_selected_records)
        self.point_colour_list.itemSelectionChanged.connect(lambda: self.on_colour_selection_changed("points"))
        self.line_colour_list.itemSelectionChanged.connect(lambda: self.on_colour_selection_changed("lines"))
        self.bulk_pick_colour_button.clicked.connect(self.pick_bulk_colour)
        self.bulk_match_colour_button.clicked.connect(self.match_bulk_colour_from_list)
        self.bulk_apply_colour_button.clicked.connect(self.main_window.apply_bulk_colour_to_selected_matching)
        self.palette_apply_visible_button.clicked.connect(self.main_window.apply_palette_to_visible_records)
        self.palette_rebuild_mapping_button.clicked.connect(self.rebuild_palette_mapping_preview)
        self.palette_auto_mapping_button.clicked.connect(self.auto_map_palette_preview)
        self.palette_apply_mapping_button.clicked.connect(self.main_window.apply_palette_mapping_preview)
        self.palette_edit_save_button.clicked.connect(self.edit_save_palette)
        self.palette_open_folder_button.clicked.connect(self.open_palettes_folder)
        self.palette_combo.currentTextChanged.connect(lambda _text: (self.update_palette_preview(), self.rebuild_palette_mapping_preview()))
        self.palette_target_combo.currentTextChanged.connect(lambda _text: self.rebuild_palette_mapping_preview())
        self.point_search_button.clicked.connect(self.rebuild_points_list)
        self.point_search_edit.returnPressed.connect(self.rebuild_points_list)
        self.point_select_matches_button.clicked.connect(self.check_matching_points)
        self.point_reset_search_button.clicked.connect(self.reset_point_search)
        self.point_bulk_pick_colour_button.clicked.connect(self.pick_point_bulk_colour)
        self.point_bulk_apply_colour_button.clicked.connect(self.main_window.bulk_recolour_checked_points)
        self.point_bulk_delete_button.clicked.connect(self.main_window.bulk_delete_checked_points)
        self.npc_choose_source_button.clicked.connect(self.main_window.choose_npc_data_source)
        self.npc_reload_source_button.clicked.connect(self.main_window.reload_npc_data_source)
        self.npc_compare_button.clicked.connect(self.main_window.compare_current_zone_to_npc_data)
        self.npc_select_all_button.clicked.connect(lambda: self.set_all_npc_match_rows_checked(True))
        self.npc_select_none_button.clicked.connect(lambda: self.set_all_npc_match_rows_checked(False))
        self.npc_match_type_filter_combo.currentIndexChanged.connect(lambda _index: self.apply_npc_match_filters())
        self.npc_presence_filter_combo.currentIndexChanged.connect(lambda _index: self.apply_npc_match_filters())
        self.npc_search_edit.textChanged.connect(lambda _text: self.apply_npc_match_filters())
        self.npc_preview_button.clicked.connect(self.main_window.preview_selected_npc_matches)
        self.npc_clear_preview_button.clicked.connect(self.main_window.clear_npc_match_preview)
        self.npc_apply_button.clicked.connect(self.main_window.apply_selected_npc_matches)
        self.npc_delete_from_map_button.clicked.connect(self.main_window.delete_selected_npc_match_points_from_map)
        self.era_save_default_button.clicked.connect(self.main_window.save_current_expansion_setting)
        self.era_scan_button.clicked.connect(self.main_window.scan_current_zone_for_era_cleanup)
        self.era_select_all_button.clicked.connect(lambda: self.set_all_era_cleanup_rows_checked(True))
        self.era_select_none_button.clicked.connect(lambda: self.set_all_era_cleanup_rows_checked(False))
        self.era_preview_button.clicked.connect(self.main_window.preview_selected_era_cleanup)
        self.era_clear_preview_button.clicked.connect(self.main_window.clear_era_cleanup_preview)
        self.era_remove_button.clicked.connect(self.main_window.remove_selected_era_cleanup_labels)
        self.missing_scan_button.clicked.connect(self.main_window.find_missing_npcs_for_current_era)
        self.missing_select_all_button.clicked.connect(lambda: self.set_all_missing_npc_rows_checked(True))
        self.missing_select_none_button.clicked.connect(lambda: self.set_all_missing_npc_rows_checked(False))
        self.missing_preview_button.clicked.connect(self.main_window.preview_selected_missing_npcs)
        self.missing_clear_preview_button.clicked.connect(self.main_window.clear_missing_npc_preview)
        self.missing_add_button.clicked.connect(self.main_window.add_selected_missing_npcs_to_map)
        self.missing_r_spin.valueChanged.connect(lambda _value: self.update_missing_colour_preview())
        self.missing_g_spin.valueChanged.connect(lambda _value: self.update_missing_colour_preview())
        self.missing_b_spin.valueChanged.connect(lambda _value: self.update_missing_colour_preview())
        self.missing_match_label_combo.currentIndexChanged.connect(lambda _index: self.update_missing_match_source_preview())
        self.missing_match_label_button.clicked.connect(self.match_missing_style_from_existing_label)
        self.update_missing_colour_preview()
        self.update_missing_match_source_preview()
        self.choose_map_folder_button.clicked.connect(self.main_window.choose_map_folder)
        self.zone_search_button.clicked.connect(self.main_window.rebuild_zone_list)
        self.zone_search_edit.returnPressed.connect(self.main_window.rebuild_zone_list)
        self.zone_search_edit.textChanged.connect(lambda _text: self.main_window.rebuild_zone_list())
        self.zone_reset_button.clicked.connect(self.main_window.reset_zone_search)
        self.open_selected_zone_button.clicked.connect(self.main_window.open_selected_zone)
        self.zones_list.itemDoubleClicked.connect(lambda item: self.main_window.open_selected_zone())
        self.pending_refresh_button.clicked.connect(self.rebuild_pending_changes)
        self.pending_revert_all_button.clicked.connect(self.main_window.reload_from_disk)
        self.pending_save_button.clicked.connect(self.main_window.save_edits)
        self.refresh_palette_combo()


    def update_missing_colour_preview(self) -> None:
        if not hasattr(self, "missing_colour_preview_label"):
            return
        r = self.missing_r_spin.value()
        g = self.missing_g_spin.value()
        b = self.missing_b_spin.value()
        border = "#222222" if (r + g + b) > 360 else "#dddddd"
        self.missing_colour_preview_label.setStyleSheet(
            f"background-color: rgb({r}, {g}, {b}); border: 1px solid {border}; border-radius: 3px;"
        )


    def _existing_label_style_icon(self, point: MapPointRecord) -> QIcon:
        pixmap = QPixmap(22, 16)
        pixmap.fill(Qt.transparent)
        painter = QPainter(pixmap)
        try:
            r, g, b = int(point.r), int(point.g), int(point.b)
            border = QColor(30, 30, 30) if (r + g + b) > 360 else QColor(220, 220, 220)
            painter.setRenderHint(QPainter.Antialiasing, True)
            painter.setPen(QPen(border, 1))
            painter.setBrush(QBrush(QColor(r, g, b)))
            painter.drawRoundedRect(QRect(1, 1, 20, 14), 3, 3)
        finally:
            painter.end()
        return QIcon(pixmap)

    def rebuild_missing_style_match_combo(self) -> None:
        if not hasattr(self, "missing_match_label_combo"):
            return
        combo = self.missing_match_label_combo
        previous_label = combo.currentText()
        combo.blockSignals(True)
        combo.clear()
        combo.addItem("Match existing label style...", None)
        seen: set[str] = set()
        for point in sorted(getattr(self.main_window.loaded_map, "points", []), key=lambda p: p.label.lower()):
            label = point.label.strip()
            if not label or label in seen:
                continue
            seen.add(label)
            combo.addItem(self._existing_label_style_icon(point), label, point)
            combo.setItemData(combo.count() - 1, f"RGB {int(point.r)}, {int(point.g)}, {int(point.b)}; size {float(point.size):g}", Qt.ToolTipRole)
        if previous_label:
            index = combo.findText(previous_label)
            if index >= 0:
                combo.setCurrentIndex(index)
        if combo.view() is not None:
            combo.view().setMinimumWidth(max(combo.width(), 360))
        combo.blockSignals(False)
        self.update_missing_match_source_preview()

    def update_missing_match_source_preview(self) -> None:
        if not hasattr(self, "missing_match_source_preview_label"):
            return
        point = self.missing_match_label_combo.currentData() if hasattr(self, "missing_match_label_combo") else None
        if point is None:
            self.missing_match_source_preview_label.setStyleSheet(
                "background-color: transparent; border: 1px dashed #888888; border-radius: 3px;"
            )
            self.missing_match_source_preview_label.setToolTip("Choose an existing label to preview its colour before matching.")
            return
        r, g, b = int(point.r), int(point.g), int(point.b)
        border = "#222222" if (r + g + b) > 360 else "#dddddd"
        self.missing_match_source_preview_label.setStyleSheet(
            f"background-color: rgb({r}, {g}, {b}); border: 1px solid {border}; border-radius: 3px;"
        )
        self.missing_match_source_preview_label.setToolTip(
            f"Selected existing label style: RGB {r}, {g}, {b}; size {float(point.size):g}"
        )

    def match_missing_style_from_existing_label(self) -> None:
        if not hasattr(self, "missing_match_label_combo"):
            return
        point = self.missing_match_label_combo.currentData()
        if point is None:
            QMessageBox.information(self, "Match Existing Label", "Choose an existing label from the style dropdown first.")
            return
        self.missing_r_spin.setValue(int(point.r))
        self.missing_g_spin.setValue(int(point.g))
        self.missing_b_spin.setValue(int(point.b))
        self.missing_size_spin.setValue(float(point.size))
        self.update_missing_colour_preview()

    def set_npc_source_path(self, path_text: str) -> None:
        if hasattr(self, "npc_source_label"):
            if path_text:
                self.npc_source_label.setText(f"NPC data: {path_text}")
            else:
                self.npc_source_label.setText("No NPC data source selected.")

    def _make_table_columns_user_resizable(self, table: QTableWidget, widths: Optional[list[int]] = None) -> None:
        """Keep table columns manually resizable after content is populated."""
        header = table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setStretchLastSection(False)
        if widths:
            for column, width in enumerate(widths):
                if column < table.columnCount():
                    table.setColumnWidth(column, width)

    def _sync_npc_match_table_edits_to_results(self) -> None:
        """Preserve checkbox and edited label state before filtering/rebuilding the match table."""
        if not hasattr(self, "npc_match_table"):
            return
        table = self.npc_match_table
        for row in range(table.rowCount()):
            check_item = table.item(row, 0)
            if check_item is None:
                continue
            result = check_item.data(Qt.UserRole)
            if not isinstance(result, NpcMatchResult):
                continue
            result.selected = check_item.checkState() == Qt.Checked
            label_item = table.item(row, 4)
            if label_item is not None and result.npc_row is not None:
                result.npc_row.npc_label = label_item.text().strip() or result.npc_row.npc_label

    def npc_match_result_matches_filters(self, result: NpcMatchResult) -> bool:
        match_filter = self.npc_match_type_filter_combo.currentText() if hasattr(self, "npc_match_type_filter_combo") else "All match types"
        if match_filter != "All match types" and result.match_type != match_filter:
            return False

        presence_filter = self.npc_presence_filter_combo.currentText() if hasattr(self, "npc_presence_filter_combo") else "All presence"
        if presence_filter == "Present on map" and result.map_record is None:
            return False
        if presence_filter == "Missing from map" and result.map_record is not None:
            return False

        query = self.npc_search_edit.text().strip().lower() if hasattr(self, "npc_search_edit") else ""
        if query:
            values: list[str] = [
                result.match_type,
                result.confidence,
            ]
            if result.map_record is not None:
                values.extend([
                    result.map_record.label,
                    f"{result.map_record.x:.4f}",
                    f"{result.map_record.y:.4f}",
                    f"{result.map_record.z:.4f}",
                ])
            if result.npc_row is not None:
                values.extend([
                    result.npc_row.npc_label,
                    result.npc_row.npc_name,
                    result.npc_row.npc_role,
                    f"{result.npc_row.x:.4f}",
                    f"{result.npc_row.y:.4f}",
                    f"{result.npc_row.z:.4f}",
                    result.npc_row.min_expansion_name,
                    result.npc_row.max_expansion_name,
                ])
            haystack = " ".join(str(value).lower() for value in values if value)
            if query not in haystack:
                return False
        return True

    def apply_npc_match_filters(self) -> None:
        if not hasattr(self, "npc_match_table"):
            return
        self._sync_npc_match_table_edits_to_results()
        all_results = list(getattr(self.main_window, "npc_match_results", []))
        filtered = [result for result in all_results if self.npc_match_result_matches_filters(result)]
        self.rebuild_npc_match_table(filtered, preserve_widths=True)
        if hasattr(self, "npc_summary_label") and all_results:
            checked_visible = sum(1 for result in filtered if result.selected)
            self.npc_summary_label.setText(
                getattr(self.main_window, "npc_match_base_summary", "")
                + f"\nVisible after filters: {len(filtered):,} / {len(all_results):,}    Checked visible: {checked_visible:,}"
            )

    def set_all_npc_match_rows_checked(self, checked: bool) -> None:
        if not hasattr(self, "npc_match_table"):
            return
        state = Qt.Checked if checked else Qt.Unchecked
        for row in range(self.npc_match_table.rowCount()):
            item = self.npc_match_table.item(row, 0)
            if item is not None and item.flags() & Qt.ItemIsUserCheckable:
                item.setCheckState(state)
                result = item.data(Qt.UserRole)
                if isinstance(result, NpcMatchResult):
                    result.selected = checked

    def checked_npc_match_results(self) -> list[NpcMatchResult]:
        results: list[NpcMatchResult] = []
        if not hasattr(self, "npc_match_table"):
            return results
        self._sync_npc_match_table_edits_to_results()
        for row in range(self.npc_match_table.rowCount()):
            check_item = self.npc_match_table.item(row, 0)
            if check_item is None or check_item.checkState() != Qt.Checked:
                continue
            result = check_item.data(Qt.UserRole)
            if not isinstance(result, NpcMatchResult):
                continue
            label_item = self.npc_match_table.item(row, 4)
            if label_item is not None and result.npc_row is not None:
                result.npc_row.npc_label = label_item.text().strip() or result.npc_row.npc_label
            results.append(result)
        return results

    def rebuild_npc_match_table(self, results: list[NpcMatchResult], preserve_widths: bool = False) -> None:
        if not hasattr(self, "npc_match_table"):
            return
        table = self.npc_match_table
        previous_widths = [table.columnWidth(column) for column in range(table.columnCount())] if preserve_widths else []
        table.setRowCount(0)
        table.setSortingEnabled(False)
        for result in results:
            row = table.rowCount()
            table.insertRow(row)

            use_item = QTableWidgetItem("")
            use_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable)
            use_item.setCheckState(Qt.Checked if result.selected else Qt.Unchecked)
            use_item.setData(Qt.UserRole, result)
            table.setItem(row, 0, use_item)

            table.setItem(row, 1, QTableWidgetItem(result.match_type))
            table.setItem(row, 2, QTableWidgetItem(result.confidence))
            current_label = result.map_record.label if result.map_record is not None else ""
            table.setItem(row, 3, QTableWidgetItem(current_label))

            npc_label = result.npc_row.npc_label if result.npc_row is not None else ""
            label_item = QTableWidgetItem(npc_label)
            if result.npc_row is None:
                label_item.setFlags(label_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(row, 4, label_item)

            if result.map_record is not None:
                map_xyz = f"{result.map_record.x:.2f}, {result.map_record.y:.2f}, {result.map_record.z:.2f}"
            else:
                map_xyz = ""
            table.setItem(row, 5, QTableWidgetItem(map_xyz))

            if result.npc_row is not None:
                npc_xyz = f"{result.npc_row.x:.2f}, {result.npc_row.y:.2f}, {result.npc_row.z:.2f}"
                role = result.npc_row.npc_role
            else:
                npc_xyz = ""
                role = ""
            table.setItem(row, 6, QTableWidgetItem(npc_xyz))
            dist_text = f"{result.distance:.2f}" if result.distance is not None else ""
            table.setItem(row, 7, QTableWidgetItem(dist_text))
            table.setItem(row, 8, QTableWidgetItem(role))

        if preserve_widths and previous_widths:
            self._make_table_columns_user_resizable(table, previous_widths)
        else:
            table.resizeColumnsToContents()
            self._make_table_columns_user_resizable(table, [46, 120, 70, 190, 210, 155, 155, 70, 140])

    def set_expansion_number(self, expansion_number: int) -> None:
        if not hasattr(self, "era_expansion_combo"):
            return
        for index in range(self.era_expansion_combo.count()):
            if int(self.era_expansion_combo.itemData(index)) == int(expansion_number):
                self.era_expansion_combo.setCurrentIndex(index)
                return

    def current_expansion_number(self) -> int:
        if not hasattr(self, "era_expansion_combo"):
            return 0
        value = self.era_expansion_combo.currentData()
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    def current_expansion_name(self) -> str:
        if not hasattr(self, "era_expansion_combo"):
            return "Classic"
        text = self.era_expansion_combo.currentText()
        return text.split(": ", 1)[1] if ": " in text else text

    def set_all_era_cleanup_rows_checked(self, checked: bool) -> None:
        if not hasattr(self, "era_cleanup_table"):
            return
        state = Qt.Checked if checked else Qt.Unchecked
        for row in range(self.era_cleanup_table.rowCount()):
            item = self.era_cleanup_table.item(row, 0)
            if item is not None and item.flags() & Qt.ItemIsUserCheckable:
                item.setCheckState(state)

    def checked_era_cleanup_results(self) -> list[EraCleanupResult]:
        results: list[EraCleanupResult] = []
        if not hasattr(self, "era_cleanup_table"):
            return results
        for row in range(self.era_cleanup_table.rowCount()):
            check_item = self.era_cleanup_table.item(row, 0)
            if check_item is None or check_item.checkState() != Qt.Checked:
                continue
            result = check_item.data(Qt.UserRole)
            if isinstance(result, EraCleanupResult):
                results.append(result)
        return results

    def rebuild_era_cleanup_table(self, results: list[EraCleanupResult]) -> None:
        if not hasattr(self, "era_cleanup_table"):
            return
        table = self.era_cleanup_table
        table.setRowCount(0)
        table.setSortingEnabled(False)
        for result in results:
            row = table.rowCount()
            table.insertRow(row)
            use_item = QTableWidgetItem("")
            use_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable)
            use_item.setCheckState(Qt.Checked if result.selected else Qt.Unchecked)
            use_item.setData(Qt.UserRole, result)
            table.setItem(row, 0, use_item)
            table.setItem(row, 1, QTableWidgetItem(result.era_status))
            table.setItem(row, 2, QTableWidgetItem(result.map_record.label))
            table.setItem(row, 3, QTableWidgetItem(result.npc_row.npc_label))
            table.setItem(row, 4, QTableWidgetItem(expansion_label(result.npc_row.min_expansion_number)))
            table.setItem(row, 5, QTableWidgetItem(expansion_label(result.npc_row.max_expansion_number)))
            table.setItem(row, 6, QTableWidgetItem(f"{result.map_record.x:.2f}, {result.map_record.y:.2f}, {result.map_record.z:.2f}"))
            table.setItem(row, 7, QTableWidgetItem(f"{result.npc_row.x:.2f}, {result.npc_row.y:.2f}, {result.npc_row.z:.2f}"))
            table.setItem(row, 8, QTableWidgetItem(result.match_type))
            table.setItem(row, 9, QTableWidgetItem(result.confidence))
        table.resizeColumnsToContents()
        self._make_table_columns_user_resizable(table, [70, 135, 210, 210, 140, 140, 155, 155, 100, 70])

    def set_all_missing_npc_rows_checked(self, checked: bool) -> None:
        if not hasattr(self, "missing_npc_table"):
            return
        state = Qt.Checked if checked else Qt.Unchecked
        for row in range(self.missing_npc_table.rowCount()):
            item = self.missing_npc_table.item(row, 0)
            if item is not None and item.flags() & Qt.ItemIsUserCheckable:
                item.setCheckState(state)

    def checked_missing_npc_results(self) -> list[MissingNpcResult]:
        results: list[MissingNpcResult] = []
        if not hasattr(self, "missing_npc_table"):
            return results
        for row in range(self.missing_npc_table.rowCount()):
            check_item = self.missing_npc_table.item(row, 0)
            if check_item is None or check_item.checkState() != Qt.Checked:
                continue
            result = check_item.data(Qt.UserRole)
            if not isinstance(result, MissingNpcResult):
                continue
            label_item = self.missing_npc_table.item(row, 1)
            if label_item is not None:
                result.npc_label = label_item.text().strip() or result.npc_label
            results.append(result)
        return results

    def rebuild_missing_npc_table(self, results: list[MissingNpcResult]) -> None:
        if not hasattr(self, "missing_npc_table"):
            return
        table = self.missing_npc_table
        table.setRowCount(0)
        table.setSortingEnabled(False)
        for result in results:
            npc = result.npc_row
            row = table.rowCount()
            table.insertRow(row)
            use_item = QTableWidgetItem("")
            use_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable)
            use_item.setCheckState(Qt.Checked if result.selected else Qt.Unchecked)
            use_item.setData(Qt.UserRole, result)
            table.setItem(row, 0, use_item)
            table.setItem(row, 1, QTableWidgetItem(result.npc_label))
            table.setItem(row, 2, QTableWidgetItem(f"{npc.x:.2f}, {npc.y:.2f}, {npc.z:.2f}"))
            table.setItem(row, 3, QTableWidgetItem(expansion_label(npc.min_expansion_number)))
            table.setItem(row, 4, QTableWidgetItem(expansion_label(npc.max_expansion_number)))
            table.setItem(row, 5, QTableWidgetItem(npc.npc_name))
            table.setItem(row, 6, QTableWidgetItem(npc.npc_role))
        table.resizeColumnsToContents()
        self._make_table_columns_user_resizable(table, [55, 220, 155, 140, 140, 180, 160])

    def coord_spin(self) -> QDoubleSpinBox:
        spin = QDoubleSpinBox()
        spin.setRange(-100000.0, 100000.0)
        spin.setDecimals(4)
        spin.setSingleStep(1.0)
        spin.setButtonSymbols(QAbstractSpinBox.UpDownArrows)
        spin.setMinimumHeight(24)
        return spin

    def rgb_spin(self) -> QSpinBox:
        spin = QSpinBox()
        spin.setRange(0, 255)
        spin.setButtonSymbols(QAbstractSpinBox.UpDownArrows)
        spin.setMinimumHeight(24)
        spin.setMinimumWidth(78)
        return spin

    def all_controls(self):
        return [
            self.label_edit,
            self.x_spin, self.y_spin, self.z_spin,
            self.x1_spin, self.y1_spin, self.z1_spin,
            self.x2_spin, self.y2_spin, self.z2_spin,
            self.r_spin, self.g_spin, self.b_spin,
            self.size_spin,
            self.pick_color_btn,
            self.apply_button,
        ]

    def edit_mode(self) -> str:
        return self.edit_mode_combo.currentText()

    def set_record(self, record: Optional[Any]) -> None:
        self._loading = True
        self.current_record = record

        for widget in self.all_controls():
            widget.setEnabled(record is not None)

        is_point = isinstance(record, MapPointRecord)
        is_line = isinstance(record, MapLineRecord)

        self.label_edit.setEnabled(is_point)
        self.size_spin.setEnabled(is_point)
        for widget in [self.x_spin, self.y_spin, self.z_spin]:
            widget.setEnabled(is_point)
        for widget in [self.x1_spin, self.y1_spin, self.z1_spin, self.x2_spin, self.y2_spin, self.z2_spin]:
            widget.setEnabled(is_line)

        if record is None:
            self.title_label.setText("No selection")
            self.source_label.setText("")
            self.label_edit.setText("")
            self.raw_label.setText("")
        elif is_point:
            self.title_label.setText("Selected Point / Label")
            self.source_label.setText(f"{record.file_path.name}:{record.line_index + 1 if record.line_index >= 0 else 'new'}")
            self.label_edit.setText(record.label)
            self.x_spin.setValue(record.x)
            self.y_spin.setValue(record.y)
            self.z_spin.setValue(record.z)
            self.r_spin.setValue(record.r)
            self.g_spin.setValue(record.g)
            self.b_spin.setValue(record.b)
            self.size_spin.setValue(record.size)
            self.raw_label.setText(record.to_map_line())
        elif is_line:
            self.title_label.setText("Selected Line")
            self.source_label.setText(f"{record.file_path.name}:{record.line_index + 1 if record.line_index >= 0 else 'new'}")
            self.label_edit.setText("")
            self.x1_spin.setValue(record.x1)
            self.y1_spin.setValue(record.y1)
            self.z1_spin.setValue(record.z1)
            self.x2_spin.setValue(record.x2)
            self.y2_spin.setValue(record.y2)
            self.z2_spin.setValue(record.z2)
            self.r_spin.setValue(record.r)
            self.g_spin.setValue(record.g)
            self.b_spin.setValue(record.b)
            self.raw_label.setText(record.to_map_line())

        if record is None:
            # Do not clear a multi-select message when show_multi_selection() intentionally sets one afterwards.
            if self.title_label.text() != "Multiple Items Selected":
                self.set_multi_selection_summary(0, 0)
        else:
            is_point_summary = 1 if isinstance(record, MapPointRecord) else 0
            is_line_summary = 1 if isinstance(record, MapLineRecord) else 0
            self.set_multi_selection_summary(is_point_summary, is_line_summary)
            if is_point:
                self.footer_summary_label.setText(f"Single selection — Point from {record.file_path.name}")
            elif is_line:
                self.footer_summary_label.setText(f"Single selection — Line from {record.file_path.name}")

        self.update_color_preview()
        self._loading = False

    def update_color_preview(self) -> None:
        color = QColor(self.r_spin.value(), self.g_spin.value(), self.b_spin.value())
        self.color_preview.setStyleSheet(
            f"background-color: rgb({color.red()}, {color.green()}, {color.blue()}); border: 1px solid #888;"
        )

    def pick_color(self) -> None:
        start = QColor(self.r_spin.value(), self.g_spin.value(), self.b_spin.value())
        color = choose_colour_dialog(start, self, "Choose colour")
        if not color.isValid():
            return
        self.r_spin.setValue(color.red())
        self.g_spin.setValue(color.green())
        self.b_spin.setValue(color.blue())
        self.update_color_preview()

    def apply_changes(self) -> None:
        record = self.current_record
        if record is None:
            if self.title_label.text() == "Multiple Items Selected":
                self.main_window.recolour_selected_records_from_panel()
            return

        if isinstance(record, MapPointRecord):
            before = snapshot_point(record)
            record.label = self.label_edit.text().strip()
            record.x = self.x_spin.value()
            record.y = self.y_spin.value()
            record.z = self.z_spin.value()
            record.r = self.r_spin.value()
            record.g = self.g_spin.value()
            record.b = self.b_spin.value()
            record.size = self.size_spin.value()
            after = snapshot_point(record)

            if before != after:
                if before["label"] != after["label"]:
                    action = "Edit point label"
                elif (before["r"], before["g"], before["b"]) != (after["r"], after["g"], after["b"]):
                    action = "Edit point colour"
                else:
                    action = "Edit point"
                self.main_window.add_undo(action, record, before, after)
                record.dirty = True
                self.main_window.update_record_items(record)

        elif isinstance(record, MapLineRecord):
            before = snapshot_line(record)
            record.x1 = self.x1_spin.value()
            record.y1 = self.y1_spin.value()
            record.z1 = self.z1_spin.value()
            record.x2 = self.x2_spin.value()
            record.y2 = self.y2_spin.value()
            record.z2 = self.z2_spin.value()
            record.r = self.r_spin.value()
            record.g = self.g_spin.value()
            record.b = self.b_spin.value()
            after = snapshot_line(record)

            if before != after:
                if (before["r"], before["g"], before["b"]) != (after["r"], after["g"], after["b"]):
                    action = "Edit line colour"
                else:
                    action = "Edit line"
                self.main_window.add_undo(action, record, before, after)
                record.dirty = True
                self.main_window.update_record_items(record)

        self.set_record(record)
        self.main_window.update_dirty_indicator()

    def rebuild_layers(self) -> None:
        while self.layers_layout.count():
            item = self.layers_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

        self.layer_checkboxes.clear()
        self.active_file_combo.blockSignals(True)
        self.active_file_combo.clear()

        if not self.main_window.loaded_files:
            self.layers_layout.addWidget(QLabel("Open files to show layer toggles."))
            self.active_file_combo.blockSignals(False)
            return

        for file_path in self.main_window.loaded_files:
            line_count = sum(1 for r in self.main_window.loaded_map.lines if r.file_path == file_path)
            point_count = sum(1 for r in self.main_window.loaded_map.points if r.file_path == file_path)
            cb = QCheckBox(f"{file_path.name}  ({line_count} L, {point_count} P)")
            cb.setChecked(self.main_window.layer_visible.get(file_path, True))
            cb.toggled.connect(lambda checked, fp=file_path: self.main_window.set_layer_visible(fp, checked))
            self.layers_layout.addWidget(cb)
            self.layer_checkboxes[file_path] = cb
            self.active_file_combo.addItem(file_path.name, str(file_path))

        self.layers_layout.addStretch(1)
        self.active_file_combo.blockSignals(False)
        self.rebuild_colour_list()
        self.rebuild_points_list()



    def rebuild_pending_changes(self) -> None:
        if not hasattr(self, "pending_changes_list"):
            return
        self.pending_changes_list.clear()
        for record in self.main_window.dirty_records():
            if isinstance(record, MapPointRecord):
                record_type = "Point"
                label = record.label
                colour = f"RGB ({record.r}, {record.g}, {record.b})"
            else:
                record_type = "Line"
                label = ""
                colour = f"RGB ({record.r}, {record.g}, {record.b})"
            action = "Deleted" if getattr(record, "deleted", False) else ("Added" if record.line_index < 0 else "Edited")
            line_num = "new" if record.line_index < 0 else str(record.line_index + 1)
            item = QListWidgetItem(f"{action:7} {record_type:5} {record.file_path.name}:{line_num} {colour} {label}")
            item.setData(Qt.UserRole, id(record))
            self.pending_changes_list.addItem(item)

    def set_multi_selection_summary(self, point_count: int, line_count: int) -> None:
        total = point_count + line_count
        if total <= 0:
            self.multi_select_label.setText("Multi-select: none")
            self.delete_selected_button.setEnabled(False)
        elif total == 1:
            self.multi_select_label.setText(f"Selected: {point_count} point(s), {line_count} line(s)")
            self.delete_selected_button.setEnabled(True)
        else:
            self.multi_select_label.setText(f"Multi-select: {total} item(s) — {point_count} point(s), {line_count} line(s)")
            self.delete_selected_button.setEnabled(True)

    def show_multi_selection(self, point_count: int, line_count: int) -> None:
        self.set_record(None)
        self.title_label.setText("Multiple Items Selected")
        self.source_label.setText(f"{point_count} point(s), {line_count} line(s)")
        self.raw_label.setText(
            "Coordinate fields are disabled for multi-select. "
            "Set RGB and click Apply Changes to recolour all selected records, "
            "or use Delete Selected Points/Lines to mark them for deletion."
        )

        # Coordinates, label, and point size stay greyed out for mixed selections.
        for widget in [
            self.label_edit,
            self.x_spin, self.y_spin, self.z_spin,
            self.x1_spin, self.y1_spin, self.z1_spin,
            self.x2_spin, self.y2_spin, self.z2_spin,
            self.size_spin,
        ]:
            widget.setEnabled(False)

        # Colour editing remains available for multi-select.
        for widget in [self.r_spin, self.g_spin, self.b_spin, self.pick_color_btn, self.apply_button]:
            widget.setEnabled(True)

        selected_records = self.main_window.selected_map_records()
        if selected_records:
            first = selected_records[0]
            if hasattr(first, "r"):
                self.r_spin.setValue(first.r)
                self.g_spin.setValue(first.g)
                self.b_spin.setValue(first.b)
                colours = {(record.r, record.g, record.b) for record in selected_records if hasattr(record, "r")}
                if len(colours) == 1:
                    self.raw_label.setText(self.raw_label.text() + f"\nCurrent selection colour: {next(iter(colours))}")
                else:
                    self.raw_label.setText(self.raw_label.text() + f"\nMixed colours selected. RGB fields start from the first selected record: {(first.r, first.g, first.b)}")
                self.update_color_preview()

        self.set_multi_selection_summary(point_count, line_count)

    def on_colour_selection_changed(self, record_type: str) -> None:
        colours = self.main_window.selected_colours_from_bulk_tab(record_type)
        if colours:
            self.bulk_old_rgb_label.setText(f"{record_type}: {', '.join(str(c) for c in colours)}")
            self.main_window.current_bulk_colour_type = record_type
            self.main_window.select_records_by_colour(record_type)

    def colour_icon(self, rgb: tuple[int, int, int]) -> QPixmap:
        pixmap = QPixmap(18, 18)
        pixmap.fill(QColor(*rgb))
        return pixmap

    def list_text_colour(self) -> QColor:
        return QColor(245, 245, 245) if getattr(self.main_window, "dark_ui", False) else QColor(0, 0, 0)


    def refresh_palette_combo(self) -> None:
        if not hasattr(self, "palette_combo"):
            return
        current = self.palette_combo.currentText()
        self.palette_combo.blockSignals(True)
        self.palette_combo.clear()
        for name in sorted(available_palettes().keys()):
            self.palette_combo.addItem(name)
        if current:
            index = self.palette_combo.findText(current)
            if index >= 0:
                self.palette_combo.setCurrentIndex(index)
        self.palette_combo.blockSignals(False)
        self.update_palette_preview()
        self.rebuild_palette_mapping_preview()

    def selected_palette(self) -> Optional[dict[str, Any]]:
        if not hasattr(self, "palette_combo"):
            return None
        return available_palettes().get(self.palette_combo.currentText())

    def update_palette_preview(self) -> None:
        palette = self.selected_palette()
        if not palette:
            self.palette_preview_label.setText("No palette selected.")
            return
        entries = palette_entries(palette)
        sample = ", ".join(entry["name"] for entry in entries[:5])
        if len(entries) > 5:
            sample += f", +{len(entries) - 5} more"
        self.palette_preview_label.setText(
            f"{palette.get('description', '')}\nEntries: {len(entries)}\n{sample}".strip()
        )


    def current_palette_mapping(self) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        palette = self.selected_palette()
        if not palette:
            return rows
        target_key = "dark" if self.palette_target_combo.currentText().lower().startswith("dark") else "light"
        for row in range(self.palette_mapping_list.count()):
            item = self.palette_mapping_list.item(row)
            data = item.data(Qt.UserRole)
            combo = self.palette_mapping_widgets.get(id(item))
            if not data or combo is None:
                continue
            role_name = combo.currentText()
            if role_name == "Skip":
                continue
            entry = palette_entry_by_name(palette, role_name)
            if not entry:
                continue
            rows.append({
                "scope": data["scope"],
                "rgb": tuple(data["rgb"]),
                "role": role_name,
                "target_rgb": tuple(entry[target_key]),
                "count": data["count"],
            })
        return rows

    def visible_colour_groups_for_mapping(self) -> list[dict[str, Any]]:
        point_groups: dict[tuple[int, int, int], dict[str, Any]] = {}
        line_groups: dict[tuple[int, int, int], dict[str, Any]] = {}

        for point in self.main_window.loaded_map.points:
            if getattr(point, "deleted", False) or not self.main_window.layer_visible.get(point.file_path, True):
                continue
            rgb = (point.r, point.g, point.b)
            group = point_groups.setdefault(rgb, {"scope": "Points", "rgb": rgb, "count": 0, "labels": []})
            group["count"] += 1
            if point.label:
                group["labels"].append(point.label)

        for line in self.main_window.loaded_map.lines:
            if getattr(line, "deleted", False) or not self.main_window.layer_visible.get(line.file_path, True):
                continue
            rgb = (line.r, line.g, line.b)
            group = line_groups.setdefault(rgb, {"scope": "Lines", "rgb": rgb, "count": 0, "labels": []})
            group["count"] += 1

        groups = list(point_groups.values()) + list(line_groups.values())
        groups.sort(key=lambda group: (group["scope"], -group["count"], group["rgb"]))
        return groups

    def best_role_for_group(self, group: dict[str, Any], palette: dict[str, Any]) -> str:
        inferred = None
        if group["scope"] == "Points":
            inferred = infer_point_role_from_labels(group.get("labels", []), palette)
        if inferred:
            return inferred

        entries = palette_entries(palette)
        if not entries:
            return "Skip"
        rgb = tuple(group["rgb"])
        best_entry = min(
            entries,
            key=lambda entry: min(rgb_distance_sq(rgb, entry["light"]), rgb_distance_sq(rgb, entry["dark"])),
        )
        return best_entry["name"]

    def rebuild_palette_mapping_preview(self) -> None:
        if not hasattr(self, "palette_mapping_list"):
            return
        palette = self.selected_palette()
        self.palette_mapping_list.clear()
        self.palette_mapping_widgets.clear()
        if not palette:
            return

        role_names = ["Skip"] + palette_role_names(palette)
        groups = self.visible_colour_groups_for_mapping()
        if not groups:
            self.palette_mapping_list.addItem(QListWidgetItem("No visible colours to map."))
            return

        for group in groups:
            rgb = tuple(group["rgb"])
            item = QListWidgetItem()
            item.setData(Qt.UserRole, group)
            self.palette_mapping_list.addItem(item)

            row_widget = QWidget()
            row_layout = QHBoxLayout(row_widget)
            row_layout.setContentsMargins(4, 2, 4, 2)
            row_layout.setSpacing(8)

            swatch = QLabel()
            swatch.setFixedSize(22, 22)
            swatch.setStyleSheet(f"background-color: rgb({rgb[0]}, {rgb[1]}, {rgb[2]}); border: 1px solid #777;")

            sample_labels = ""
            if group["scope"] == "Points" and group.get("labels"):
                sample = ", ".join(group["labels"][:3])
                sample_labels = f" — {sample}"

            label = QLabel(f'{group["scope"]} RGB {rgb}   Count: {group["count"]}{sample_labels}')
            label.setWordWrap(True)

            combo = QComboBox()
            combo.addItems(role_names)
            combo.setCurrentText(self.best_role_for_group(group, palette))
            combo.setMinimumWidth(160)

            row_layout.addWidget(swatch)
            row_layout.addWidget(label, 1)
            row_layout.addWidget(combo)

            item.setSizeHint(row_widget.sizeHint())
            self.palette_mapping_list.setItemWidget(item, row_widget)
            self.palette_mapping_widgets[id(item)] = combo

    def auto_map_palette_preview(self) -> None:
        palette = self.selected_palette()
        if not palette:
            return
        for row in range(self.palette_mapping_list.count()):
            item = self.palette_mapping_list.item(row)
            group = item.data(Qt.UserRole)
            combo = self.palette_mapping_widgets.get(id(item))
            if not group or combo is None:
                continue
            combo.setCurrentText(self.best_role_for_group(group, palette))

    def edit_save_palette(self) -> None:
        palette = self.selected_palette() or DEFAULT_MAP_PALETTES["EQ Map Standard"]
        dialog = PaletteEditorDialog(self, palette)
        if dialog.exec() == QDialog.Accepted and dialog.saved_palette:
            path = save_user_palette(dialog.saved_palette)
            self.refresh_palette_combo()
            index = self.palette_combo.findText(dialog.saved_palette["name"])
            if index >= 0:
                self.palette_combo.setCurrentIndex(index)
            self.main_window.status_label.setText(f"Saved palette: {path.name}")

    def open_palettes_folder(self) -> None:
        try:
            PALETTES_DIR.mkdir(parents=True, exist_ok=True)
            os.startfile(str(PALETTES_DIR))
        except Exception as exc:
            QMessageBox.warning(self, "Palettes Folder", f"Could not open palettes folder:\n{exc}")

    def pick_bulk_colour(self) -> None:
        c = choose_colour_dialog(QColor(self.bulk_new_r.value(), self.bulk_new_g.value(), self.bulk_new_b.value()), self, "Choose new colour")
        if c.isValid():
            self.bulk_new_r.setValue(c.red())
            self.bulk_new_g.setValue(c.green())
            self.bulk_new_b.setValue(c.blue())

    def match_bulk_colour_from_list(self) -> None:
        colours = self.main_window.all_colour_counts()
        dialog = ColorChoiceDialog(colours, self)
        if dialog.exec() == QDialog.Accepted and dialog.selected_rgb is not None:
            r, g, b = dialog.selected_rgb
            self.bulk_new_r.setValue(r)
            self.bulk_new_g.setValue(g)
            self.bulk_new_b.setValue(b)

    def pick_point_bulk_colour(self) -> None:
        c = choose_colour_dialog(QColor(self.point_bulk_r.value(), self.point_bulk_g.value(), self.point_bulk_b.value()), self, "Choose new point colour")
        if c.isValid():
            self.point_bulk_r.setValue(c.red())
            self.point_bulk_g.setValue(c.green())
            self.point_bulk_b.setValue(c.blue())

    def rebuild_colour_list(self) -> None:
        self.point_colour_list.clear()
        self.line_colour_list.clear()

        point_counts: dict[tuple[int, int, int], int] = {}
        line_counts: dict[tuple[int, int, int], int] = {}

        for point in self.main_window.loaded_map.points:
            if getattr(point, "deleted", False) or not self.main_window.layer_visible.get(point.file_path, True):
                continue
            rgb = (point.r, point.g, point.b)
            point_counts[rgb] = point_counts.get(rgb, 0) + 1

        for line in self.main_window.loaded_map.lines:
            if getattr(line, "deleted", False) or not self.main_window.layer_visible.get(line.file_path, True):
                continue
            rgb = (line.r, line.g, line.b)
            line_counts[rgb] = line_counts.get(rgb, 0) + 1

        for rgb, count in sorted(point_counts.items(), key=lambda item: item[1], reverse=True):
            item = QListWidgetItem(f"RGB {rgb}    Points: {count}")
            item.setIcon(self.colour_icon(rgb))
            item.setData(Qt.UserRole, rgb)
            item.setForeground(self.list_text_colour())
            self.point_colour_list.addItem(item)

        for rgb, count in sorted(line_counts.items(), key=lambda item: item[1], reverse=True):
            item = QListWidgetItem(f"RGB {rgb}    Lines: {count}")
            item.setIcon(self.colour_icon(rgb))
            item.setData(Qt.UserRole, rgb)
            item.setForeground(self.list_text_colour())
            self.line_colour_list.addItem(item)

        if hasattr(self, "palette_mapping_list"):
            self.rebuild_palette_mapping_preview()

    def reset_point_search(self) -> None:
        self.point_search_edit.clear()
        self.rebuild_points_list()

    def rebuild_points_list(self) -> None:
        search = self.point_search_edit.text().strip().lower()
        self.points_list.clear()
        for point in self.main_window.loaded_map.points:
            if getattr(point, "deleted", False) or not self.main_window.layer_visible.get(point.file_path, True):
                continue
            if search and search not in point.label.lower():
                continue
            item = QListWidgetItem(f"{point.label}    RGB ({point.r}, {point.g}, {point.b})    {point.file_path.name}:{point.line_index + 1 if point.line_index >= 0 else 'new'}")
            item.setData(Qt.UserRole, id(point))
            item.setIcon(self.colour_icon((point.r, point.g, point.b)))
            item.setForeground(self.list_text_colour())
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            self.points_list.addItem(item)

    def check_matching_points(self) -> None:
        self.rebuild_points_list()
        for index in range(self.points_list.count()):
            self.points_list.item(index).setCheckState(Qt.Checked)

    def checked_point_records(self) -> list[MapPointRecord]:
        ids = set()
        for index in range(self.points_list.count()):
            item = self.points_list.item(index)
            if item.checkState() == Qt.Checked:
                ids.add(item.data(Qt.UserRole))
        return [point for point in self.main_window.loaded_map.points if id(point) in ids and not getattr(point, "deleted", False)]



class OptionOneExplorerPanel(QWidget):
    """Option 1 style left explorer panel with a navigation rail."""
    PAGE_ORDER = ["Zones", "Layers", "Bulk Colours", "Points", "Pending Changes"]

    def __init__(self, main_window: "EqMapMainWindow") -> None:
        super().__init__(main_window)
        self.main_window = main_window
        self.setMinimumWidth(310)
        self.setMaximumWidth(420)
        self.nav_buttons: dict[str, QToolButton] = {}
        self._layers_checkbox_widgets: list[QCheckBox] = []
        self._build_ui()

    def _build_ui(self) -> None:
        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        self.rail_widget = QWidget()
        self.rail_widget.setObjectName("explorerRail")
        self.rail_widget.setFixedWidth(84)
        rail_layout = QVBoxLayout(self.rail_widget)
        rail_layout.setContentsMargins(10, 14, 10, 14)
        rail_layout.setSpacing(8)

        self.explorer_logo = QLabel("Explorer")
        self.explorer_logo.setObjectName("explorerRailTitle")
        self.explorer_logo.setWordWrap(True)
        rail_layout.addWidget(self.explorer_logo)

        button_specs = [
            ("Zones", "🗺\nZones"),
            ("Layers", "▤\nLayers"),
            ("Bulk Colours", "◔\nBulk\nColours"),
            ("Points", "⌖\nPoints"),
            ("Pending Changes", "↺\nPending"),
        ]
        for page_name, label in button_specs:
            button = QToolButton()
            button.setObjectName("explorerNavButton")
            button.setText(label)
            button.setCheckable(True)
            button.setToolButtonStyle(Qt.ToolButtonTextOnly)
            button.clicked.connect(lambda checked=False, name=page_name: self.set_page(name))
            rail_layout.addWidget(button)
            self.nav_buttons[page_name] = button

        rail_layout.addStretch(1)
        root.addWidget(self.rail_widget)

        self.content_widget = QWidget()
        self.content_widget.setObjectName("explorerContent")
        content_layout = QVBoxLayout(self.content_widget)
        content_layout.setContentsMargins(12, 12, 12, 12)
        content_layout.setSpacing(10)

        self.section_title = QLabel("Zones")
        self.section_title.setObjectName("explorerSectionTitle")
        content_layout.addWidget(self.section_title)

        self.page_stack = QStackedWidget()
        self.page_stack.setObjectName("explorerPageStack")
        content_layout.addWidget(self.page_stack, 1)

        # Zones page
        self.zones_page = QWidget()
        zones_layout = QVBoxLayout(self.zones_page)
        zones_layout.setContentsMargins(0, 0, 0, 0)
        zones_layout.setSpacing(8)

        self.zone_quick_search = QLineEdit()
        self.zone_quick_search.setPlaceholderText("Search zones...")
        zones_layout.addWidget(self.zone_quick_search)

        self.zone_preview_list = QListWidget()
        zones_layout.addWidget(self.zone_preview_list, 1)

        self.open_zones_button = QPushButton("Open Zone Browser")
        zones_layout.addWidget(self.open_zones_button)

        self.current_zone_label = QLabel("No zone loaded")
        self.current_zone_label.setWordWrap(True)
        self.current_zone_label.setObjectName("explorerSubtleText")
        zones_layout.addWidget(self.current_zone_label)
        self.page_stack.addWidget(self.zones_page)

        # Layers page
        self.layers_page = QWidget()
        layers_layout = QVBoxLayout(self.layers_page)
        layers_layout.setContentsMargins(0, 0, 0, 0)
        layers_layout.setSpacing(8)

        self.layers_intro = QLabel("Toggle which loaded files are visible.")
        self.layers_intro.setWordWrap(True)
        self.layers_intro.setObjectName("explorerSubtleText")
        layers_layout.addWidget(self.layers_intro)

        self.layers_summary = QLabel("No files loaded")
        self.layers_summary.setWordWrap(True)
        self.layers_summary.setObjectName("explorerSummaryText")
        layers_layout.addWidget(self.layers_summary)

        self.layers_container = QWidget()
        self.layers_container_layout = QVBoxLayout(self.layers_container)
        self.layers_container_layout.setContentsMargins(0, 0, 0, 0)
        self.layers_container_layout.setSpacing(6)
        layers_layout.addWidget(self.layers_container)
        layers_layout.addStretch(1)
        self.page_stack.addWidget(self.layers_page)

        # Bulk colours page
        self.bulk_page = QWidget()
        bulk_layout = QVBoxLayout(self.bulk_page)
        bulk_layout.setContentsMargins(0, 0, 0, 0)
        bulk_layout.setSpacing(8)

        self.colour_swatches = QLabel("No colours loaded")
        self.colour_swatches.setWordWrap(True)
        self.colour_swatches.setObjectName("explorerSwatches")
        bulk_layout.addWidget(self.colour_swatches)

        self.bulk_summary = QLabel("Visible colour groups: 0")
        self.bulk_summary.setObjectName("explorerSummaryText")
        bulk_layout.addWidget(self.bulk_summary)

        self.open_bulk_button = QPushButton("Open Bulk Colour Tool")
        bulk_layout.addWidget(self.open_bulk_button)
        bulk_layout.addStretch(1)
        self.page_stack.addWidget(self.bulk_page)

        # Points page
        self.points_page = QWidget()
        points_layout = QVBoxLayout(self.points_page)
        points_layout.setContentsMargins(0, 0, 0, 0)
        points_layout.setSpacing(8)

        self.points_summary = QLabel("Points: 0\nLines: 0")
        self.points_summary.setObjectName("explorerSummaryText")
        self.points_summary.setWordWrap(True)
        points_layout.addWidget(self.points_summary)

        self.points_hint = QLabel("Search and bulk-edit points from the Points tool.")
        self.points_hint.setObjectName("explorerSubtleText")
        self.points_hint.setWordWrap(True)
        points_layout.addWidget(self.points_hint)

        self.open_points_button = QPushButton("Open Points Tool")
        points_layout.addWidget(self.open_points_button)
        points_layout.addStretch(1)
        self.page_stack.addWidget(self.points_page)

        # Pending page
        self.pending_page = QWidget()
        pending_layout = QVBoxLayout(self.pending_page)
        pending_layout.setContentsMargins(0, 0, 0, 0)
        pending_layout.setSpacing(8)

        self.pending_summary = QLabel("Pending: 0")
        self.pending_summary.setObjectName("explorerSummaryText")
        pending_layout.addWidget(self.pending_summary)

        self.pending_hint = QLabel("Review unsaved changes before saving.")
        self.pending_hint.setObjectName("explorerSubtleText")
        self.pending_hint.setWordWrap(True)
        pending_layout.addWidget(self.pending_hint)

        self.open_pending_button = QPushButton("Review Pending Changes")
        pending_layout.addWidget(self.open_pending_button)
        pending_layout.addStretch(1)
        self.page_stack.addWidget(self.pending_page)

        root.addWidget(self.content_widget, 1)

        self.open_zones_button.clicked.connect(lambda: self.main_window.show_side_tool("Zones"))
        self.open_bulk_button.clicked.connect(lambda: self.main_window.show_side_tool("Bulk Colours"))
        self.open_points_button.clicked.connect(lambda: self.main_window.show_side_tool("Points"))
        self.open_pending_button.clicked.connect(lambda: self.main_window.show_side_tool("Pending Changes"))
        self.zone_quick_search.textChanged.connect(self.rebuild_zone_preview)
        self.zone_preview_list.itemDoubleClicked.connect(self.open_selected_preview_zone)

        self.set_page("Zones")

    def clear_layout(self, layout: QVBoxLayout) -> None:
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

    def set_page(self, page_name: str) -> None:
        if page_name not in self.PAGE_ORDER:
            return
        for name, button in self.nav_buttons.items():
            button.setChecked(name == page_name)
        self.section_title.setText(page_name)
        self.page_stack.setCurrentIndex(self.PAGE_ORDER.index(page_name))

    def available_zone_names(self) -> list[tuple[str, str]]:
        names = sorted(ZONE_SHORTNAME_TO_FULLNAME.items(), key=lambda item: item[1].lower())
        if names:
            return names
        # Fallback sample list
        return [
            ("qeynos", "South Qeynos"),
            ("qey2hh1", "North Qeynos"),
            ("qeytoqrg", "Qeynos Hills"),
            ("freportw", "West Freeport"),
            ("commons", "West Commonlands"),
            ("ecommons", "East Commonlands"),
            ("nro", "North Ro"),
            ("misty", "Misty Thicket"),
        ]

    def rebuild_zone_preview(self) -> None:
        search = self.zone_quick_search.text().strip().lower()
        self.zone_preview_list.clear()
        count = 0
        for shortname, fullname in self.available_zone_names():
            label = f"{fullname} ({shortname})"
            if search and search not in label.lower():
                continue
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, shortname)
            self.zone_preview_list.addItem(item)
            count += 1
            if count >= 20:
                break
        if count == 0:
            self.zone_preview_list.addItem(QListWidgetItem("No matching zones"))

    def open_selected_preview_zone(self) -> None:
        item = self.zone_preview_list.currentItem()
        if item is None:
            return
        shortname = item.data(Qt.UserRole)
        if not shortname:
            return
        if hasattr(self.main_window.side_panel, "zone_search_edit"):
            self.main_window.side_panel.zone_search_edit.setText(shortname)
            self.main_window.rebuild_zone_list()
            for row in range(self.main_window.side_panel.zones_list.count()):
                zone_item = self.main_window.side_panel.zones_list.item(row)
                if zone_item.data(Qt.UserRole) == shortname:
                    self.main_window.side_panel.zones_list.setCurrentItem(zone_item)
                    break
        self.main_window.show_side_tool("Zones")

    def refresh(self) -> None:
        loaded = self.main_window.loaded_files
        if loaded:
            short = loaded[0].stem.split("_", 1)[0]
            display = self.main_window.zone_display_name(short) if hasattr(self.main_window, "zone_display_name") else short
            self.current_zone_label.setText(display)
        else:
            self.current_zone_label.setText("No zone loaded")

        self.rebuild_zone_preview()

        self.clear_layout(self.layers_container_layout)
        if not loaded:
            self.layers_summary.setText("No files loaded")
            self.layers_container_layout.addWidget(QLabel("Open map files to show layer visibility."))
        else:
            total_points = sum(1 for p in self.main_window.loaded_map.points if not getattr(p, "deleted", False))
            total_lines = sum(1 for l in self.main_window.loaded_map.lines if not getattr(l, "deleted", False))
            self.layers_summary.setText(f"{len(loaded)} loaded file(s)\n{total_points} point(s), {total_lines} line(s)")
            for file_path in loaded:
                row = QCheckBox(file_path.name)
                row.setChecked(self.main_window.layer_visible.get(file_path, True))
                row.toggled.connect(lambda checked, fp=file_path: self.main_window.set_layer_visible(fp, checked))
                self.layers_container_layout.addWidget(row)
        self.layers_container_layout.addStretch(1)

        point_count = sum(1 for p in self.main_window.loaded_map.points if not getattr(p, "deleted", False))
        line_count = sum(1 for l in self.main_window.loaded_map.lines if not getattr(l, "deleted", False))
        self.points_summary.setText(f"Points: {point_count:,}\nLines: {line_count:,}")

        dirty_count = len(self.main_window.dirty_records()) if hasattr(self.main_window, "dirty_records") else 0
        self.pending_summary.setText(f"Pending: {dirty_count:,}")

        colours = {}
        for point in self.main_window.loaded_map.points:
            if getattr(point, "deleted", False):
                continue
            rgb = (point.r, point.g, point.b)
            colours[rgb] = colours.get(rgb, 0) + 1
        for line in self.main_window.loaded_map.lines:
            if getattr(line, "deleted", False):
                continue
            rgb = (line.r, line.g, line.b)
            colours[rgb] = colours.get(rgb, 0) + 1

        if colours:
            top_colours = sorted(colours.items(), key=lambda item: item[1], reverse=True)[:12]
            swatches = " ".join(
                f'<span style="background-color: rgb({r},{g},{b}); color: rgb({r},{g},{b}); border: 1px solid #777;">■■</span>'
                for (r, g, b), _count in top_colours
            )
            self.colour_swatches.setText(swatches)
            self.bulk_summary.setText(f"Visible colour groups: {len(colours):,}")
        else:
            self.colour_swatches.setText("No colours loaded")
            self.bulk_summary.setText("Visible colour groups: 0")


class EqMapMainWindow(QMainWindow):
    def __init__(self, initial_files: Optional[list[Path]] = None) -> None:
        super().__init__()
        self.setWindowTitle(f"EQ Map Editor")
        icon = app_icon()
        if not icon.isNull():
            self.setWindowIcon(icon)
        self.mapper = CoordinateMapper(flip_display_y=False)

        self.scene = QGraphicsScene(self)
        self.view = EqMapView(self)
        self.view.setScene(self.scene)

        self.loaded_files: list[Path] = []
        self.loaded_file_mtimes: dict[Path, float] = {}
        self.loaded_map = LoadedMap(lines=[], points=[])
        self.layer_visible: dict[Path, bool] = {}

        self.line_items_by_record: dict[int, MovableLineItem] = {}
        self.point_items_by_record: dict[int, MovablePointMarker] = {}
        self.label_items_by_record: dict[int, QGraphicsTextItem] = {}
        self.endpoint_handles: list[EndpointHandle] = []
        self.selection_highlights: list[QGraphicsItem] = []
        self.bulk_selected_records: list[Any] = []
        self.search_matches: list[MapPointRecord] = []
        self.search_index: int = -1

        self.show_labels = True
        self.show_points = True
        self.dark_ui = False
        self.current_bulk_colour_type = "points"
        self.map_folder = ""
        self.remember_last_loaded_map = True
        self.auto_fit_restored_map = True
        self.autosave_settings_on_exit = True
        self.show_beta_warning = False
        self.welcome_seen = False
        self.default_background_mode = "remember"
        self.max_highlights_before_warning = 5000
        self.confirm_bulk_edit_over = 1
        self.confirm_bulk_delete_over = 1
        self.confirm_bulk_actions = True
        self._fit_last_map_after_show = False
        self._drag_update_lock = False
        self.pending_line_start: Optional[QPointF] = None
        self.npc_data_source_path = ""
        self.current_expansion_number_setting = 0
        self.npc_data_rows: list[NpcDataRow] = []
        self.npc_match_results: list[NpcMatchResult] = []
        self.era_cleanup_results: list[EraCleanupResult] = []
        self.missing_npc_results: list[MissingNpcResult] = []
        self.npc_preview_items: list[QGraphicsItem] = []
        self.era_preview_items: list[QGraphicsItem] = []
        self.missing_preview_items: list[QGraphicsItem] = []

        self.undo_stack: list[UndoCommand] = []
        self.redo_stack: list[UndoCommand] = []

        self.side_panel = SidePanel(self)

        self._build_ui()

        if initial_files:
            self.load_files(initial_files)
        else:
            self.load_last_session()

    def _build_ui(self) -> None:
        toolbar = QToolBar("Main Toolbar", self)
        toolbar.setObjectName("topToolbar")
        toolbar.setMovable(False)
        toolbar.setFloatable(False)
        toolbar.setIconSize(QSize(16, 16))
        toolbar.setToolButtonStyle(Qt.ToolButtonTextOnly)
        toolbar.setFixedHeight(64)
        self.addToolBar(toolbar)
        self.top_toolbar = toolbar

        file_menu = QMenu("File", self)
        file_menu.setMinimumWidth(190)

        file_button = QToolButton(self)
        file_button.setObjectName("fileMenuButton")
        file_button.setText("☰  File")
        file_button.setMenu(file_menu)
        file_button.setPopupMode(QToolButton.InstantPopup)
        file_button.setToolButtonStyle(Qt.ToolButtonTextOnly)
        toolbar.addWidget(file_button)
        self.file_button = file_button

        open_action = QAction("Open Map File(s)", self)
        open_action.setShortcut("Ctrl+O")
        open_action.triggered.connect(self.open_files_dialog)
        file_menu.addAction(open_action)

        save_action = QAction("Save Edits", self)
        save_action.setShortcut("Ctrl+S")
        save_action.triggered.connect(self.save_edits)
        file_menu.addAction(save_action)

        save_as_action = QAction("Save As...", self)
        save_as_action.triggered.connect(self.save_as_folder)
        file_menu.addAction(save_as_action)

        file_menu.addSeparator()

        revert_action = QAction("Revert Unsaved", self)
        revert_action.triggered.connect(self.reload_from_disk)
        file_menu.addAction(revert_action)

        restore_backup_action = QAction("Restore Backup...", self)
        restore_backup_action.triggered.connect(self.restore_from_backup)
        file_menu.addAction(restore_backup_action)

        open_backup_action = QAction("Open Backup Folder", self)
        open_backup_action.triggered.connect(self.open_backup_folder)
        file_menu.addAction(open_backup_action)

        file_menu.addSeparator()

        preferences_action = QAction("Preferences", self)
        preferences_action.triggered.connect(self.open_preferences)
        file_menu.addAction(preferences_action)

        help_action = QAction("Controls", self)
        help_action.triggered.connect(self.show_controls)
        file_menu.addAction(help_action)

        keyboard_action = QAction("Keyboard Shortcuts", self)
        keyboard_action.triggered.connect(self.show_keyboard_shortcuts)
        file_menu.addAction(keyboard_action)

        quick_start_action = QAction("Quick Start", self)
        quick_start_action.triggered.connect(self.show_quick_start)
        file_menu.addAction(quick_start_action)

        open_logs_action = QAction("Open Logs Folder", self)
        open_logs_action.triggered.connect(self.open_logs_folder)
        file_menu.addAction(open_logs_action)

        about_action = QAction("About", self)
        about_action.triggered.connect(self.show_about)
        file_menu.addAction(about_action)

        file_menu.addSeparator()

        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.confirm_exit)
        file_menu.addAction(exit_action)

        toolbar.addSeparator()

        fit_action = QAction("Fit Map", self)
        self.fit_action = fit_action
        fit_action.setShortcut("Ctrl+F")
        fit_action.triggered.connect(self.fit_map)
        toolbar.addAction(fit_action)

        fit_selected_action = QAction("Fit Selected", self)
        self.fit_selected_action = fit_selected_action
        fit_selected_action.triggered.connect(self.fit_selected)
        toolbar.addAction(fit_selected_action)

        clear_selection_action = QAction("Clear Selection", self)
        self.clear_selection_action = clear_selection_action
        clear_selection_action.triggered.connect(self.clear_selection_and_highlights)
        toolbar.addAction(clear_selection_action)

        self.toggle_labels_action = QAction("Show Labels", self)
        self.toggle_labels_action.setCheckable(True)
        self.toggle_labels_action.setChecked(True)
        self.toggle_labels_action.triggered.connect(self.toggle_labels)
        toolbar.addAction(self.toggle_labels_action)

        self.toggle_points_action = QAction("Show Points", self)
        self.toggle_points_action.setCheckable(True)
        self.toggle_points_action.setChecked(True)
        self.toggle_points_action.triggered.connect(self.toggle_points)
        toolbar.addAction(self.toggle_points_action)

        self.background_action_group = QActionGroup(self)
        self.background_action_group.setExclusive(True)

        self.light_bg_action = QAction("Light Background", self)
        self.light_bg_action.setCheckable(True)
        self.light_bg_action.setChecked(True)
        self.light_bg_action.triggered.connect(lambda: self.set_background("light"))
        self.background_action_group.addAction(self.light_bg_action)
        toolbar.addAction(self.light_bg_action)

        self.dark_bg_action = QAction("Dark Background", self)
        self.dark_bg_action.setCheckable(True)
        self.dark_bg_action.triggered.connect(lambda: self.set_background("dark"))
        self.background_action_group.addAction(self.dark_bg_action)
        toolbar.addAction(self.dark_bg_action)

        toggle_sidebar_action = QAction("Toggle Sidebar", self)
        self.toggle_sidebar_action = toggle_sidebar_action
        toggle_sidebar_action.triggered.connect(self.toggle_sidebar)
        toolbar.addAction(toggle_sidebar_action)

        toolbar.addSeparator()
        self.search_label = QLabel("Search Labels")
        self.search_label.setObjectName("toolbarSearchLabel")
        toolbar.addWidget(self.search_label)
        self.global_search_edit = QLineEdit()
        self.global_search_edit.setObjectName("toolbarSearchEdit")
        self.global_search_edit.setPlaceholderText("Point label text...")
        self.global_search_edit.setFixedWidth(220)
        self.global_search_edit.setFixedHeight(38)
        self.global_search_edit.returnPressed.connect(self.search_select_first_label_match)
        toolbar.addWidget(self.global_search_edit)

        search_first_action = QAction("Find First", self)
        self.search_first_action = search_first_action
        search_first_action.triggered.connect(self.search_select_first_label_match)
        toolbar.addAction(search_first_action)

        search_all_action = QAction("Select Matches", self)
        self.search_all_action = search_all_action
        search_all_action.triggered.connect(self.search_select_all_label_matches)
        toolbar.addAction(search_all_action)

        center_selected_action = QAction("Center Selected", self)
        self.center_selected_action = center_selected_action
        center_selected_action.triggered.connect(self.center_selected)
        toolbar.addAction(center_selected_action)

        self.setup_shortcuts(
            open_action,
            save_action,
            self.undo,
            self.redo,
            self.fit_map,
            self.clear_selection_and_highlights,
            self.delete_selected_records,
            self.select_all_visible_records,
            self.global_search_edit,
        )

        self.canvas_container = MapCanvasContainer(self)
        canvas_layout = QVBoxLayout(self.canvas_container)
        canvas_layout.setContentsMargins(0, 0, 0, 0)
        canvas_layout.setSpacing(0)
        canvas_layout.addWidget(self.view)

        self.canvas_controls = CanvasControlsOverlay(self)
        self.canvas_controls.setParent(self.canvas_container)
        self.minimap_overlay = MiniMapOverlay(self)
        self.minimap_overlay.setParent(self.canvas_container)

        self.splitter = QSplitter(Qt.Horizontal, self)
        self.splitter.addWidget(self.canvas_container)
        self.splitter.addWidget(self.side_panel)
        self.splitter.setStretchFactor(0, 5)
        self.splitter.setStretchFactor(1, 1)
        self.splitter.setSizes([1260, 440])
        self.setCentralWidget(self.splitter)

        self.view.horizontalScrollBar().valueChanged.connect(self.update_canvas_overlays)
        self.view.verticalScrollBar().valueChanged.connect(self.update_canvas_overlays)

        self.status_label = QLabel("Open one or more EQ map text files.")
        self.dirty_label = QLabel("Clean")
        status = QStatusBar(self)
        status.addWidget(self.status_label)
        status.addPermanentWidget(self.dirty_label)
        self.setStatusBar(status)

        self.set_background("dark")

    def edit_mode(self) -> str:
        return self.side_panel.edit_mode()

    def show_controls(self) -> None:
        QMessageBox.information(
            self,
            "Controls",
            (
                f"EQ Map Editor {VERSION} controls:\n\n"
                "- Mouse wheel: zoom in/out\n"
                "- Middle/right mouse drag: pan\n"
                "- Select Only: inspect records without moving them\n"
                "- Move Points: drag point markers\n"
                "- Move Lines: drag whole line segments\n"
                "- Move Line Endpoints: select a line, then drag yellow endpoint handles\n"
                "- Add New Point: double-click the map to create a P record\n"
                "- Add New Line: double-click once for endpoint 1, again for endpoint 2\n"
                "- Side panel checkboxes toggle file/layer visibility\n"
                "- Save Edits: writes dirty L/P records to source files with timestamped backups\n"
                "- Flip Display Y: flips display only; saved coordinate meaning does not change"
            ),
        )

    def set_background(self, mode: str) -> None:
        self.dark_ui = mode == "dark"
        canvas_color = QColor(18, 18, 18) if self.dark_ui else QColor(245, 246, 248)
        self.view.setBackgroundBrush(QBrush(canvas_color))
        self.scene.setBackgroundBrush(QBrush(canvas_color))
        if hasattr(self, "light_bg_action"):
            self.light_bg_action.setChecked(mode == "light")
        if hasattr(self, "dark_bg_action"):
            self.dark_bg_action.setChecked(mode == "dark")

        up_arrow, down_arrow = ensure_spinbox_arrow_images()

        if self.dark_ui:
            window_bg = "#242424"
            panel_bg = "#282828"
            field_bg = "#1b1b1b"
            text_color = "#f0f0f0"
            border = "#555"
            strong_border = "#6f6f6f"
            button_bg = "#1f1f1f"
            button_hover = "#2b2b2b"
            button_checked = "#3f5f91"
            toolbar_bg = "#1d1f22"
            toolbar_border = "#36393d"
            toolbar_button = "#1f1f21"
            toolbar_button_hover = "#2d3036"
            toolbar_button_checked = "#3b5c91"
            toolbar_text = "#f0f0f0"
            search_bg = "#17181b"
        else:
            window_bg = "#f4f5f7"
            panel_bg = "#fbfbfc"
            field_bg = "#ffffff"
            text_color = "#1b1d20"
            border = "#c8ccd2"
            strong_border = "#b8bec8"
            button_bg = "#ffffff"
            button_hover = "#eef2f8"
            button_checked = "#dbe7fb"
            toolbar_bg = "#eef1f5"
            toolbar_border = "#c8ced8"
            toolbar_button = "#ffffff"
            toolbar_button_hover = "#f3f6fb"
            toolbar_button_checked = "#dbe7fb"
            toolbar_text = "#1f2630"
            search_bg = "#ffffff"

        self.setStyleSheet(f"""
            QWidget {{
                background-color: {window_bg};
                color: {text_color};
            }}
            QGroupBox {{
                border: 1px solid {border};
                border-radius: 8px;
                margin-top: 10px;
                padding: 10px;
                background-color: {panel_bg};
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 4px;
                color: {text_color};
                font-weight: bold;
            }}
            QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox, QListWidget {{
                background-color: {field_bg};
                color: {text_color};
                border: 1px solid {border};
                border-radius: 4px;
            }}
            QSpinBox, QDoubleSpinBox {{
                padding-right: 22px;
                min-height: 22px;
            }}
            QSpinBox::up-button, QDoubleSpinBox::up-button {{
                subcontrol-origin: border;
                subcontrol-position: top right;
                width: 20px;
                border-left: 1px solid {strong_border};
                border-bottom: 1px solid {border};
                background-color: {button_bg};
            }}
            QSpinBox::down-button, QDoubleSpinBox::down-button {{
                subcontrol-origin: border;
                subcontrol-position: bottom right;
                width: 20px;
                border-left: 1px solid {strong_border};
                border-top: 1px solid {border};
                background-color: {button_bg};
            }}
            QSpinBox::up-button:hover, QDoubleSpinBox::up-button:hover,
            QSpinBox::down-button:hover, QDoubleSpinBox::down-button:hover {{
                background-color: {button_hover};
            }}
            QSpinBox::up-arrow, QDoubleSpinBox::up-arrow {{
                image: url("{up_arrow}");
                width: 9px;
                height: 9px;
            }}
            QSpinBox::down-arrow, QDoubleSpinBox::down-arrow {{
                image: url("{down_arrow}");
                width: 9px;
                height: 9px;
            }}
            QPushButton {{
                background-color: {button_bg};
                color: {text_color};
                border: 1px solid {border};
                border-radius: 4px;
                padding: 4px 8px;
            }}
            QPushButton:hover {{
                background-color: {button_hover};
            }}
            QTabWidget::pane {{
                border: 1px solid {border};
                background-color: {panel_bg};
            }}
            QTabBar::tab {{
                background: {button_bg};
                color: {text_color};
                padding: 7px 10px;
                border: 1px solid {border};
                border-bottom: none;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
                margin-right: 2px;
            }}
            QTabBar::tab:selected {{
                background: {button_checked};
                font-weight: bold;
            }}
            QMenu {{
                background-color: {panel_bg};
                color: {text_color};
                border: 1px solid {border};
                padding: 4px;
            }}
            QMenu::item {{
                padding: 6px 24px 6px 14px;
                border-radius: 4px;
            }}
            QMenu::item:selected {{
                background-color: {button_checked};
            }}
            QToolBar#topToolbar {{
                background-color: {toolbar_bg};
                border: none;
                border-bottom: 1px solid {toolbar_border};
                spacing: 6px;
                padding: 8px 8px;
            }}
            QToolBar#topToolbar::separator {{
                background: {toolbar_border};
                width: 1px;
                margin: 5px 6px;
            }}
            QToolButton, QToolBar#topToolbar QToolButton, QToolButton#fileMenuButton {{
                background-color: {toolbar_button};
                color: {toolbar_text};
                border: 1px solid {toolbar_border};
                border-radius: 6px;
                padding: 6px 12px;
                margin: 2px 2px;
                min-height: 36px;
            }}
            QToolButton:hover, QToolBar#topToolbar QToolButton:hover, QToolButton#fileMenuButton:hover {{
                background-color: {toolbar_button_hover};
            }}
            QToolButton:checked, QToolBar#topToolbar QToolButton:checked {{
                background-color: {toolbar_button_checked};
                border: 1px solid #7ba1e6;
                font-weight: bold;
            }}
            QToolButton::menu-indicator {{
                subcontrol-origin: padding;
                subcontrol-position: right center;
                width: 10px;
            }}
            QLabel#toolbarSearchLabel {{
                background: transparent;
                border: none;
                margin-left: 6px;
                margin-right: 2px;
                padding-top: 1px;
                font-weight: 600;
                color: {toolbar_text};
            }}
            QLineEdit#toolbarSearchEdit {{
                background-color: {search_bg};
                color: {toolbar_text};
                border: 1px solid {toolbar_border};
                border-radius: 6px;
                padding: 0px 10px;
                min-height: 38px;
                max-height: 38px;
            }}
            QWidget#explorerRail {{
                background-color: {panel_bg};
                border-right: 1px solid {border};
            }}
            QLabel#explorerRailTitle {{
                color: {text_color};
                font-weight: bold;
                font-size: 13px;
                padding: 4px 2px 8px 2px;
            }}
            QWidget#explorerContent {{
                background-color: {window_bg};
            }}
            QLabel#explorerSectionTitle {{
                font-size: 15px;
                font-weight: bold;
                color: {text_color};
                padding-left: 2px;
            }}
            QToolButton#explorerNavButton {{
                background-color: transparent;
                color: {text_color};
                border: 1px solid transparent;
                border-radius: 8px;
                padding: 8px 4px;
                min-height: 58px;
                text-align: center;
            }}
            QToolButton#explorerNavButton:hover {{
                background-color: {button_hover};
                border: 1px solid {border};
            }}
            QToolButton#explorerNavButton:checked {{
                background-color: {button_checked};
                border: 1px solid #7ba1e6;
                font-weight: bold;
            }}
            QLabel#explorerSubtleText {{
                color: #9aa2ae;
                font-size: 11px;
            }}
            QLabel#explorerSummaryText {{
                color: {text_color};
                font-size: 12px;
                font-weight: 600;
            }}
            QLabel#explorerSwatches {{
                padding: 6px 0;
            }}
            QWidget#canvasControlsOverlay {{
                background-color: rgba(35, 38, 43, 0.94);
                border: 1px solid {border};
                border-radius: 8px;
            }}
            QToolButton#canvasControlButton {{
                background-color: {button_bg};
                color: {text_color};
                border: 1px solid {border};
                border-radius: 5px;
                padding: 0px;
                margin: 0px;
                min-width: 42px;
                max-width: 42px;
                min-height: 38px;
                max-height: 38px;
            }}
            QToolButton#canvasControlButton:hover {{
                background-color: {button_hover};
            }}
            QLabel#canvasOverlayText {{
                background-color: transparent;
                color: {text_color};
                font-weight: 600;
            }}
            QWidget#miniMapOverlay {{
                background-color: rgba(35, 38, 43, 0.90);
                border: 1px solid {border};
                border-radius: 8px;
            }}
            QLabel#inspectorHeaderLabel {{
                font-size: 16px;
                font-weight: bold;
                color: {text_color};
            }}
            QToolButton#inspectorUtilityButton {{
                background-color: transparent;
                color: {text_color};
                border: 1px solid {border};
                border-radius: 4px;
                min-width: 24px;
                min-height: 24px;
                padding: 2px;
            }}
            QToolButton#inspectorUtilityButton:hover {{
                background-color: {button_hover};
            }}
            QLabel#inspectorSummaryTitle {{
                font-weight: bold;
                font-size: 14px;
            }}
            QLabel#inspectorSummaryMeta {{
                color: #9aa2ae;
            }}
            QLabel#inspectorMultiSummary, QLabel#inspectorFooterSummary {{
                color: {text_color};
                font-size: 11px;
            }}
            QLabel#inspectorHelperText {{
                color: #9aa2ae;
                font-size: 10px;
                line-height: 1.3em;
            }}
            QLabel#compactFieldLabel {{
                font-size: 11px;
                color: {text_color};
            }}
            QLabel#colourPreviewSwatch {{
                border: 1px solid {strong_border};
                border-radius: 4px;
                background-color: {field_bg};
            }}
            QPushButton#primaryButton {{
                background-color: #2f63d8;
                color: #ffffff;
                border: 1px solid #4173e3;
                border-radius: 6px;
                font-weight: bold;
                min-height: 34px;
            }}
            QPushButton#primaryButton:hover {{
                background-color: #3d71e6;
            }}
            QPushButton#dangerButton {{
                background-color: transparent;
                color: #ff8c8c;
                border: 1px solid #9b4a4a;
                border-radius: 6px;
                min-height: 34px;
                padding: 4px 10px;
            }}
            QPushButton#dangerButton:hover {{
                background-color: rgba(160, 70, 70, 0.15);
            }}
        """)

        if hasattr(self, "side_panel"):
            self.side_panel.rebuild_colour_list()
            self.side_panel.rebuild_points_list()
            try:
                self.side_panel.rebuild_pending_changes()
            except Exception:
                pass


    def log_event(self, message: str) -> None:
        try:
            LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with LOG_PATH.open("a", encoding="utf-8") as log_file:
                log_file.write(f"[{timestamp}] {message}\n")
        except Exception:
            pass

    def open_log_location(self) -> None:
        try:
            LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
            if not LOG_PATH.exists():
                LOG_PATH.write_text("", encoding="utf-8")
            os.startfile(str(LOG_PATH.parent))
        except Exception as exc:
            QMessageBox.warning(self, "Log Location", f"Could not open log location:\n{exc}")

    def apply_preferences_to_panel(self) -> None:
        return

    def read_preferences_from_panel(self) -> None:
        return

    def confirm_bulk_action(self, title: str, message: str, record_count: int, action_type: str = "edit") -> bool:
        if not self.confirm_bulk_actions:
            return True
        threshold = self.confirm_bulk_delete_over if action_type == "delete" else self.confirm_bulk_edit_over
        if record_count <= 0 or record_count < threshold:
            return True
        result = QMessageBox.question(
            self,
            title,
            message,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        return result == QMessageBox.Yes

    def collect_settings(self) -> dict[str, Any]:
        self.read_preferences_from_panel()
        return {
            "last_files": [str(path) for path in self.loaded_files] if self.remember_last_loaded_map else [],
            "dark_mode": bool(getattr(self, "dark_ui", False)),
            "sidebar_sizes": self.splitter.sizes() if hasattr(self, "splitter") else [],
            "map_folder": getattr(self, "map_folder", ""),
            "remember_last_loaded_map": bool(self.remember_last_loaded_map),
            "auto_fit_restored_map": bool(self.auto_fit_restored_map),
            "autosave_settings_on_exit": bool(self.autosave_settings_on_exit),
            "show_beta_warning": bool(self.show_beta_warning),
            "welcome_seen": bool(getattr(self, "welcome_seen", False)),
            "default_background_mode": self.default_background_mode,
            "flip_display_y": bool(self.mapper.flip_display_y),
            "max_highlights_before_warning": int(self.max_highlights_before_warning),
            "confirm_bulk_edit_over": int(self.confirm_bulk_edit_over),
            "confirm_bulk_delete_over": int(self.confirm_bulk_delete_over),
            "confirm_bulk_actions": bool(self.confirm_bulk_actions),
            "npc_data_source_path": getattr(self, "npc_data_source_path", ""),
            "current_expansion_number": int(getattr(self, "current_expansion_number_setting", 0)),
        }

    def save_settings(self) -> None:
        try:
            SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
            SETTINGS_PATH.write_text(json.dumps(self.collect_settings(), indent=2), encoding="utf-8")
            self.log_event("Settings saved.")
        except Exception as exc:
            self.log_event(f"Settings save failed: {exc}")

    def save_last_session(self) -> None:
        self.save_settings()

    def load_settings_data(self) -> dict[str, Any]:
        try:
            if SETTINGS_PATH.exists():
                return json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
        return {}

    def load_last_session(self) -> None:
        data = self.load_settings_data()
        try:
            self.remember_last_loaded_map = data.get("remember_last_loaded_map", True)
            self.auto_fit_restored_map = data.get("auto_fit_restored_map", True)
            self.autosave_settings_on_exit = data.get("autosave_settings_on_exit", True)
            self.show_beta_warning = data.get("show_beta_warning", True)
            self.welcome_seen = data.get("welcome_seen", False)
            self.default_background_mode = data.get("default_background_mode", "remember")
            self.max_highlights_before_warning = int(data.get("max_highlights_before_warning", 5000))
            self.confirm_bulk_edit_over = int(data.get("confirm_bulk_edit_over", 1))
            self.confirm_bulk_delete_over = int(data.get("confirm_bulk_delete_over", 1))
            self.confirm_bulk_actions = data.get("confirm_bulk_actions", True)
            self.npc_data_source_path = data.get("npc_data_source_path", "")
            self.current_expansion_number_setting = int(data.get("current_expansion_number", 0))
            if hasattr(self, "side_panel"):
                self.side_panel.set_npc_source_path(self.npc_data_source_path)
                self.side_panel.set_expansion_number(self.current_expansion_number_setting)
            self.mapper.flip_display_y = bool(data.get("flip_display_y", False))
            if hasattr(self, "flip_y_action"):
                self.flip_y_action.setChecked(self.mapper.flip_display_y)
            self.apply_preferences_to_panel()

            if self.default_background_mode == "light":
                self.set_background("dark")
            elif self.default_background_mode == "dark":
                self.set_background("dark")
            elif data.get("dark_mode"):
                self.set_background("dark")
            else:
                self.set_background("dark")

            self.map_folder = data.get("map_folder", "")
            if self.map_folder and hasattr(self, "side_panel"):
                self.side_panel.map_folder_edit.setText(self.map_folder)
                self.rebuild_zone_list()

            files = [Path(path) for path in data.get("last_files", [])] if self.remember_last_loaded_map else []
            files = [path for path in files if path.exists()]
            if files:
                self.load_files(files, remember=False)
                self._fit_last_map_after_show = bool(self.auto_fit_restored_map)
                self.status_label.setText("Restored last loaded map files." + (" Fit will run after the window is shown." if self.auto_fit_restored_map else ""))

            sizes = data.get("sidebar_sizes", [])
            if sizes and hasattr(self, "splitter"):
                self.splitter.setSizes([int(x) for x in sizes])
        except Exception:
            pass

    def open_preferences(self) -> None:
        dialog = PreferencesDialog(self)
        dialog.exec()

    def open_backup_folder(self) -> None:
        try:
            BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
            os.startfile(str(BACKUPS_DIR))
        except Exception as exc:
            QMessageBox.warning(self, "Backup Folder", f"Could not open backup folder:\n{exc}")

    def maybe_show_startup_warning(self) -> None:
        if not self.show_beta_warning:
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("Beta Safety Warning")
        layout = QVBoxLayout(dialog)
        warning = QLabel(
            "This tool edits EverQuest map .txt files.\n\n"
            "Before saving edits, make sure you have backups of your map folder.\n"
            "For beta testing, use Save As... or work from a copied map folder."
        )
        warning.setWordWrap(True)
        layout.addWidget(warning)
        checkbox = QCheckBox("Do not show this again")
        layout.addWidget(checkbox)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)
        dialog.exec()
        if checkbox.isChecked():
            self.show_beta_warning = False
            self.save_settings()


    def open_logs_folder(self) -> None:
        try:
            LOGS_DIR.mkdir(parents=True, exist_ok=True)
            os.startfile(str(LOGS_DIR))
        except Exception as exc:
            QMessageBox.warning(self, "Logs Folder", f"Could not open logs folder:\n{exc}")

    def show_quick_start(self) -> None:
        QMessageBox.information(
            self,
            "EQ Map Editor Quick Start",
            (
                "Suggested beta workflow:\n\n"
                "1. Work from a copied map folder while testing.\n"
                "2. Use File > Open Map File(s) to load one zone's map files.\n"
                "3. Use Fit Map / Fit Selected and the mini-map to navigate.\n"
                "4. Select points or lines, then edit them in the Inspector.\n"
                "5. Use Bulk Colours > Palette Conversion to switch light/dark map palettes.\n"
                "6. Use Save As... for a safe test export, or Save Edits when ready.\n"
                "7. Backups are written to the local backups folder before source files are changed."
            ),
        )

    def show_keyboard_shortcuts(self) -> None:
        QMessageBox.information(
            self,
            "Keyboard Shortcuts",
            (
                "Keyboard shortcuts:\n\n"
                "Ctrl+O  Open map file(s)\n"
                "Ctrl+S  Save edits\n"
                "Ctrl+Z  Undo\n"
                "Ctrl+Y  Redo\n"
                "Esc     Clear selection\n"
                "Delete  Delete selected records\n"
                "Ctrl+H  Hide selected layer/records where supported\n\n"
                "Mouse:\n"
                "Wheel   Zoom\n"
                "Right or middle drag   Pan\n"
                "Double-left-click      Add point/line based on edit mode"
            ),
        )

    def show_about(self) -> None:
        QMessageBox.information(
            self,
            "About EQ Map Editor",
            (
                f"EQ Map Editor {VERSION}\n\n"
                "A desktop editor for EverQuest map .txt files.\n\n"
                "Local support folders:\n"
                f"- Settings: {SETTINGS_DIR}\n"
                f"- Backups: {BACKUPS_DIR}\n"
                f"- Logs: {LOGS_DIR}\n"
                f"- Palettes: {PALETTES_DIR}"
            ),
        )

    def maybe_show_welcome(self) -> None:
        if getattr(self, "welcome_seen", False):
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("Welcome to EQ Map Editor")
        layout = QVBoxLayout(dialog)
        text = QLabel(
            "Welcome to EQ Map Editor.\n\n"
            "Quick start:\n"
            "1. Open one or more EQ map .txt files.\n"
            "2. Use the map canvas to inspect, pan, zoom, and select records.\n"
            "3. Edit selected points/lines in the Inspector.\n"
            "4. Use Bulk Colours > Palette Conversion to switch light/dark palettes.\n"
            "5. Use Save As... first while testing, or Save Edits when you are ready.\n\n"
            "Tip: beta testers should work from a copied map folder."
        )
        text.setWordWrap(True)
        layout.addWidget(text)
        checkbox = QCheckBox("Do not show this again")
        layout.addWidget(checkbox)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)
        dialog.exec()
        if checkbox.isChecked():
            self.welcome_seen = True
            self.save_settings()

    def open_files_dialog(self) -> None:
        if not self.confirm_discard_unsaved():
            return
        file_names, _ = QFileDialog.getOpenFileNames(
            self,
            "Open EQ Map Text File(s)",
            "",
            "Text files (*.txt);;All files (*.*)",
        )
        if file_names:
            self.load_files([Path(name) for name in file_names])

    def load_files(self, file_paths: list[Path], remember: bool = True) -> None:
        try:
            loaded_map = load_map_files(file_paths)
        except OSError as exc:
            QMessageBox.critical(self, "Load Error", f"Could not load map files:\n{exc}")
            return

        self.loaded_files = [Path(path) for path in file_paths]
        self.loaded_file_mtimes = {path: path.stat().st_mtime for path in self.loaded_files if path.exists()}
        self.layer_visible = {path: True for path in self.loaded_files}
        self.loaded_map = loaded_map
        self.undo_stack.clear()
        self.redo_stack.clear()
        self.pending_line_start = None
        self.render_map()
        self.side_panel.rebuild_layers()
        if hasattr(self.side_panel, "rebuild_missing_style_match_combo"):
            self.side_panel.rebuild_missing_style_match_combo()
        self.update_dirty_indicator()
        if remember:
            self.save_last_session()

        file_list = ", ".join(path.name for path in self.loaded_files)
        self.status_label.setText(
            f"Loaded {len(loaded_map.lines):,} lines and {len(loaded_map.points):,} points from {file_list}"
        )
        self.log_event(f"Loaded files: {file_list}")
        if hasattr(self, "nav_panel"):
            self.nav_panel.refresh()


    def update_scene_pan_padding(self) -> None:
        rect = self.scene.itemsBoundingRect()
        if rect.isNull():
            return

        # Use at least one viewport-width/height of padding in scene units.
        # This makes the left/right/top/bottom map edges reachable near the center of the view.
        viewport_rect = self.view.mapToScene(self.view.viewport().rect()).boundingRect()
        pad_x = max(rect.width() * 0.25, viewport_rect.width() * 0.55, 250.0)
        pad_y = max(rect.height() * 0.25, viewport_rect.height() * 0.55, 250.0)

        padded = QRectF(
            rect.x() - pad_x,
            rect.y() - pad_y,
            rect.width() + pad_x * 2,
            rect.height() + pad_y * 2,
        )
        self.scene.setSceneRect(padded)

    def render_map(self, keep_view: bool = False) -> None:
        old_transform = self.view.transform()
        old_center = self.view.mapToScene(self.view.viewport().rect().center())

        self.scene.clear()
        self.line_items_by_record.clear()
        self.point_items_by_record.clear()
        self.label_items_by_record.clear()
        self.endpoint_handles.clear()
        self.selection_highlights.clear()
        self.npc_preview_items.clear()
        self.era_preview_items.clear()
        self.missing_preview_items.clear()
        self.bulk_selected_records = []
        self.side_panel.set_record(None)

        for record in self.loaded_map.lines:
            if not getattr(record, "deleted", False) and self.layer_visible.get(record.file_path, True):
                self.add_or_update_line_item(record)

        for record in self.loaded_map.points:
            if not getattr(record, "deleted", False) and self.layer_visible.get(record.file_path, True):
                self.add_or_update_point_items(record)

        self.scene.selectionChanged.connect(self.on_selection_changed)
        self.on_edit_mode_changed(self.edit_mode())
        self.update_scene_pan_padding()

        if keep_view:
            self.view.setTransform(old_transform)
            self.view.centerOn(old_center)
        else:
            self.fit_map()
        self.update_scene_pan_padding()
        self.update_canvas_overlays()

    def line_tooltip(self, record: MapLineRecord) -> str:
        return f"{record.file_path.name}:{record.line_index + 1 if record.line_index >= 0 else 'new'}\n{record.to_map_line()}"

    def point_tooltip(self, record: MapPointRecord) -> str:
        return f"{record.file_path.name}:{record.line_index + 1 if record.line_index >= 0 else 'new'}\n{record.to_map_line()}"

    def add_or_update_line_item(self, record: MapLineRecord) -> None:
        if not self.layer_visible.get(record.file_path, True):
            self.render_map(keep_view=True)
            return

        key = id(record)
        p1 = self.mapper.map_to_scene(record.x1, record.y1)
        p2 = self.mapper.map_to_scene(record.x2, record.y2)

        item = self.line_items_by_record.get(key)
        if item is None:
            item = MovableLineItem(self, record)
            self.scene.addItem(item)
            self.line_items_by_record[key] = item

        self._drag_update_lock = True
        item.setLine(0, 0, p2.x() - p1.x(), p2.y() - p1.y())
        item.setPos(p1)
        self._drag_update_lock = False

        item.setPen(QPen(record.color, 1.0))
        item.setToolTip(self.line_tooltip(record))
        item.setFlag(QGraphicsItem.ItemIsMovable, self.edit_mode() == "Move Lines")

    def add_or_update_point_items(self, record: MapPointRecord) -> None:
        if not self.layer_visible.get(record.file_path, True):
            self.render_map(keep_view=True)
            return

        key = id(record)
        point = self.mapper.map_to_scene(record.x, record.y)
        radius = max(2.0, record.size * 1.8)

        marker = self.point_items_by_record.get(key)
        if marker is None:
            marker = MovablePointMarker(self, record)
            self.scene.addItem(marker)
            self.point_items_by_record[key] = marker

        self._drag_update_lock = True
        marker.setRect(-radius, -radius, radius * 2, radius * 2)
        marker.setPos(point)
        self._drag_update_lock = False

        marker.setPen(QPen(record.color, 0.8))
        marker.setBrush(QBrush(record.color))
        marker.setVisible(self.show_points)
        marker.setToolTip(self.point_tooltip(record))
        marker.setFlag(QGraphicsItem.ItemIsMovable, self.edit_mode() == "Move Points")

        text = self.label_items_by_record.get(key)
        if text is None:
            text = QGraphicsTextItem()
            text.setData(0, record)
            text.setFlag(QGraphicsItem.ItemIsSelectable, True)
            self.scene.addItem(text)
            self.label_items_by_record[key] = text

        text.setPlainText(record.label.replace("_", " "))
        text.setDefaultTextColor(record.color)
        text.setFont(QFont("Arial", 6))
        text.setPos(point.x() + radius + 1, point.y() - 5)
        text.setVisible(self.show_labels)
        text.setToolTip(self.point_tooltip(record))

    def update_record_items(self, record: Any) -> None:
        if isinstance(record, MapPointRecord):
            self.add_or_update_point_items(record)
        elif isinstance(record, MapLineRecord):
            self.add_or_update_line_item(record)
            if self.side_panel.current_record is record and self.edit_mode() == "Move Line Endpoints":
                self.show_endpoint_handles(record)
        self.update_dirty_indicator()

    def setup_shortcuts(
        self,
        open_action: QAction,
        save_action: QAction,
        undo_callable,
        redo_callable,
        fit_callable,
        clear_callable,
        delete_callable,
        select_all_callable,
        search_widget: QLineEdit,
    ) -> None:
        open_action.setShortcut("Ctrl+O")
        save_action.setShortcut("Ctrl+S")

        undo_action = QAction("Undo", self)
        undo_action.setShortcut("Ctrl+Z")
        undo_action.triggered.connect(undo_callable)
        self.addAction(undo_action)

        redo_action = QAction("Redo", self)
        redo_action.setShortcut("Ctrl+Y")
        redo_action.triggered.connect(redo_callable)
        self.addAction(redo_action)

        focus_search_action = QAction("Focus Search", self)
        focus_search_action.setShortcut("Ctrl+F")
        focus_search_action.triggered.connect(lambda: (search_widget.setFocus(), search_widget.selectAll()))
        self.addAction(focus_search_action)

        fit_action = QAction("Fit Map Shortcut", self)
        fit_action.setShortcut("F")
        fit_action.triggered.connect(fit_callable)
        self.addAction(fit_action)

        clear_action = QAction("Clear Selection Shortcut", self)
        clear_action.setShortcut("Esc")
        clear_action.triggered.connect(clear_callable)
        self.addAction(clear_action)

        delete_action = QAction("Delete Selected Shortcut", self)
        delete_action.setShortcut("Del")
        delete_action.triggered.connect(delete_callable)
        self.addAction(delete_action)

        select_all_action = QAction("Select All Visible Records Shortcut", self)
        select_all_action.setShortcut("Ctrl+A")
        select_all_action.triggered.connect(select_all_callable)
        self.addAction(select_all_action)

    def visible_point_records(self) -> list[MapPointRecord]:
        return [
            point for point in self.loaded_map.points
            if not getattr(point, "deleted", False)
            and self.layer_visible.get(point.file_path, True)
        ]

    def label_search_text(self) -> str:
        return self.global_search_edit.text().strip().lower() if hasattr(self, "global_search_edit") else ""

    def find_label_matches(self) -> list[MapPointRecord]:
        search = self.label_search_text()
        if not search:
            return []
        return [point for point in self.visible_point_records() if search in point.label.lower()]

    def search_select_first_label_match(self) -> None:
        matches = self.find_label_matches()
        self.search_matches = matches
        self.search_index = 0 if matches else -1
        if not matches:
            self.status_label.setText("No point labels matched the search text.")
            return

        record = matches[0]
        self.scene.blockSignals(True)
        self.scene.clearSelection()
        item = self.point_items_by_record.get(id(record))
        if item:
            item.setSelected(True)
        self.scene.blockSignals(False)
        self.refresh_selection_highlights()
        self.side_panel.set_record(record)
        self.center_on_record(record)
        self.status_label.setText(f"Found 1 of {len(matches)} matching point label(s): {record.label}")

    def search_select_all_label_matches(self) -> None:
        matches = self.find_label_matches()
        self.search_matches = matches
        if not matches:
            self.status_label.setText("No point labels matched the search text.")
            return

        self.scene.blockSignals(True)
        self.scene.clearSelection()
        for record in matches:
            item = self.point_items_by_record.get(id(record))
            if item:
                item.setSelected(True)
        self.scene.blockSignals(False)
        self.refresh_selection_highlights()
        point_count = len(matches)
        self.side_panel.show_multi_selection(point_count, 0)
        self.fit_selected()
        self.status_label.setText(f"Selected {len(matches)} matching point label(s).")

    def center_on_record(self, record: Any) -> None:
        if isinstance(record, MapPointRecord):
            point = self.mapper.map_to_scene(record.x, record.y)
            self.view.centerOn(point)
        elif isinstance(record, MapLineRecord):
            p1 = self.mapper.map_to_scene(record.x1, record.y1)
            p2 = self.mapper.map_to_scene(record.x2, record.y2)
            self.view.centerOn((p1.x() + p2.x()) / 2, (p1.y() + p2.y()) / 2)

    def center_selected(self) -> None:
        records = self.selected_map_records() or self.bulk_selected_records
        if not records:
            self.status_label.setText("No selected records to center.")
            return

        rect = self.bounding_rect_for_records(records)
        if rect.isNull():
            self.center_on_record(records[0])
        else:
            self.view.centerOn(rect.center())
        self.status_label.setText(f"Centered on {len(records)} selected record(s).")

    def bounding_rect_for_records(self, records: list[Any]) -> QRectF:
        rect = QRectF()
        for record in records:
            if isinstance(record, MapPointRecord):
                point = self.mapper.map_to_scene(record.x, record.y)
                radius = max(10.0, record.size * 4.0)
                item_rect = QRectF(point.x() - radius, point.y() - radius, radius * 2, radius * 2)
            elif isinstance(record, MapLineRecord):
                p1 = self.mapper.map_to_scene(record.x1, record.y1)
                p2 = self.mapper.map_to_scene(record.x2, record.y2)
                item_rect = QRectF(p1, p2).normalized()
            else:
                continue
            rect = item_rect if rect.isNull() else rect.united(item_rect)
        return rect

    def select_all_visible_records(self) -> None:
        records: list[Any] = []
        records.extend([
            line for line in self.loaded_map.lines
            if not getattr(line, "deleted", False)
            and self.layer_visible.get(line.file_path, True)
        ])
        records.extend(self.visible_point_records())

        if not records:
            self.status_label.setText("No visible records to select.")
            return

        if self.confirm_bulk_actions and len(records) > 2500:
            if QMessageBox.question(
                self,
                "Select All Visible Records",
                f"Select {len(records)} visible records? This may take a moment.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            ) != QMessageBox.Yes:
                return

        self.scene.blockSignals(True)
        self.scene.clearSelection()
        for record in records:
            item = self.line_items_by_record.get(id(record)) or self.point_items_by_record.get(id(record))
            if item:
                item.setSelected(True)
        self.scene.blockSignals(False)
        self.refresh_selection_highlights()
        point_count = sum(1 for record in records if isinstance(record, MapPointRecord))
        line_count = sum(1 for record in records if isinstance(record, MapLineRecord))
        self.side_panel.show_multi_selection(point_count, line_count)
        self.status_label.setText(f"Selected all visible records: {point_count} point(s), {line_count} line(s).")

    def clear_selection_and_highlights(self) -> None:
        self.scene.clearSelection()
        self.clear_selection_highlights()
        self.bulk_selected_records = []
        self.side_panel.set_record(None)
        self.status_label.setText("Selection and highlights cleared.")

    def fit_selected(self) -> None:
        records = self.selected_map_records() or self.bulk_selected_records
        rect = self.bounding_rect_for_records(records)
        if rect.isNull():
            self.status_label.setText("No selected records to fit.")
            return

        margin = 75
        padded = QRectF(rect.x() - margin, rect.y() - margin, rect.width() + margin * 2, rect.height() + margin * 2)
        self.view.fitInView(padded, Qt.KeepAspectRatio)
        self.status_label.setText("Fit selected records.")

    def save_as_folder(self) -> None:
        if not self.loaded_files:
            QMessageBox.information(self, "No Files", "Open map files before using Save As.")
            return
        folder = QFileDialog.getExistingDirectory(self, "Choose Output Folder for Edited Map Files", str(self.loaded_files[0].parent))
        if not folder:
            return
        try:
            self.write_map_files_to_folder(Path(folder))
            self.log_event(f"Save As completed to {folder}")
            QMessageBox.information(self, "Save As Complete", f"Edited map files were written to:\n{folder}\n\nYour original map files were not changed.")
        except Exception as exc:
            self.log_event(f"Save As failed: {exc}")
            QMessageBox.critical(self, "Save As Error", f"Could not save edited copies:\n{exc}")

    def write_map_files_to_folder(self, output_folder: Path) -> None:
        output_folder.mkdir(parents=True, exist_ok=True)
        dirty_records = self.dirty_records()
        by_file: dict[Path, list[Any]] = {}
        for record in dirty_records:
            by_file.setdefault(record.file_path, []).append(record)

        for file_path in self.loaded_files:
            original_lines = file_path.read_text(encoding="utf-8", errors="replace").splitlines() if file_path.exists() else []
            records = by_file.get(file_path, [])
            append_records = []
            for record in records:
                if record.line_index < 0:
                    append_records.append(record)
                    continue
                if record.line_index >= len(original_lines):
                    raise IndexError(f"Line index out of range for {record.file_path.name}:{record.line_index + 1}")
                if getattr(record, "deleted", False):
                    original_lines[record.line_index] = None
                else:
                    original_lines[record.line_index] = record.to_map_line()

            original_lines = [line for line in original_lines if line is not None]
            for record in append_records:
                if not getattr(record, "deleted", False):
                    original_lines.append(record.to_map_line())

            (output_folder / file_path.name).write_text("\n".join(original_lines) + "\n", encoding="utf-8")

    def fit_map(self) -> None:
        rect = self.scene.itemsBoundingRect()
        if rect.isNull():
            return
        margin = 50
        padded = QRectF(
            rect.x() - margin,
            rect.y() - margin,
            rect.width() + margin * 2,
            rect.height() + margin * 2,
        )
        self.view.fitInView(padded, Qt.KeepAspectRatio)
        self.update_scene_pan_padding()
        self.update_canvas_overlays()

    def toggle_labels(self, checked: bool) -> None:
        self.show_labels = checked
        for item in self.label_items_by_record.values():
            item.setVisible(checked)

    def toggle_points(self, checked: bool) -> None:
        self.show_points = checked
        for item in self.point_items_by_record.values():
            item.setVisible(checked)

    def toggle_display_y_flip(self, checked: bool) -> None:
        self.mapper.flip_display_y = checked
        self.render_map(keep_view=False)
        self.status_label.setText("Display Y flipped." if checked else "Display Y normal.")

    def set_layer_visible(self, file_path: Path, visible: bool) -> None:
        self.layer_visible[file_path] = visible
        self.render_map(keep_view=True)
        self.side_panel.rebuild_layers()
        if hasattr(self, "nav_panel"):
            self.nav_panel.refresh()
        self.status_label.setText(f"{file_path.name} visible: {visible}")

    def on_edit_mode_changed(self, mode: str) -> None:
        if hasattr(self, "canvas_controls"):
            self.canvas_controls.set_active_mode(mode)
        for item in self.point_items_by_record.values():
            item.setFlag(QGraphicsItem.ItemIsMovable, mode == "Move Points")
        for item in self.line_items_by_record.values():
            item.setFlag(QGraphicsItem.ItemIsMovable, mode == "Move Lines")

        self.clear_endpoint_handles()
        selected = self.scene.selectedItems()
        if mode == "Move Line Endpoints" and selected:
            record = selected[0].data(0)
            if isinstance(record, MapLineRecord):
                self.show_endpoint_handles(record)

        self.pending_line_start = None
        self.status_label.setText(f"Edit mode: {mode}")

    def clear_selection_highlights(self) -> None:
        for item in self.selection_highlights:
            try:
                self.scene.removeItem(item)
            except RuntimeError:
                pass
        self.selection_highlights.clear()

    def add_highlights_for_records(self, records: list[Any], max_items: int = 0) -> int:
        self.clear_selection_highlights()
        glow_colour = QColor(255, 230, 80, 130)
        glow_pen = QPen(glow_colour, 7.0)
        glow_brush = QBrush(QColor(255, 230, 80, 70))

        highlighted = 0
        records_to_draw = records if not max_items else records[:max_items]
        for record in records_to_draw:
            if isinstance(record, MapPointRecord):
                point = self.mapper.map_to_scene(record.x, record.y)
                radius = max(7.0, record.size * 3.2)
                glow = QGraphicsEllipseItem(point.x() - radius, point.y() - radius, radius * 2, radius * 2)
                glow.setBrush(glow_brush)
                glow.setPen(QPen(glow_colour, 2.0))
                glow.setZValue(-500)
                self.scene.addItem(glow)
                self.selection_highlights.append(glow)
                highlighted += 1

            elif isinstance(record, MapLineRecord):
                p1 = self.mapper.map_to_scene(record.x1, record.y1)
                p2 = self.mapper.map_to_scene(record.x2, record.y2)
                glow = QGraphicsLineItem(p1.x(), p1.y(), p2.x(), p2.y())
                glow.setPen(glow_pen)
                glow.setZValue(-500)
                self.scene.addItem(glow)
                self.selection_highlights.append(glow)
                highlighted += 1

        return highlighted

    def refresh_selection_highlights(self) -> None:
        records = []
        records_seen = set()
        for item in self.scene.selectedItems():
            record = item.data(0)
            if id(record) in records_seen:
                continue
            if isinstance(record, (MapPointRecord, MapLineRecord)):
                records_seen.add(id(record))
                records.append(record)
        self.add_highlights_for_records(records)

    def on_selection_changed(self) -> None:
        selected = self.scene.selectedItems()
        self.refresh_selection_highlights()
        if not selected:
            self.side_panel.set_record(None)
            self.side_panel.set_multi_selection_summary(0, 0)
            self.clear_endpoint_handles()
            return

        if len(selected) > 1:
            point_count = sum(1 for item in selected if isinstance(item.data(0), MapPointRecord))
            line_count = sum(1 for item in selected if isinstance(item.data(0), MapLineRecord))
            self.side_panel.show_multi_selection(point_count, line_count)
            self.clear_endpoint_handles()
            self.status_label.setText(f"Selected {point_count} point(s) and {line_count} line(s).")
            return

        record = selected[0].data(0)
        self.side_panel.set_record(record)
        self.clear_endpoint_handles()

        if isinstance(record, MapPointRecord):
            self.status_label.setText(
                f"Point: {record.label} | X={record.x:.4f}, Y={record.y:.4f}, Z={record.z:.4f} | "
                f"RGB=({record.r}, {record.g}, {record.b}) | {record.file_path.name}:{record.line_index + 1 if record.line_index >= 0 else 'new'}"
            )
        elif isinstance(record, MapLineRecord):
            self.status_label.setText(
                f"Line | RGB=({record.r}, {record.g}, {record.b}) | {record.file_path.name}:{record.line_index + 1 if record.line_index >= 0 else 'new'}"
            )
            if self.edit_mode() == "Move Line Endpoints":
                self.show_endpoint_handles(record)

    def clear_endpoint_handles(self) -> None:
        for handle in self.endpoint_handles:
            self.scene.removeItem(handle)
        self.endpoint_handles.clear()

    def show_endpoint_handles(self, record: MapLineRecord) -> None:
        self.clear_endpoint_handles()
        for endpoint, x, y in [(1, record.x1, record.y1), (2, record.x2, record.y2)]:
            handle = EndpointHandle(self, record, endpoint)
            handle.setPos(self.mapper.map_to_scene(x, y))
            self.scene.addItem(handle)
            self.endpoint_handles.append(handle)

    def on_marker_dragged(self, point: MapPointRecord, scene_pos: QPointF) -> None:
        if self._drag_update_lock or self.edit_mode() != "Move Points":
            return

        point.x, point.y = self.mapper.scene_to_map(scene_pos)
        point.dirty = True

        label = self.label_items_by_record.get(id(point))
        if label is not None:
            radius = max(2.0, point.size * 1.8)
            label.setPos(scene_pos.x() + radius + 1, scene_pos.y() - 5)

        if self.side_panel.current_record is point:
            self.side_panel.set_record(point)

        self.update_dirty_indicator()
        self.status_label.setText(f"Moved point: {point.label} | X={point.x:.4f}, Y={point.y:.4f}")

    def on_line_dragged(self, line: MapLineRecord, scene_pos: QPointF) -> None:
        if self._drag_update_lock or self.edit_mode() != "Move Lines":
            return

        new_x1, new_y1 = self.mapper.scene_to_map(scene_pos)
        dx = new_x1 - line.x1
        dy = new_y1 - line.y1

        line.x1 += dx
        line.y1 += dy
        line.x2 += dx
        line.y2 += dy
        line.dirty = True

        if self.side_panel.current_record is line:
            self.side_panel.set_record(line)

        self.update_dirty_indicator()
        self.status_label.setText(
            f"Moved line | ({line.x1:.2f}, {line.y1:.2f}) -> ({line.x2:.2f}, {line.y2:.2f})"
        )

    def on_endpoint_dragged(self, line: MapLineRecord, endpoint: int, scene_pos: QPointF) -> None:
        if self._drag_update_lock or self.edit_mode() != "Move Line Endpoints":
            return

        x, y = self.mapper.scene_to_map(scene_pos)
        if endpoint == 1:
            line.x1 = x
            line.y1 = y
        else:
            line.x2 = x
            line.y2 = y

        line.dirty = True
        self.add_or_update_line_item(line)

        if self.side_panel.current_record is line:
            self.side_panel.set_record(line)

        self.update_dirty_indicator()
        self.status_label.setText(f"Moved line endpoint {endpoint} | X={x:.2f}, Y={y:.2f}")

    def active_file_for_new_records(self) -> Optional[Path]:
        if not self.loaded_files:
            return None
        index = self.side_panel.active_file_combo.currentIndex()
        if index >= 0:
            data = self.side_panel.active_file_combo.itemData(index)
            if data:
                return Path(data)
        return self.loaded_files[0]

    def handle_canvas_double_click(self, scene_pos: QPointF) -> None:
        mode = self.edit_mode()
        active_file = self.active_file_for_new_records()
        if active_file is None:
            return

        if mode == "Add New Point":
            label, ok = QInputDialog.getText(self, "Add New Point", "Point label:")
            if not ok:
                return
            x, y = self.mapper.scene_to_map(scene_pos)
            point = MapPointRecord(
                file_path=active_file,
                line_index=-1,
                raw_text="",
                x=x,
                y=y,
                z=0.0,
                r=255,
                g=255,
                b=255,
                size=2.0,
                label=label.strip() or "New_Point",
                dirty=True,
            )
            self.loaded_map.points.append(point)
            self.add_or_update_point_items(point)
            self.side_panel.rebuild_layers()
            self.update_dirty_indicator()
            self.status_label.setText(f"Added point to {target_file.name}. Save Edits will append it.")

        elif mode == "Add New Line":
            if self.pending_line_start is None:
                self.pending_line_start = scene_pos
                self.status_label.setText("Add New Line: double-click the second endpoint.")
                return

            x1, y1 = self.mapper.scene_to_map(self.pending_line_start)
            x2, y2 = self.mapper.scene_to_map(scene_pos)
            line = MapLineRecord(
                file_path=active_file,
                line_index=-1,
                raw_text="",
                x1=x1,
                y1=y1,
                z1=0.0,
                x2=x2,
                y2=y2,
                z2=0.0,
                r=255,
                g=255,
                b=255,
                dirty=True,
            )
            self.pending_line_start = None
            self.loaded_map.lines.append(line)
            self.add_or_update_line_item(line)
            self.side_panel.rebuild_layers()
            self.update_dirty_indicator()
            self.status_label.setText(f"Added line to {target_file.name}. Save Edits will append it.")

    def show_map_context_menu(self, global_pos, scene_pos: QPointF) -> None:
        if not self.loaded_files:
            return
        menu = QMenu(self)
        add_point_action = menu.addAction("Add Point Here")
        add_line_start_action = menu.addAction("Start New Line Here")
        if self.pending_line_start is not None:
            add_line_finish_action = menu.addAction("Finish New Line Here")
        else:
            add_line_finish_action = None

        chosen = menu.exec(global_pos)
        if chosen == add_point_action:
            self.add_point_at(scene_pos)
        elif chosen == add_line_start_action:
            self.start_line_at(scene_pos)
        elif add_line_finish_action is not None and chosen == add_line_finish_action:
            self.finish_line_at(scene_pos)

    def add_point_at(self, scene_pos: QPointF) -> None:
        active_file = self.active_file_for_new_records()
        if active_file is None:
            return

        dialog = PointDetailsDialog(self, title="Add Point", files=self.loaded_files, active_file=active_file)
        if dialog.exec() != QDialog.Accepted:
            return
        values = dialog.values()
        target_file = values.get("file_path") or active_file
        x, y = self.mapper.scene_to_map(scene_pos)
        point = MapPointRecord(
            file_path=target_file,
            line_index=-1,
            raw_text="",
            x=x,
            y=y,
            z=0.0,
            r=values["r"],
            g=values["g"],
            b=values["b"],
            size=values["size"],
            label=values["label"],
            dirty=True,
        )
        self.loaded_map.points.append(point)
        self.undo_stack.append(AddRecordCommand("Add point", self, point, "points"))
        self.redo_stack.clear()
        self.add_or_update_point_items(point)
        self.side_panel.rebuild_layers()
        self.update_dirty_indicator()
        self.status_label.setText(f"Added point to {target_file.name}. Save Edits will append it.")

    def start_line_at(self, scene_pos: QPointF) -> None:
        self.pending_line_start = scene_pos
        QMessageBox.information(
            self,
            "Add Line",
            "First endpoint has been set. Double-left-click the second endpoint and choose 'Finish New Line Here'.",
        )
        self.status_label.setText("New line endpoint 1 set. Double-left-click endpoint 2 and choose Finish New Line Here.")

    def finish_line_at(self, scene_pos: QPointF) -> None:
        active_file = self.active_file_for_new_records()
        if active_file is None or self.pending_line_start is None:
            return

        dialog = LineColorDialog(self, files=self.loaded_files, active_file=active_file)
        if dialog.exec() != QDialog.Accepted:
            return
        values = dialog.values()
        target_file = values.get("file_path") or active_file

        x1, y1 = self.mapper.scene_to_map(self.pending_line_start)
        x2, y2 = self.mapper.scene_to_map(scene_pos)
        self.pending_line_start = None

        line = MapLineRecord(
            file_path=target_file,
            line_index=-1,
            raw_text="",
            x1=x1,
            y1=y1,
            z1=0.0,
            x2=x2,
            y2=y2,
            z2=0.0,
            r=values["r"],
            g=values["g"],
            b=values["b"],
            dirty=True,
        )
        self.loaded_map.lines.append(line)
        self.undo_stack.append(AddRecordCommand("Add line", self, line, "lines"))
        self.redo_stack.clear()
        self.add_or_update_line_item(line)
        self.side_panel.rebuild_layers()
        self.update_dirty_indicator()
        self.status_label.setText(f"Added line to {target_file.name}. Save Edits will append it.")

    def selected_colours_from_bulk_tab(self, record_type: Optional[str] = None) -> list[tuple[int, int, int]]:
        record_type = record_type or self.current_bulk_colour_type
        list_widget = self.side_panel.point_colour_list if record_type == "points" else self.side_panel.line_colour_list
        colours = [item.data(Qt.UserRole) for item in list_widget.selectedItems()]
        return [colour for colour in colours if colour is not None]

    def selected_colour_from_bulk_tab(self) -> Optional[tuple[int, int, int]]:
        colours = self.selected_colours_from_bulk_tab()
        return colours[0] if colours else None

    def all_colour_counts(self) -> list[tuple[tuple[int, int, int], int, int]]:
        counts: dict[tuple[int, int, int], dict[str, int]] = {}
        for point in self.loaded_map.points:
            if getattr(point, "deleted", False) or not self.layer_visible.get(point.file_path, True):
                continue
            rgb = (point.r, point.g, point.b)
            counts.setdefault(rgb, {"points": 0, "lines": 0})["points"] += 1
        for line in self.loaded_map.lines:
            if getattr(line, "deleted", False) or not self.layer_visible.get(line.file_path, True):
                continue
            rgb = (line.r, line.g, line.b)
            counts.setdefault(rgb, {"points": 0, "lines": 0})["lines"] += 1

        rows = [(rgb, values["points"], values["lines"]) for rgb, values in counts.items()]
        return sorted(rows, key=lambda row: row[1] + row[2], reverse=True)

    def select_records_by_colour(self, record_type: str) -> None:
        self.current_bulk_colour_type = record_type
        colours = set(self.selected_colours_from_bulk_tab(record_type))
        if not colours:
            return

        # Do not call item.setSelected(True) for thousands of records. It triggers many
        # selection-change events and can be extremely slow/crashy with large EQ map files.
        self.scene.blockSignals(True)
        self.scene.clearSelection()
        self.scene.blockSignals(False)

        records: list[Any] = []
        if record_type == "points":
            for point in self.loaded_map.points:
                if getattr(point, "deleted", False) or not self.layer_visible.get(point.file_path, True):
                    continue
                if (point.r, point.g, point.b) in colours:
                    records.append(point)
        elif record_type == "lines":
            for line in self.loaded_map.lines:
                if getattr(line, "deleted", False) or not self.layer_visible.get(line.file_path, True):
                    continue
                if (line.r, line.g, line.b) in colours:
                    records.append(line)

        self.bulk_selected_records = records
        if (
            self.max_highlights_before_warning
            and len(records) > self.max_highlights_before_warning
            and QMessageBox.question(
                self,
                "Highlight Many Records",
                f"This will highlight {len(records)} {record_type}. Continue?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            ) != QMessageBox.Yes
        ):
            self.status_label.setText(f"Prepared {len(records)} {record_type}, but highlights were not drawn.")
            return
        highlighted = self.add_highlights_for_records(records)
        self.status_label.setText(f"Prepared and highlighted {highlighted} {record_type} matching {len(colours)} colour(s).")



    def apply_palette_mapping_preview(self) -> None:
        mappings = self.side_panel.current_palette_mapping()
        if not mappings:
            self.status_label.setText("No palette mappings to apply.")
            return

        mapping_lookup = {(row["scope"], tuple(row["rgb"])): tuple(row["target_rgb"]) for row in mappings}

        changes: list[tuple[Any, tuple[int, int, int]]] = []
        for point in self.loaded_map.points:
            if getattr(point, "deleted", False) or not self.layer_visible.get(point.file_path, True):
                continue
            key = ("Points", (point.r, point.g, point.b))
            if key in mapping_lookup and mapping_lookup[key] != key[1]:
                changes.append((point, mapping_lookup[key]))

        for line in self.loaded_map.lines:
            if getattr(line, "deleted", False) or not self.layer_visible.get(line.file_path, True):
                continue
            key = ("Lines", (line.r, line.g, line.b))
            if key in mapping_lookup and mapping_lookup[key] != key[1]:
                changes.append((line, mapping_lookup[key]))

        if not changes:
            self.status_label.setText("No visible records would change from the current mapping.")
            return

        point_count = sum(1 for record, _rgb in changes if isinstance(record, MapPointRecord))
        line_count = sum(1 for record, _rgb in changes if isinstance(record, MapLineRecord))
        target = self.side_panel.palette_target_combo.currentText()
        palette = self.side_panel.selected_palette() or {}

        if not self.confirm_bulk_action(
            "Apply Palette Mapping",
            f"Apply palette mapping to {len(changes)} visible record(s)?\\n\\n"
            f"Palette: {palette.get('name', 'Palette')}\\nTarget: {target}\\n"
            f"Points: {point_count}\\nLines: {line_count}\\n\\n"
            "Only mapped point/line colour groups will be changed. Groups set to Skip will be ignored.",
            len(changes),
            action_type="edit",
        ):
            return

        affected = [record for record, _rgb in changes]
        before = [snapshot_point(r) if isinstance(r, MapPointRecord) else snapshot_line(r) for r in affected]
        for record, new_rgb in changes:
            record.r, record.g, record.b = new_rgb
            record.dirty = True
        after = [snapshot_point(r) if isinstance(r, MapPointRecord) else snapshot_line(r) for r in affected]

        self.undo_stack.append(BulkEditCommand(f"Apply {target} palette mapping", affected, before, after))
        self.redo_stack.clear()
        self.render_map(keep_view=True)
        self.update_dirty_indicator()
        self.status_label.setText(f"Applied palette mapping to {len(changes)} visible record(s).")

    def apply_palette_to_visible_records(self) -> None:
        palette = self.side_panel.selected_palette()
        if not palette:
            self.status_label.setText("No palette selected.")
            return

        target_mode = self.side_panel.palette_target_combo.currentText()
        records: list[Any] = []
        for record in list(self.loaded_map.lines) + list(self.loaded_map.points):
            if getattr(record, "deleted", False):
                continue
            if not self.layer_visible.get(record.file_path, True):
                continue
            records.append(record)

        if not records:
            self.status_label.setText("No visible records to recolour.")
            return

        changes: list[tuple[Any, tuple[int, int, int]]] = []
        for record in records:
            current_rgb = (record.r, record.g, record.b)
            new_rgb = map_rgb_to_palette(current_rgb, palette, target_mode)
            if new_rgb != current_rgb:
                changes.append((record, new_rgb))

        if not changes:
            self.status_label.setText(f"Visible records already match the {target_mode.lower()} palette.")
            return

        point_count = sum(1 for record, _rgb in changes if isinstance(record, MapPointRecord))
        line_count = sum(1 for record, _rgb in changes if isinstance(record, MapLineRecord))
        if not self.confirm_bulk_action(
            "Apply Colour Palette",
            f"Apply palette '{palette.get('name', 'Palette')}' to {len(changes)} visible record(s)?\n\n"
            f"Target: {target_mode}\nPoints: {point_count}\nLines: {line_count}\n\n"
            "This maps each current colour to the nearest palette entry and applies that entry's light/dark version.",
            len(changes),
            action_type="edit",
        ):
            return

        affected = [record for record, _rgb in changes]
        before = [snapshot_point(r) if isinstance(r, MapPointRecord) else snapshot_line(r) for r in affected]
        for record, new_rgb in changes:
            record.r, record.g, record.b = new_rgb
            record.dirty = True
        after = [snapshot_point(r) if isinstance(r, MapPointRecord) else snapshot_line(r) for r in affected]

        self.undo_stack.append(BulkEditCommand(f"Apply {target_mode} palette", affected, before, after))
        self.redo_stack.clear()
        self.render_map(keep_view=True)
        self.update_dirty_indicator()
        self.status_label.setText(f"Applied {target_mode.lower()} palette to {len(changes)} visible record(s).")

    def apply_bulk_colour_to_selected_matching(self) -> None:
        colours = set(self.selected_colours_from_bulk_tab(self.current_bulk_colour_type))
        if not colours:
            return

        records = [
            record for record in self.bulk_selected_records
            if isinstance(record, (MapPointRecord, MapLineRecord))
            and not getattr(record, "deleted", False)
            and (record.r, record.g, record.b) in colours
        ]

        # Fallback for normal manually selected records.
        if not records:
            for item in self.scene.selectedItems():
                record = item.data(0)
                if isinstance(record, (MapPointRecord, MapLineRecord)) and not getattr(record, "deleted", False):
                    if (record.r, record.g, record.b) in colours and record not in records:
                        records.append(record)

        if not records:
            self.status_label.setText("No prepared or selected records match the selected colour(s).")
            return

        point_count = sum(1 for record in records if isinstance(record, MapPointRecord))
        line_count = sum(1 for record in records if isinstance(record, MapLineRecord))
        new_rgb_preview = (self.side_panel.bulk_new_r.value(), self.side_panel.bulk_new_g.value(), self.side_panel.bulk_new_b.value())
        if not self.confirm_bulk_action(
            "Confirm Bulk Recolour",
            f"Recolour {len(records)} record(s)?\n\nPoints: {point_count}\nLines: {line_count}\nNew colour: RGB {new_rgb_preview}",
            len(records),
            action_type="edit",
        ):
            return

        before = [snapshot_point(r) if isinstance(r, MapPointRecord) else snapshot_line(r) for r in records]
        new_rgb = (self.side_panel.bulk_new_r.value(), self.side_panel.bulk_new_g.value(), self.side_panel.bulk_new_b.value())
        for record in records:
            record.r, record.g, record.b = new_rgb
            record.dirty = True
        after = [snapshot_point(r) if isinstance(r, MapPointRecord) else snapshot_line(r) for r in records]
        self.undo_stack.append(BulkEditCommand("Bulk recolour selected matching records", records, before, after))
        self.redo_stack.clear()
        self.bulk_selected_records = []
        self.render_map(keep_view=True)
        self.update_dirty_indicator()
        self.status_label.setText(f"Recoloured {len(records)} record(s).")

    def recolour_selected_records_from_panel(self) -> None:
        records = self.selected_map_records()
        if not records:
            self.status_label.setText("No selected points or lines to recolour.")
            return

        point_count = sum(1 for record in records if isinstance(record, MapPointRecord))
        line_count = sum(1 for record in records if isinstance(record, MapLineRecord))
        preview_rgb = (
            self.side_panel.r_spin.value(),
            self.side_panel.g_spin.value(),
            self.side_panel.b_spin.value(),
        )
        if not self.confirm_bulk_action(
            "Confirm Selected Recolour",
            f"Recolour {len(records)} selected record(s)?\n\nPoints: {point_count}\nLines: {line_count}\nNew colour: RGB {preview_rgb}",
            len(records),
            action_type="edit",
        ):
            return

        before = []
        for record in records:
            if isinstance(record, MapPointRecord):
                before.append(snapshot_point(record))
            else:
                before.append(snapshot_line(record))

        new_rgb = (
            self.side_panel.r_spin.value(),
            self.side_panel.g_spin.value(),
            self.side_panel.b_spin.value(),
        )

        for record in records:
            record.r, record.g, record.b = new_rgb
            record.dirty = True

        after = []
        for record in records:
            if isinstance(record, MapPointRecord):
                after.append(snapshot_point(record))
            else:
                after.append(snapshot_line(record))

        self.undo_stack.append(BulkEditCommand("Recolour selected records", records, before, after))
        self.redo_stack.clear()
        self.render_map(keep_view=True)
        self.update_dirty_indicator()
        self.status_label.setText(f"Recoloured {len(records)} selected record(s) to RGB {new_rgb}.")

    def selected_map_records(self) -> list[Any]:
        records: list[Any] = []
        seen = set()
        for item in self.scene.selectedItems():
            record = item.data(0)
            if isinstance(record, (MapPointRecord, MapLineRecord)) and not getattr(record, "deleted", False):
                if id(record) not in seen:
                    seen.add(id(record))
                    records.append(record)
        return records

    def delete_selected_records(self) -> None:
        records = self.selected_map_records()
        if not records:
            self.status_label.setText("No selected points or lines to delete.")
            return

        point_count = sum(1 for record in records if isinstance(record, MapPointRecord))
        line_count = sum(1 for record in records if isinstance(record, MapLineRecord))

        if not self.confirm_bulk_action(
            "Delete Selected Records",
            f"Mark {len(records)} selected record(s) for deletion?\n\n"
            f"Points: {point_count}\nLines: {line_count}\n\n"
            "Save Edits will remove them from the map text files.",
            len(records),
            action_type="delete",
        ):
            return

        before = []
        for record in records:
            if isinstance(record, MapPointRecord):
                before.append(snapshot_point(record) | {"deleted": getattr(record, "deleted", False)})
            else:
                before.append(snapshot_line(record) | {"deleted": getattr(record, "deleted", False)})

        for record in records:
            record.deleted = True
            record.dirty = True

        after = []
        for record in records:
            if isinstance(record, MapPointRecord):
                after.append(snapshot_point(record) | {"deleted": getattr(record, "deleted", False)})
            else:
                after.append(snapshot_line(record) | {"deleted": getattr(record, "deleted", False)})

        self.undo_stack.append(BulkEditCommand("Delete selected records", records, before, after))
        self.redo_stack.clear()
        self.render_map(keep_view=True)
        self.side_panel.rebuild_layers()
        self.update_dirty_indicator()
        self.status_label.setText(f"Marked {point_count} point(s) and {line_count} line(s) for deletion.")

    def bulk_recolour_checked_points(self) -> None:
        points = self.side_panel.checked_point_records()
        if not points:
            self.status_label.setText("No checked points to recolour.")
            return
        before = [snapshot_point(point) for point in points]
        new_rgb = (self.side_panel.point_bulk_r.value(), self.side_panel.point_bulk_g.value(), self.side_panel.point_bulk_b.value())
        for point in points:
            point.r, point.g, point.b = new_rgb
            point.dirty = True
            self.update_record_items(point)
        after = [snapshot_point(point) for point in points]
        self.undo_stack.append(BulkEditCommand("Bulk recolour checked points", points, before, after))
        self.redo_stack.clear()
        self.render_map(keep_view=True)
        self.update_dirty_indicator()
        self.status_label.setText(f"Recoloured {len(points)} checked point(s).")

    def bulk_delete_checked_points(self) -> None:
        points = self.side_panel.checked_point_records()
        if not points:
            self.status_label.setText("No checked points to delete.")
            return
        if not self.confirm_bulk_action(
            "Delete Checked Points",
            f"Mark {len(points)} checked point(s) for deletion?",
            len(points),
            action_type="delete",
        ):
            return
        before = [snapshot_point(point) | {"deleted": getattr(point, "deleted", False)} for point in points]
        for point in points:
            point.deleted = True
            point.dirty = True
        after = [snapshot_point(point) | {"deleted": getattr(point, "deleted", False)} for point in points]
        self.undo_stack.append(BulkEditCommand("Bulk delete checked points", points, before, after))
        self.redo_stack.clear()
        self.render_map(keep_view=True)
        self.side_panel.rebuild_layers()
        self.update_dirty_indicator()
        self.status_label.setText(f"Marked {len(points)} point(s) for deletion. Save Edits will remove them.")

    def choose_npc_data_source(self) -> None:
        start_dir = str(Path(self.npc_data_source_path).parent) if self.npc_data_source_path else str(APP_ROOT)
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Choose NPC Data Source",
            start_dir,
            "NPC data (*.xlsx *.csv);;Excel workbooks (*.xlsx);;CSV files (*.csv);;All files (*.*)",
        )
        if not file_name:
            return
        self.npc_data_source_path = file_name
        self.side_panel.set_npc_source_path(file_name)
        self.npc_data_rows = []
        self.save_settings()
        self.reload_npc_data_source()

    def reload_npc_data_source(self) -> None:
        if not self.npc_data_source_path:
            QMessageBox.information(self, "NPC Data", "Choose an NPC data source first.")
            return
        try:
            self.npc_data_rows = self.load_npc_data_rows(Path(self.npc_data_source_path))
        except Exception as exc:
            QMessageBox.critical(self, "NPC Data Load Error", f"Could not load NPC data:\n{exc}")
            self.log_event(f"NPC data load failed: {exc}\n{traceback.format_exc()}")
            return
        self.status_label.setText(f"Loaded {len(self.npc_data_rows):,} NPC data row(s).")
        if hasattr(self.side_panel, "npc_summary_label"):
            self.side_panel.npc_summary_label.setText(f"Loaded {len(self.npc_data_rows):,} NPC data row(s). Click Compare Current Zone.")

    def load_npc_data_rows(self, path: Path) -> list[NpcDataRow]:
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(str(path))
        raw_rows: list[dict[str, Any]] = []
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                raw_rows = [dict(row) for row in reader]
        elif path.suffix.lower() == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("Reading .xlsx NPC data requires openpyxl. Install it with: pip install openpyxl") from exc
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            headers = [clean_cell(value) for value in next(rows_iter, [])]
            for values in rows_iter:
                raw_rows.append({headers[index]: values[index] if index < len(values) else None for index in range(len(headers))})
        else:
            raise RuntimeError("NPC data source must be a .xlsx or .csv file.")

        npc_rows: list[NpcDataRow] = []
        for index, row in enumerate(raw_rows):
            zone = row_get(row, "zone_shortname", "zone_short_name", "zone")
            npc_name = row_get(row, "npc_name", "name")
            npc_role = row_get(row, "npc_role", "lastname", "role")
            if not zone or not npc_name:
                continue
            x = row_get_number(row, "npc_map_x", "map_x", "inverted_spawn_y", "spawn_y")
            y = row_get_number(row, "npc_map_y", "map_y", "inverted_spawn_x", "spawn_x")
            z = row_get_number(row, "npc_map_z", "map_z", "spawn_z")
            if x is None or y is None or z is None:
                continue
            min_expansion_number = row_get_int(row, "min_expansion_number", "spawn_min_expansion", "min_expansion", "min_expansion_id")
            max_expansion_number = row_get_int(row, "max_expansion_number", "spawn_max_expansion", "max_expansion", "max_expansion_id")
            min_expansion_name = row_get(row, "min_expansion", "spawn_min_expansion_name", "min_expansion_name")
            max_expansion_name = row_get(row, "max_expansion", "spawn_max_expansion_name", "max_expansion_name")
            npc_rows.append(NpcDataRow(
                source_index=index,
                zone_shortname=zone.lower(),
                npc_name=npc_name,
                npc_role=npc_role,
                npc_label=generated_npc_label(npc_name, npc_role),
                x=float(x), y=float(y), z=float(z),
                min_expansion_number=min_expansion_number,
                max_expansion_number=max_expansion_number,
                min_expansion_name=min_expansion_name,
                max_expansion_name=max_expansion_name,
                raw=row,
            ))
        return npc_rows

    def current_loaded_zone_shortname(self) -> str:
        return current_zone_shortname_from_files(self.loaded_files)

    def compare_current_zone_to_npc_data(self) -> None:
        if not self.loaded_files or not self.loaded_map.points:
            QMessageBox.information(self, "NPC Match", "Load a zone map before comparing NPC data.")
            return
        if not self.npc_data_rows:
            if self.npc_data_source_path:
                self.reload_npc_data_source()
            else:
                self.choose_npc_data_source()
            if not self.npc_data_rows:
                return

        zone = self.current_loaded_zone_shortname()
        npc_candidates = [row for row in self.npc_data_rows if row.zone_shortname == zone]
        if not npc_candidates:
            QMessageBox.information(self, "NPC Match", f"No NPC rows found for zone '{zone}'.")
            return

        point_candidates = [point for point in self.loaded_map.points if not getattr(point, "deleted", False)]
        results: list[NpcMatchResult] = []
        used_point_ids: set[int] = set()

        # Match NPC-first, mirroring extract_eq_map_points.py choose_best_map_match().
        for npc in npc_candidates:
            exact_matches: list[tuple[float, MapPointRecord]] = []
            possible_matches: list[tuple[float, MapPointRecord]] = []
            for point in point_candidates:
                score, match_type, dist = score_point_to_npc(point, npc)
                if match_type == "Yes":
                    exact_matches.append((dist if dist is not None else math.inf, point))
                elif match_type in {"Possible", "Coordinate Match"}:
                    possible_matches.append((dist if dist is not None else math.inf, point))

            if exact_matches:
                dist, best_point = min(exact_matches, key=lambda item: item[0])
                match_type = "Yes"
                score = 100.0
            elif possible_matches:
                dist, best_point = min(possible_matches, key=lambda item: item[0])
                if dist != math.inf and -20 < dist < 20:
                    match_type = "Coordinate Match"
                    score = 96.0
                else:
                    match_type = "Possible"
                    score = 82.0
            else:
                best_point = None
                dist = None
                match_type = "NPC only"
                score = 0.0

            if best_point is None:
                results.append(NpcMatchResult(None, npc, "NPC only", "", 0.0, None, False))
                continue

            used_point_ids.add(id(best_point))
            confidence = confidence_from_score(score)
            selected = match_type in {"Yes", "Coordinate Match"}
            results.append(NpcMatchResult(best_point, npc, match_type, confidence, score, dist, selected))

        for point in point_candidates:
            if id(point) not in used_point_ids:
                results.append(NpcMatchResult(point, None, "Map only", "", 0.0, None, False))

        type_order = {"Yes": 0, "Coordinate Match": 1, "Possible": 2, "Map only": 5, "NPC only": 6, "No": 7}
        results.sort(key=lambda result: (type_order.get(result.match_type, 9), -(result.score or 0), result.map_record.label if result.map_record else result.npc_row.npc_label if result.npc_row else ""))
        self.npc_match_results = results
        self.side_panel.rebuild_npc_match_table(results)

        matched = sum(1 for result in results if result.map_record is not None and result.npc_row is not None)
        selected = sum(1 for result in results if result.selected)
        map_only = sum(1 for result in results if result.map_record is not None and result.npc_row is None)
        npc_only = sum(1 for result in results if result.map_record is None and result.npc_row is not None)
        yes_count = sum(1 for result in results if result.match_type == "Yes")
        coord_count = sum(1 for result in results if result.match_type == "Coordinate Match")
        possible_count = sum(1 for result in results if result.match_type == "Possible")
        summary = (
            f"Zone: {zone}\n"
            f"Map points: {len(point_candidates):,}    NPC rows: {len(npc_candidates):,}\n"
            f"Yes: {yes_count:,}    Coordinate Match: {coord_count:,}    Possible: {possible_count:,}\n"
            f"Matched: {matched:,}    Selected: {selected:,}    Map only: {map_only:,}    NPC only: {npc_only:,}"
        )
        self.npc_match_base_summary = summary
        self.side_panel.npc_summary_label.setText(summary)
        self.side_panel.apply_npc_match_filters()
        self.status_label.setText(f"NPC comparison complete for {zone}: {matched:,} matched row(s).")
        self.show_side_tool("NPC Match")

    def clear_npc_match_preview(self) -> None:
        for item in list(getattr(self, "npc_preview_items", [])):
            try:
                self.scene.removeItem(item)
            except RuntimeError:
                pass
        self.npc_preview_items = []
        self.status_label.setText("NPC match preview cleared.")

    def preview_selected_npc_matches(self) -> None:
        self.clear_npc_match_preview()
        results = [result for result in self.side_panel.checked_npc_match_results() if result.map_record is not None and result.npc_row is not None]
        if not results:
            self.status_label.setText("No selected matched rows to preview.")
            return
        preview_pen = QPen(QColor(255, 180, 0), 1.2)
        preview_brush = QBrush(QColor(255, 180, 0, 180))
        for result in results:
            point = result.map_record
            npc = result.npc_row
            old_scene = self.mapper.map_to_scene(point.x, point.y)
            new_scene = self.mapper.map_to_scene(npc.x, npc.y)
            line = QGraphicsLineItem(old_scene.x(), old_scene.y(), new_scene.x(), new_scene.y())
            line.setPen(preview_pen)
            line.setZValue(9000)
            self.scene.addItem(line)
            marker = QGraphicsEllipseItem(-5, -5, 10, 10)
            marker.setPos(new_scene)
            marker.setPen(preview_pen)
            marker.setBrush(preview_brush)
            marker.setZValue(9001)
            marker.setToolTip(f"Preview: {npc.npc_label}")
            self.scene.addItem(marker)
            text = QGraphicsTextItem(npc.npc_label.replace("_", " "))
            text.setDefaultTextColor(QColor(255, 180, 0))
            text.setFont(QFont("Arial", 7, QFont.Bold))
            text.setPos(new_scene.x() + 7, new_scene.y() - 7)
            text.setZValue(9002)
            self.scene.addItem(text)
            self.npc_preview_items.extend([line, marker, text])
        self.status_label.setText(f"Previewing {len(results):,} selected NPC swap(s).")

    def apply_selected_npc_matches(self) -> None:
        results = [result for result in self.side_panel.checked_npc_match_results() if result.map_record is not None and result.npc_row is not None]
        if not results:
            self.status_label.setText("No selected matched rows to apply.")
            return
        if not self.confirm_bulk_action(
            "Apply NPC Match Swaps",
            f"Apply label + XYZ from NPC data to {len(results)} selected map point(s)?\n\nRGB, size, source file, and layer are preserved. The map will be marked unsaved until you click Save Edits.",
            len(results),
            action_type="edit",
        ):
            return
        before = [snapshot_point(result.map_record) for result in results]
        records = [result.map_record for result in results]
        for result in results:
            point = result.map_record
            npc = result.npc_row
            point.label = npc.npc_label
            point.x = npc.x
            point.y = npc.y
            point.z = npc.z
            point.dirty = True
        after = [snapshot_point(record) for record in records]
        self.undo_stack.append(BulkEditCommand("NPC Match & Swap", records, before, after))
        self.redo_stack.clear()
        self.clear_npc_match_preview()
        self.render_map(keep_view=True)
        self.update_dirty_indicator()
        self.status_label.setText(f"Applied NPC label + XYZ to {len(records):,} point(s). Save Edits when ready.")

    def delete_selected_npc_match_points_from_map(self) -> None:
        results = self.side_panel.checked_npc_match_results()
        records: list[MapPointRecord] = []
        seen: set[int] = set()
        for result in results:
            point = result.map_record
            if point is None or getattr(point, "deleted", False):
                continue
            if id(point) in seen:
                continue
            records.append(point)
            seen.add(id(point))

        if not records:
            self.status_label.setText("No selected NPC-match rows that are present on the map to delete.")
            return

        if not self.confirm_bulk_action(
            "Delete Selected NPC Match Labels",
            f"Mark {len(records)} selected map label(s) from the NPC Match table for deletion?\n\n"
            "Only rows that are already present on the loaded map are included. NPC-only rows are ignored.\n\n"
            "Save Edits will remove them from the map text files.",
            len(records),
            action_type="delete",
        ):
            return

        before = [snapshot_point(point) | {"deleted": getattr(point, "deleted", False)} for point in records]
        for point in records:
            point.deleted = True
            point.dirty = True
        after = [snapshot_point(point) | {"deleted": getattr(point, "deleted", False)} for point in records]

        self.undo_stack.append(BulkEditCommand("NPC Match delete labels", records, before, after))
        self.redo_stack.clear()
        self.clear_npc_match_preview()
        self.render_map(keep_view=True)
        self.side_panel.rebuild_layers()
        self.side_panel.rebuild_missing_style_match_combo()
        self.update_dirty_indicator()
        self.status_label.setText(f"Marked {len(records):,} NPC-match map label(s) for deletion. Save Edits when ready.")

    def save_current_expansion_setting(self) -> None:
        self.current_expansion_number_setting = self.side_panel.current_expansion_number()
        self.save_settings()
        self.status_label.setText(f"Saved current expansion setting: {self.side_panel.current_expansion_name()} ({self.current_expansion_number_setting}).")

    def _best_npc_match_for_point(self, point: MapPointRecord, npc_candidates: list[NpcDataRow]) -> Optional[tuple[float, str, Optional[float], NpcDataRow]]:
        best: Optional[tuple[float, str, Optional[float], NpcDataRow]] = None
        for npc in npc_candidates:
            score, match_type, dist = score_point_to_npc(point, npc)
            if match_type not in {"Yes", "Coordinate Match", "Possible"}:
                continue
            if best is None or score > best[0] or (score == best[0] and (dist or math.inf) < (best[2] or math.inf)):
                best = (score, match_type, dist, npc)
        return best

    def scan_current_zone_for_era_cleanup(self) -> None:
        if not self.loaded_files or not self.loaded_map.points:
            QMessageBox.information(self, "Era Cleanup", "Load a zone map before scanning for expansion-era labels.")
            return
        if not self.npc_data_rows:
            if self.npc_data_source_path:
                self.reload_npc_data_source()
            else:
                self.choose_npc_data_source()
            if not self.npc_data_rows:
                return

        selected_expansion = self.side_panel.current_expansion_number()
        selected_name = self.side_panel.current_expansion_name()
        zone = self.current_loaded_zone_shortname()
        npc_candidates = [row for row in self.npc_data_rows if row.zone_shortname == zone]
        if not npc_candidates:
            QMessageBox.information(self, "Era Cleanup", f"No NPC rows found for zone '{zone}'.")
            return

        results: list[EraCleanupResult] = []
        checked_points = [point for point in self.loaded_map.points if not getattr(point, "deleted", False)]
        matched_count = 0
        unknown_count = 0
        safe_count = 0
        for point in checked_points:
            best = self._best_npc_match_for_point(point, npc_candidates)
            if best is None:
                continue
            score, match_type, dist, npc = best
            matched_count += 1
            min_num = npc.min_expansion_number
            max_num = npc.max_expansion_number
            status = "Available"
            flagged = False
            if min_num is None or min_num < 0:
                status = "Unknown min expansion"
                unknown_count += 1
            elif min_num > selected_expansion:
                status = "Too new"
                flagged = True
            elif max_num is not None and max_num >= 0 and max_num < selected_expansion:
                status = "Expired / removed"
                flagged = True
            else:
                safe_count += 1
            if flagged:
                results.append(EraCleanupResult(
                    map_record=point,
                    npc_row=npc,
                    era_status=status,
                    match_type=match_type,
                    confidence=confidence_from_score(score),
                    score=score,
                    distance=dist,
                    selected=True,
                ))

        results.sort(key=lambda result: (0 if result.era_status == "Too new" else 1, result.npc_row.min_expansion_number if result.npc_row.min_expansion_number is not None else 999, result.map_record.label))
        self.era_cleanup_results = results
        self.side_panel.rebuild_era_cleanup_table(results)
        summary = (
            f"Zone: {zone}   Selected era: {selected_name} ({selected_expansion})\n"
            f"Map labels checked: {len(checked_points):,}   Matched to NPC data: {matched_count:,}\n"
            f"Flagged for review: {len(results):,}   Safe/current-era: {safe_count:,}   Unknown min expansion: {unknown_count:,}\n"
            "Flags include min expansion greater than selected era and max expansion earlier than selected era."
        )
        self.side_panel.era_summary_label.setText(summary)
        self.status_label.setText(f"Era cleanup scan complete for {zone}: {len(results):,} label(s) flagged.")
        self.show_side_tool("NPC Match")

    def clear_era_cleanup_preview(self) -> None:
        for item in list(getattr(self, "era_preview_items", [])):
            try:
                self.scene.removeItem(item)
            except RuntimeError:
                pass
        self.era_preview_items = []
        self.status_label.setText("Era cleanup preview cleared.")

    def preview_selected_era_cleanup(self) -> None:
        self.clear_era_cleanup_preview()
        results = self.side_panel.checked_era_cleanup_results()
        if not results:
            self.status_label.setText("No selected era cleanup rows to preview.")
            return
        preview_pen = QPen(QColor(255, 80, 80), 1.8)
        preview_brush = QBrush(QColor(255, 80, 80, 170))
        for result in results:
            point = result.map_record
            scene_pos = self.mapper.map_to_scene(point.x, point.y)
            marker = QGraphicsEllipseItem(-7, -7, 14, 14)
            marker.setPos(scene_pos)
            marker.setPen(preview_pen)
            marker.setBrush(preview_brush)
            marker.setZValue(9101)
            marker.setToolTip(f"Era cleanup: {point.label} - {result.era_status}")
            self.scene.addItem(marker)
            text = QGraphicsTextItem(f"REMOVE: {point.label.replace('_', ' ')}")
            text.setDefaultTextColor(QColor(255, 80, 80))
            text.setFont(QFont("Arial", 7, QFont.Bold))
            text.setPos(scene_pos.x() + 8, scene_pos.y() - 8)
            text.setZValue(9102)
            self.scene.addItem(text)
            self.era_preview_items.extend([marker, text])
        self.status_label.setText(f"Previewing {len(results):,} era cleanup removal(s).")

    def remove_selected_era_cleanup_labels(self) -> None:
        results = self.side_panel.checked_era_cleanup_results()
        records: list[MapPointRecord] = []
        seen: set[int] = set()
        for result in results:
            if id(result.map_record) not in seen and not getattr(result.map_record, "deleted", False):
                records.append(result.map_record)
                seen.add(id(result.map_record))
        if not records:
            self.status_label.setText("No selected era cleanup labels to remove.")
            return
        selected_expansion = self.side_panel.current_expansion_number()
        selected_name = self.side_panel.current_expansion_name()
        if not self.confirm_bulk_action(
            "Remove Future/Expired Era Labels",
            f"Mark {len(records)} selected map label(s) for deletion?\n\nSelected era: {selected_name} ({selected_expansion})\nThis removes labels whose matched NPC min expansion is later than your selected era, or whose max expansion is earlier than your selected era.\n\nSave Edits will remove them from the map text files.",
            len(records),
            action_type="delete",
        ):
            return
        before = [snapshot_point(point) | {"deleted": getattr(point, "deleted", False)} for point in records]
        for point in records:
            point.deleted = True
            point.dirty = True
        after = [snapshot_point(point) | {"deleted": getattr(point, "deleted", False)} for point in records]
        self.undo_stack.append(BulkEditCommand("Era cleanup remove labels", records, before, after))
        self.redo_stack.clear()
        self.clear_era_cleanup_preview()
        self.render_map(keep_view=True)
        self.side_panel.rebuild_layers()
        self.update_dirty_indicator()
        self.status_label.setText(f"Marked {len(records):,} era cleanup label(s) for deletion. Save Edits when ready.")

    def _used_npc_indices_for_current_map(self, points: list[MapPointRecord], npc_candidates: list[NpcDataRow]) -> set[int]:
        scored: list[tuple[float, int, int]] = []
        for point_index, point in enumerate(points):
            for npc_index, npc in enumerate(npc_candidates):
                score, match_type, _dist = score_point_to_npc(point, npc)
                if match_type in {"Yes", "Coordinate Match", "Possible"}:
                    scored.append((score, point_index, npc_index))
        scored.sort(key=lambda item: item[0], reverse=True)

        used_points: set[int] = set()
        used_npcs: set[int] = set()
        for score, point_index, npc_index in scored:
            if point_index in used_points or npc_index in used_npcs:
                continue
            used_points.add(point_index)
            used_npcs.add(npc_index)
        return used_npcs

    def find_missing_npcs_for_current_era(self) -> None:
        if not self.loaded_files or self.loaded_map is None:
            QMessageBox.information(self, "Add Missing NPCs", "Load a zone map before finding missing NPCs.")
            return
        if not self.npc_data_rows:
            if self.npc_data_source_path:
                self.reload_npc_data_source()
            else:
                self.choose_npc_data_source()
            if not self.npc_data_rows:
                return

        selected_expansion = self.side_panel.current_expansion_number()
        selected_name = self.side_panel.current_expansion_name()
        zone = self.current_loaded_zone_shortname()
        zone_rows = [row for row in self.npc_data_rows if row.zone_shortname == zone]
        if not zone_rows:
            QMessageBox.information(self, "Add Missing NPCs", f"No NPC rows found for zone '{zone}'.")
            return

        valid_rows = [row for row in zone_rows if npc_is_valid_for_expansion(row, selected_expansion)]
        current_points = [point for point in self.loaded_map.points if not getattr(point, "deleted", False)]
        used_valid_indices = self._used_npc_indices_for_current_map(current_points, valid_rows)

        missing_results: list[MissingNpcResult] = []
        for npc_index, npc in enumerate(valid_rows):
            if npc_index in used_valid_indices:
                continue
            missing_results.append(MissingNpcResult(npc_row=npc, npc_label=npc.npc_label, selected=False))

        missing_results.sort(key=lambda result: (
            result.npc_row.min_expansion_number if result.npc_row.min_expansion_number is not None and result.npc_row.min_expansion_number >= 0 else -1,
            result.npc_label.lower(),
            result.npc_row.source_index,
        ))
        self.missing_npc_results = missing_results
        self.side_panel.rebuild_missing_npc_table(missing_results)
        summary = (
            f"Zone: {zone}   Selected era: {selected_name} ({selected_expansion})\n"
            f"NPC rows in zone: {len(zone_rows):,}   Valid for era: {len(valid_rows):,}\n"
            f"Already on map: {len(used_valid_indices):,}   Missing/addable: {len(missing_results):,}\n"
            "Rows are not checked by default; edit labels in the table, then check the NPCs to add."
        )
        self.side_panel.missing_summary_label.setText(summary)
        self.status_label.setText(f"Missing NPC scan complete for {zone}: {len(missing_results):,} addable NPC(s).")
        self.show_side_tool("NPC Match")

    def clear_missing_npc_preview(self) -> None:
        for item in list(getattr(self, "missing_preview_items", [])):
            try:
                self.scene.removeItem(item)
            except RuntimeError:
                pass
        self.missing_preview_items = []
        self.status_label.setText("Missing NPC preview cleared.")

    def preview_selected_missing_npcs(self) -> None:
        self.clear_missing_npc_preview()
        results = self.side_panel.checked_missing_npc_results()
        if not results:
            self.status_label.setText("No selected missing NPC rows to preview.")
            return
        r = self.side_panel.missing_r_spin.value()
        g = self.side_panel.missing_g_spin.value()
        b = self.side_panel.missing_b_spin.value()
        preview_colour = QColor(r, g, b, 185)
        preview_pen = QPen(QColor(r, g, b), 1.8)
        preview_brush = QBrush(preview_colour)
        for result in results:
            npc = result.npc_row
            scene_pos = self.mapper.map_to_scene(npc.x, npc.y)
            marker = QGraphicsEllipseItem(-7, -7, 14, 14)
            marker.setPos(scene_pos)
            marker.setPen(preview_pen)
            marker.setBrush(preview_brush)
            marker.setZValue(9201)
            marker.setToolTip(f"Add NPC: {result.npc_label}")
            self.scene.addItem(marker)
            text = QGraphicsTextItem(f"ADD: {result.npc_label.replace('_', ' ')}")
            text.setDefaultTextColor(QColor(r, g, b))
            text.setFont(QFont("Arial", 7, QFont.Bold))
            text.setPos(scene_pos.x() + 8, scene_pos.y() - 8)
            text.setZValue(9202)
            self.scene.addItem(text)
            self.missing_preview_items.extend([marker, text])
        self.status_label.setText(f"Previewing {len(results):,} missing NPC addition(s).")

    def add_selected_missing_npcs_to_map(self) -> None:
        results = self.side_panel.checked_missing_npc_results()
        if not results:
            self.status_label.setText("No selected missing NPC rows to add.")
            return
        active_file = self.active_file_for_new_records()
        if active_file is None:
            QMessageBox.information(self, "Add Missing NPCs", "No active map file is available for new NPC labels.")
            return
        selected_expansion = self.side_panel.current_expansion_number()
        selected_name = self.side_panel.current_expansion_name()
        if not self.confirm_bulk_action(
            "Add Missing NPC Labels",
            f"Add {len(results)} selected NPC label(s) to {active_file.name}?\n\nSelected era: {selected_name} ({selected_expansion})\nThe new points will be appended to the active map file when you click Save Edits.",
            len(results),
            action_type="edit",
        ):
            return

        r = self.side_panel.missing_r_spin.value()
        g = self.side_panel.missing_g_spin.value()
        b = self.side_panel.missing_b_spin.value()
        size = self.side_panel.missing_size_spin.value()
        new_points: list[MapPointRecord] = []
        seen_labels: set[str] = set()
        for result in results:
            label = result.npc_label.strip() or result.npc_row.npc_label
            key = normalize_label(label)
            if key in seen_labels:
                continue
            seen_labels.add(key)
            npc = result.npc_row
            new_points.append(MapPointRecord(
                file_path=active_file,
                line_index=-1,
                raw_text="",
                x=npc.x,
                y=npc.y,
                z=npc.z,
                r=r,
                g=g,
                b=b,
                size=size,
                label=label,
                dirty=True,
            ))

        if not new_points:
            self.status_label.setText("No unique selected missing NPC labels to add.")
            return
        self.loaded_map.points.extend(new_points)
        self.undo_stack.append(AddRecordsCommand("Add missing NPC labels", self, new_points, "points"))
        self.redo_stack.clear()
        self.clear_missing_npc_preview()
        self.render_map(keep_view=True)
        self.side_panel.rebuild_layers()
        self.update_dirty_indicator()
        self.side_panel.rebuild_missing_style_match_combo()
        self.status_label.setText(f"Added {len(new_points):,} missing NPC label(s) to {active_file.name}. Save Edits when ready.")

    def choose_map_folder(self) -> None:
        start_dir = getattr(self, "map_folder", "") or (str(self.loaded_files[0].parent) if self.loaded_files else "")
        folder = QFileDialog.getExistingDirectory(self, "Choose EQ Maps Folder", start_dir)
        if not folder:
            return
        self.map_folder = folder
        self.side_panel.map_folder_edit.setText(folder)
        self.rebuild_zone_list()
        self.save_settings()

    def available_zone_shortnames(self) -> list[str]:
        folder = getattr(self, "map_folder", "") or self.side_panel.map_folder_edit.text().strip()
        if not folder:
            return []
        path = Path(folder)
        if not path.exists() or not path.is_dir():
            return []

        shortnames = set()
        for file_path in path.glob("*.txt"):
            stem = file_path.stem
            shortname = stem.split("_", 1)[0]
            if shortname:
                shortnames.add(shortname.lower())
        return sorted(shortnames)

    def zone_display_name(self, shortname: str) -> str:
        full = ZONE_SHORTNAME_TO_FULLNAME.get(shortname.lower(), shortname)
        return f"{full} ({shortname})"

    def zone_files_for_shortname(self, shortname: str) -> list[Path]:
        folder = Path(getattr(self, "map_folder", "") or self.side_panel.map_folder_edit.text().strip())
        if not folder.exists():
            return []
        files = []
        base = folder / f"{shortname}.txt"
        if base.exists():
            files.append(base)
        for idx in range(1, 5):
            layer = folder / f"{shortname}_{idx}.txt"
            if layer.exists():
                files.append(layer)
        # Include any other matching layers without duplicate.
        for file_path in sorted(folder.glob(f"{shortname}_*.txt")):
            if file_path not in files:
                files.append(file_path)
        return files

    def rebuild_zone_list(self) -> None:
        if not hasattr(self.side_panel, "zones_list"):
            return
        search = self.side_panel.zone_search_edit.text().strip().lower()
        self.side_panel.zones_list.clear()

        for shortname in self.available_zone_shortnames():
            display = self.zone_display_name(shortname)
            if search and search not in display.lower() and search not in shortname.lower():
                continue
            item = QListWidgetItem(display)
            item.setData(Qt.UserRole, shortname)
            self.side_panel.zones_list.addItem(item)

        self.status_label.setText(f"Found {self.side_panel.zones_list.count()} zone(s) in map folder.")

    def reset_zone_search(self) -> None:
        self.side_panel.zone_search_edit.clear()
        self.rebuild_zone_list()

    def open_selected_zone(self) -> None:
        item = self.side_panel.zones_list.currentItem()
        if item is None:
            return
        shortname = item.data(Qt.UserRole)
        files = self.zone_files_for_shortname(shortname)
        if not files:
            QMessageBox.warning(self, "No Map Files", f"No map files found for {shortname}.")
            return
        if not self.confirm_discard_unsaved():
            return
        self.load_files(files)
        self.status_label.setText(f"Opened {self.zone_display_name(shortname)}.")



    def set_edit_mode_from_overlay(self, mode_name: str) -> None:
        if hasattr(self.side_panel, "edit_mode_combo"):
            self.side_panel.edit_mode_combo.setCurrentText(mode_name)
        if hasattr(self, "canvas_controls"):
            self.canvas_controls.set_active_mode(mode_name)

    def zoom_in(self) -> None:
        self.view.scale(1.2, 1.2)
        self.update_canvas_overlays()

    def zoom_out(self) -> None:
        self.view.scale(1 / 1.2, 1 / 1.2)
        self.update_canvas_overlays()

    def update_canvas_cursor_position(self, scene_point: QPointF) -> None:
        x, y = self.mapper.scene_to_map(scene_point)
        self.canvas_cursor_text = f"X: {x:.2f}   Y: {y:.2f}   Z: 0.00"
        self.update_canvas_overlays()

    def update_canvas_overlays(self) -> None:
        if not hasattr(self, "canvas_container"):
            return
        container_rect = self.canvas_container.rect()
        if container_rect.width() <= 0 or container_rect.height() <= 0:
            return

        if hasattr(self, "canvas_controls"):
            self.canvas_controls.setGeometry(14, container_rect.height() - 82, min(1120, max(840, container_rect.width() - 210)), 66)
            zoom = int(round(self.view.transform().m11() * 100))
            cursor_text = getattr(self, "canvas_cursor_text", "X: 0.00   Y: 0.00   Z: 0.00")
            self.canvas_controls.update_status(cursor_text, f"Zoom: {zoom}%")
            self.canvas_controls.raise_()

        if hasattr(self, "minimap_overlay"):
            width, height = 160, 132
            self.minimap_overlay.setGeometry(container_rect.width() - width - 18, container_rect.height() - height - 18, width, height)
            self.minimap_overlay.update()
            self.minimap_overlay.raise_()

    def show_side_tool(self, tab_name: str) -> None:
        if not hasattr(self.side_panel, "tabs"):
            return
        for index in range(self.side_panel.tabs.count()):
            if self.side_panel.tabs.tabText(index) == tab_name:
                self.side_panel.tabs.setCurrentIndex(index)
                sizes = self.splitter.sizes() if hasattr(self, "splitter") else []
                if len(sizes) >= 2 and sizes[1] < 120:
                    total = max(sum(sizes), 1)
                    self.splitter.setSizes([max(total - 440, 300), 440])
                self.status_label.setText(f"Opened {tab_name} tool.")
                return

    def toggle_sidebar(self) -> None:
        if not hasattr(self, "splitter"):
            return
        sizes = self.splitter.sizes()
        if len(sizes) < 2:
            return
        total = max(sum(sizes), 1)
        if sizes[1] < 80:
            self.splitter.setSizes([max(total - 440, 300), 440])
        else:
            self.splitter.setSizes([total, 0])
        self.update_canvas_overlays()
        self.save_settings()

    def confirm_exit(self) -> None:
        self.save_settings()
        self.close()

    def add_undo(self, label: str, record: Any, before: dict[str, Any], after: dict[str, Any]) -> None:
        if before == after:
            return
        self.undo_stack.append(UndoCommand(label, record, before.copy(), after.copy()))
        self.redo_stack.clear()
        if len(self.undo_stack) > 100:
            self.undo_stack.pop(0)
        self.update_dirty_indicator()

    def undo(self) -> None:
        if not self.undo_stack:
            self.status_label.setText("Nothing to undo.")
            return
        command = self.undo_stack.pop()
        command.undo()
        self.redo_stack.append(command)
        if isinstance(command, BulkEditCommand):
            self.render_map(keep_view=True)
            self.side_panel.rebuild_layers()
            self.update_dirty_indicator()
        elif hasattr(command, "record") and command.record in (self.loaded_map.lines + self.loaded_map.points):
            self.update_record_items(command.record)
            self.side_panel.set_record(command.record)
        self.status_label.setText(f"Undo: {command.label}")

    def redo(self) -> None:
        if not self.redo_stack:
            self.status_label.setText("Nothing to redo.")
            return
        command = self.redo_stack.pop()
        command.redo()
        self.undo_stack.append(command)
        if isinstance(command, BulkEditCommand):
            self.render_map(keep_view=True)
            self.side_panel.rebuild_layers()
            self.update_dirty_indicator()
        elif hasattr(command, "record") and command.record in (self.loaded_map.lines + self.loaded_map.points):
            self.update_record_items(command.record)
            self.side_panel.set_record(command.record)
        self.status_label.setText(f"Redo: {command.label}")

    def dirty_records(self) -> list[Any]:
        return [record for record in self.loaded_map.lines + self.loaded_map.points if record.dirty]

    def dirty_files(self) -> list[Path]:
        return sorted({record.file_path for record in self.dirty_records()})

    def update_dirty_indicator(self) -> None:
        if hasattr(self, "side_panel"):
            try:
                self.side_panel.rebuild_colour_list()
                self.side_panel.rebuild_points_list()
                self.side_panel.rebuild_pending_changes()
                if hasattr(self, "nav_panel"):
                    self.nav_panel.refresh()
            except Exception:
                pass
        dirty_files = self.dirty_files()
        if dirty_files:
            names = ", ".join(path.name for path in dirty_files)
            self.dirty_label.setText(f"Unsaved: {names}")
            self.setWindowTitle(f"EQ Map Editor *")
        else:
            self.dirty_label.setText("Clean")
            self.setWindowTitle(f"EQ Map Editor")


    def externally_modified_files(self, files: Optional[list[Path]] = None) -> list[Path]:
        candidates = files or self.loaded_files
        modified: list[Path] = []
        for file_path in candidates:
            try:
                loaded_mtime = self.loaded_file_mtimes.get(file_path)
                if loaded_mtime is None or not file_path.exists():
                    continue
                current_mtime = file_path.stat().st_mtime
                if current_mtime > loaded_mtime + 0.01:
                    modified.append(file_path)
            except Exception:
                continue
        return modified

    def confirm_external_overwrite(self, files: Optional[list[Path]] = None) -> bool:
        modified = self.externally_modified_files(files)
        if not modified:
            return True
        file_list = "\n".join(f"- {path.name}" for path in modified[:12])
        if len(modified) > 12:
            file_list += f"\n...and {len(modified) - 12} more"
        result = QMessageBox.warning(
            self,
            "Files Changed Outside EQ Map Editor",
            "One or more map files were changed on disk after they were loaded.\n\n"
            f"{file_list}\n\n"
            "Saving now may overwrite those outside changes. Continue saving anyway?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        return result == QMessageBox.Yes

    def save_edits(self) -> None:
        dirty_records = self.dirty_records()
        if not dirty_records:
            QMessageBox.information(self, "No Changes", "There are no edited records to save.")
            return

        by_file: dict[Path, list[Any]] = {}
        for record in dirty_records:
            by_file.setdefault(record.file_path, []).append(record)

        if not self.confirm_external_overwrite(list(by_file.keys())):
            self.status_label.setText("Save cancelled because files changed externally.")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        saved_files: list[str] = []

        try:
            for file_path, records in by_file.items():
                original_lines = file_path.read_text(encoding="utf-8", errors="replace").splitlines() if file_path.exists() else []
                backup_path = BACKUPS_DIR / f"{file_path.name}.{timestamp}.bak"
                if file_path.exists():
                    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(file_path, backup_path)

                append_records = []
                for record in records:
                    if record.line_index < 0:
                        append_records.append(record)
                        continue
                    if record.line_index >= len(original_lines):
                        raise IndexError(f"Line index out of range for {record.file_path.name}:{record.line_index + 1}")
                    if getattr(record, "deleted", False):
                        original_lines[record.line_index] = None
                    else:
                        original_lines[record.line_index] = record.to_map_line()

                original_lines = [line for line in original_lines if line is not None]

                for record in append_records:
                    if getattr(record, "deleted", False):
                        continue
                    record.line_index = len(original_lines)
                    original_lines.append(record.to_map_line())

                file_path.write_text("\n".join(original_lines) + "\n", encoding="utf-8")
                saved_files.append(f"{file_path.name} (backup: {backup_path.name if backup_path.exists() else 'none'})")

            for record in dirty_records:
                record.dirty = False
                record.raw_text = "" if getattr(record, "deleted", False) else record.to_map_line()

            self.loaded_map.lines = [record for record in self.loaded_map.lines if not getattr(record, "deleted", False)]
            self.loaded_map.points = [record for record in self.loaded_map.points if not getattr(record, "deleted", False)]

        except Exception as exc:
            QMessageBox.critical(self, "Save Error", f"Could not save edits:\n{exc}")
            return

        self.undo_stack.clear()
        self.redo_stack.clear()
        self.side_panel.rebuild_layers()
        self.side_panel.rebuild_pending_changes()
        self.update_dirty_indicator()
        self.log_event(f"Saved {len(dirty_records)} edited record(s): " + ", ".join(saved_files))
        QMessageBox.information(self, "Saved", "Saved edits:\n\n" + "\n".join(saved_files))
        self.status_label.setText(f"Saved {len(dirty_records)} edited record(s).")

    def confirm_discard_unsaved(self) -> bool:
        if not self.dirty_records():
            return True

        result = QMessageBox.question(
            self,
            "Unsaved Changes",
            "You have unsaved map edits. Continue and discard them?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        return result == QMessageBox.Yes

    def reload_from_disk(self) -> None:
        if not self.loaded_files:
            return
        if not self.confirm_discard_unsaved():
            return
        self.load_files(self.loaded_files)

    def restore_from_backup(self) -> None:
        if not self.loaded_files:
            QMessageBox.information(self, "No Files", "Open a map file before restoring from backup.")
            return
        if not self.confirm_discard_unsaved():
            return

        backup_name, _ = QFileDialog.getOpenFileName(
            self,
            "Choose Backup File to Restore",
            str(BACKUPS_DIR),
            "Backup files (*.bak);;All files (*.*)",
        )
        if not backup_name:
            return

        backup_path = Path(backup_name)

        target_candidates = []
        backup_name_text = backup_path.name
        for source in self.loaded_files:
            if backup_name_text.startswith(source.name + ".") or backup_name_text == source.name + ".bak":
                target_candidates.append(source)

        if len(target_candidates) == 1:
            target_path = target_candidates[0]
        else:
            target_name, _ = QFileDialog.getOpenFileName(
                self,
                "Choose Source File to Overwrite with Backup",
                str(backup_path.parent),
                "Text files (*.txt);;All files (*.*)",
            )
            if not target_name:
                return
            target_path = Path(target_name)

        result = QMessageBox.question(
            self,
            "Restore Backup",
            f"Restore:\n{backup_path.name}\n\nover:\n{target_path.name}?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if result != QMessageBox.Yes:
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pre_restore_backup = BACKUPS_DIR / f"{target_path.name}.pre_restore_{timestamp}.bak"
        BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
        shutil.copy2(target_path, pre_restore_backup)
        shutil.copy2(backup_path, target_path)

        QMessageBox.information(
            self,
            "Restored",
            f"Restored {backup_path.name} over {target_path.name}.\n\n"
            f"Previous current file was backed up as:\n{pre_restore_backup.name}",
        )
        self.reload_from_disk()

    def run_deferred_startup_fit(self) -> None:
        if self._fit_last_map_after_show:
            self._fit_last_map_after_show = False
            self.fit_map()
            self.status_label.setText("Restored last loaded map files and fit map to screen.")

    def showEvent(self, event) -> None:
        super().showEvent(event)
        if self._fit_last_map_after_show:
            QTimer.singleShot(0, self.run_deferred_startup_fit)
            QTimer.singleShot(150, self.run_deferred_startup_fit)

    def closeEvent(self, event) -> None:
        if self.dirty_records():
            result = QMessageBox.question(
                self,
                "Unsaved Changes",
                "You have unsaved map edits. Close anyway?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if result != QMessageBox.Yes:
                event.ignore()
                return
        if self.autosave_settings_on_exit:
            self.save_settings()
        event.accept()


def expand_input_paths(paths: list[str]) -> list[Path]:
    expanded: list[Path] = []
    for raw_path in paths:
        path = Path(raw_path)
        if path.is_dir():
            expanded.extend(sorted(path.glob("*.txt")))
        else:
            expanded.append(path)
    return expanded


def install_exception_hook() -> None:
    def handle_exception(exc_type, exc_value, exc_traceback):
        try:
            LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
            with LOG_PATH.open("a", encoding="utf-8") as log_file:
                log_file.write("\n--- Unhandled exception ---\n")
                traceback.print_exception(exc_type, exc_value, exc_traceback, file=log_file)
        except Exception:
            pass

        try:
            QMessageBox.critical(
                None,
                "EQ Map Editor Error",
                f"An unexpected error occurred:\n\n{exc_value}\n\nA log was written to:\n{LOG_PATH}",
            )
        except Exception:
            pass

        sys.__excepthook__(exc_type, exc_value, exc_traceback)

    sys.excepthook = handle_exception


def main() -> int:
    parser = argparse.ArgumentParser(description=f"EQ Map Editor {VERSION}")
    parser.add_argument("paths", nargs="*", help="One or more EQ map .txt files, or a folder containing .txt files.")
    args = parser.parse_args()

    set_windows_app_user_model_id()
    app = QApplication(sys.argv)
    icon = app_icon()
    if not icon.isNull():
        app.setWindowIcon(icon)
    app.setStyle("Fusion")
    install_exception_hook()
    initial_files = expand_input_paths(args.paths) if args.paths else []
    window = EqMapMainWindow(initial_files=initial_files)
    window.resize(1650, 950)
    window.show()
    QTimer.singleShot(0, window.run_deferred_startup_fit)
    QTimer.singleShot(250, window.run_deferred_startup_fit)
    QTimer.singleShot(300, window.maybe_show_welcome)
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
