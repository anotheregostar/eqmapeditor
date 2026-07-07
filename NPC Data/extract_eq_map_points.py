import csv
import math
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


NPC_LIST_FILENAME = "EQMapEditor NPC Label List.csv"
COORDINATE_MATCH_DISTANCE = 20


def make_safe_filename(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "_", name)


def parse_float(value):
    try:
        return float(value)
    except Exception:
        return None


def format_float(value):
    if value is None or value == math.inf:
        return ""
    return round(value, 4)


def normalize_text(value: str) -> str:
    if value is None:
        return ""
    value = str(value).strip().lower()
    return re.sub(r"[^a-z0-9]", "", value)


def clean_npc_role(value: str) -> str:
    """Remove EQEmu NULL marker values from npc_role."""
    if value is None:
        return ""
    value = str(value).strip()
    if value in {"\\N", r"\N"}:
        return ""
    return value


def should_exclude_npc_row(npc_row) -> bool:
    """Exclude NPC rows where npc_name includes '_Fabled_'."""
    npc_name = str(npc_row.get("npc_name", "") or "")
    return "_Fabled_" in npc_name


def extract_parenthetical_role_from_map_label(label: str) -> str:
    """Extract trailing parenthetical role from a map label."""
    if not label:
        return ""
    label = str(label).strip()
    match = re.search(r"_\(([^)]*)\)$", label)
    if not match:
        return ""
    role = match.group(1).strip().replace("_", " ")
    return clean_npc_role(role)


def format_role_for_npc_map_label(role: str) -> str:
    """
    Apply requested role substitutions before building npc_map_label:
      Guildmaster -> GM
      " Supplies" -> ""
      spaces -> underscores
    """
    role = clean_npc_role(role)
    if not role:
        return ""
    role = role.replace("Guildmaster", "GM")
    role = role.replace(" Supplies", "")
    role = role.strip().replace(" ", "_")
    return role


def build_npc_map_label(npc_name: str, npc_role: str) -> str:
    """Build labels like Doggle_Pitt_(Banker). If role is blank, returns npc_name only."""
    npc_name = (npc_name or "").strip()
    role = format_role_for_npc_map_label(npc_role)
    if not npc_name:
        return ""
    if not role:
        return npc_name
    return f"{npc_name}_({role})"


def clean_map_label_for_npc_match(label: str) -> str:
    """Remove trailing map role notes like _(Banker), _(Hunter), etc."""
    if label is None:
        return ""
    label = str(label).strip()
    label = re.sub(r"(?:_\([^)]*\))+$", "", label)
    return label.strip()


def get_name_aliases(value: str):
    """
    Build alternate name forms for matching.

    Helps:
      Ceridan_(Banker)  -> Ceridan
      Banker_Ceridan    -> Ceridan
      Granger_(Banker)  -> Granger
      Banker_Granger    -> Granger
    """
    if value is None:
        return set()

    raw = str(value).strip()
    cleaned = clean_map_label_for_npc_match(raw)
    aliases = set()

    common_prefix_roles = {
        "banker", "merchant", "general", "parcels", "parcel", "guildmaster",
        "gm", "trainer", "vendor", "keeper", "priest", "high", "grandmaster",
        "master", "mistress", "lord", "lady", "warlord", "grave", "savage",
        "oracle", "phantasmist", "virtuoso", "warder", "assassin", "warlock",
        "shaman", "druid", "wizard", "enchanter", "bard", "necromancer",
        "shadow", "knight", "beastlord", "ranger", "cleric", "warrior",
        "rogue", "monk", "paladin", "magician", "berserker",
    }

    for candidate in [raw, cleaned]:
        candidate = candidate.strip()
        if not candidate:
            continue

        aliases.add(candidate)
        parts = [p for p in candidate.split("_") if p]

        if len(parts) >= 2:
            first = parts[0].lower()
            if first in common_prefix_roles:
                aliases.add("_".join(parts[1:]))
            if len(parts[-1]) >= 4:
                aliases.add(parts[-1])

    role_match = re.search(r"_\(([^)]*)\)$", raw)
    if role_match and cleaned:
        role_text = role_match.group(1).strip().split(",")[0].strip()
        if role_text:
            aliases.add(f"{role_text}_{cleaned}")

    return {alias for alias in aliases if alias}


def get_zone_shortname(file_path: Path) -> str:
    name = file_path.stem
    name = re.sub(r"_\d+$", "", name)
    return name.lower()


def map_to_db_coords(map_x, map_y, map_z):
    """Convert EQ map point coordinates to database-style coordinates: -x, -y, z."""
    x = parse_float(map_x)
    y = parse_float(map_y)
    z = parse_float(map_z)
    if x is None or y is None or z is None:
        return None, None, None
    return -x, -y, z


def npc_to_map_coords(spawn_x, spawn_y, spawn_z):
    """Convert NPC database spawn coordinates to map point-style coordinates: -x, -y, z."""
    x = parse_float(spawn_x)
    y = parse_float(spawn_y)
    z = parse_float(spawn_z)
    if x is None or y is None or z is None:
        return None, None, None
    return -x, -y, z


def distance_between_npc_and_map(npc_row, map_row):
    npc_x = parse_float(npc_row.get("spawn_x"))
    npc_y = parse_float(npc_row.get("spawn_y"))
    npc_z = parse_float(npc_row.get("spawn_z"))

    map_db_x = parse_float(map_row.get("map_db_x"))
    map_db_y = parse_float(map_row.get("map_db_y"))
    map_db_z = parse_float(map_row.get("map_db_z"))

    if None in (npc_x, npc_y, npc_z, map_db_x, map_db_y, map_db_z):
        return {
            "distance": math.inf,
            "distance_x": None,
            "distance_y": None,
            "distance_z": None,
        }

    dx = map_db_x - npc_x
    dy = map_db_y - npc_y
    dz = map_db_z - npc_z

    return {
        "distance": math.sqrt(dx ** 2 + dy ** 2 + dz ** 2),
        "distance_x": dx,
        "distance_y": dy,
        "distance_z": dz,
    }


def load_csv_rows(csv_path: Path):
    if not csv_path.exists():
        raise FileNotFoundError(f"Could not find file: {csv_path}")
    with csv_path.open("r", newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        columns = reader.fieldnames or []
    return rows, columns


def parse_p_line(line: str):
    line = line.strip()
    if not line.startswith("P "):
        return None

    data = line[2:].strip()
    parts = data.split(", ")
    if len(parts) < 8:
        return None

    x = parts[0].strip()
    y = parts[1].strip()
    z = parts[2].strip()
    r = parts[3].strip()
    g = parts[4].strip()
    b = parts[5].strip()
    size = parts[6].strip()
    label = parts[-1].strip()
    return x, y, z, r, g, b, size, label


def extract_map_points_from_folder(input_folder: Path):
    rows = []
    for txt_file in sorted(input_folder.glob("*.txt")):
        zone_shortname = get_zone_shortname(txt_file)
        with txt_file.open("r", encoding="utf-8", errors="replace") as f:
            for line in f:
                parsed = parse_p_line(line)
                if parsed is None:
                    continue

                x, y, z, r, g, b, size, label = parsed
                map_db_x, map_db_y, map_db_z = map_to_db_coords(x, y, z)

                rows.append({
                    "map_zone_shortname": zone_shortname,
                    "map_x": x,
                    "map_y": y,
                    "map_z": z,
                    "map_db_x": format_float(map_db_x),
                    "map_db_y": format_float(map_db_y),
                    "map_db_z": format_float(map_db_z),
                    "map_r": r,
                    "map_g": g,
                    "map_b": b,
                    "map_size": size,
                    "map_label": label,
                    "map_match_label": clean_map_label_for_npc_match(label),
                    "map_source_file": txt_file.name,
                })
    return rows


def true_exact_name_match(npc_name, map_match_label):
    return normalize_text(npc_name) == normalize_text(map_match_label)


def possible_name_match(npc_name, map_match_label):
    npc_aliases = get_name_aliases(npc_name)
    map_aliases = get_name_aliases(map_match_label)

    for npc_alias in npc_aliases:
        for map_alias in map_aliases:
            npc_norm = normalize_text(npc_alias)
            map_norm = normalize_text(map_alias)
            if not npc_norm or not map_norm:
                continue
            if npc_norm == map_norm:
                return True
            if npc_norm.startswith(map_norm) and len(map_norm) >= 4:
                return True
            if map_norm.startswith(npc_norm) and len(npc_norm) >= 4:
                return True
            ratio = SequenceMatcher(None, npc_norm, map_norm).ratio()
            if ratio >= 0.88:
                return True
    return False


def choose_best_map_match(npc_row, map_candidates):
    """
    Matching priority:
      1. Exact name match -> Yes
      2. Alias/fuzzy name match -> Possible
      3. If the Possible match is also within 20 units -> Coordinate Match
      4. No name-based possible match -> No

    Coordinate Match is only assigned to rows that were already Possible.
    It does not coordinate-match unrelated nearby map labels.
    """
    npc_name = npc_row.get("npc_name", "")
    exact_matches = []
    possible_matches = []

    for map_row in map_candidates:
        map_label = map_row.get("map_match_label", "")
        if true_exact_name_match(npc_name, map_label):
            exact_matches.append(map_row)
        elif possible_name_match(npc_name, map_label):
            possible_matches.append(map_row)

    if exact_matches:
        best = min(exact_matches, key=lambda m: distance_between_npc_and_map(npc_row, m)["distance"])
        return best, "Yes"

    if possible_matches:
        best = min(possible_matches, key=lambda m: distance_between_npc_and_map(npc_row, m)["distance"])
        distance = distance_between_npc_and_map(npc_row, best)["distance"]
        if distance != math.inf and -COORDINATE_MATCH_DISTANCE < distance < COORDINATE_MATCH_DISTANCE:
            return best, "Coordinate Match"
        return best, "Possible"

    return None, "No"


def blank_map_columns():
    return {
        "map_label": "",
        "map_match_label": "",
        "map_x": "",
        "map_y": "",
        "map_z": "",
        "map_db_x": "",
        "map_db_y": "",
        "map_db_z": "",
        "map_r": "",
        "map_g": "",
        "map_b": "",
        "map_size": "",
        "map_source_file": "",
        "match_distance": "",
        "match_distance_x": "",
        "match_distance_y": "",
        "match_distance_z": "",
    }


def blank_npc_columns():
    return {
        "zone_shortname": "",
        "npc_name": "",
        "npc_role": "",
        "npc_map_label": "",
        "spawn_x": "",
        "spawn_y": "",
        "spawn_z": "",
        "npc_map_x": "",
        "npc_map_y": "",
        "npc_map_z": "",
        "is_merchant": "",
        "min_expansion_number": "",
        "min_expansion": "",
        "max_expansion_number": "",
        "max_expansion": "",
        "scripted_npc": "",
        "spawned_npc": "",
    }


def npc_columns_from_row(npc_row):
    npc_role = clean_npc_role(npc_row.get("npc_role", ""))
    npc_map_x, npc_map_y, npc_map_z = npc_to_map_coords(
        npc_row.get("spawn_x"),
        npc_row.get("spawn_y"),
        npc_row.get("spawn_z"),
    )
    npc_name = npc_row.get("npc_name", "")

    return {
        "zone_shortname": npc_row.get("zone_short_name", ""),
        "npc_name": npc_name,
        "npc_role": npc_role,
        "npc_map_label": build_npc_map_label(npc_name, npc_role),
        "spawn_x": npc_row.get("spawn_x", ""),
        "spawn_y": npc_row.get("spawn_y", ""),
        "spawn_z": npc_row.get("spawn_z", ""),
        "npc_map_x": format_float(npc_map_x),
        "npc_map_y": format_float(npc_map_y),
        "npc_map_z": format_float(npc_map_z),
        "is_merchant": npc_row.get("is_merchant", ""),
        "min_expansion_number": npc_row.get("min_expansion_number", ""),
        "min_expansion": npc_row.get("min_expansion", ""),
        "max_expansion_number": npc_row.get("max_expansion_number", ""),
        "max_expansion": npc_row.get("max_expansion", ""),
        "scripted_npc": npc_row.get("scripted_npc", ""),
        "spawned_npc": npc_row.get("spawned_npc", ""),
    }


def map_columns_from_row(map_row):
    return {
        "map_label": map_row.get("map_label", ""),
        "map_match_label": map_row.get("map_match_label", ""),
        "map_x": map_row.get("map_x", ""),
        "map_y": map_row.get("map_y", ""),
        "map_z": map_row.get("map_z", ""),
        "map_db_x": map_row.get("map_db_x", ""),
        "map_db_y": map_row.get("map_db_y", ""),
        "map_db_z": map_row.get("map_db_z", ""),
        "map_r": map_row.get("map_r", ""),
        "map_g": map_row.get("map_g", ""),
        "map_b": map_row.get("map_b", ""),
        "map_size": map_row.get("map_size", ""),
        "map_source_file": map_row.get("map_source_file", ""),
    }


def apply_matched_row_adjustments(combined):
    """
    For matched rows:
      - If Yes or Coordinate Match
      - If merchant
      - If npc_role blank
      - Copy role from map label parentheses
      - Rebuild npc_map_label
    """
    match_status = combined.get("npc_matched", "")
    is_merchant = str(combined.get("is_merchant", "")).strip().lower()
    npc_role = clean_npc_role(combined.get("npc_role", ""))

    if match_status in {"Yes", "Coordinate Match"}:
        if is_merchant == "yes" and not npc_role:
            inferred_role = extract_parenthetical_role_from_map_label(combined.get("map_label", ""))
            if inferred_role:
                combined["npc_role"] = inferred_role

    combined["npc_role"] = clean_npc_role(combined.get("npc_role", ""))
    combined["npc_map_label"] = build_npc_map_label(
        combined.get("npc_name", ""),
        combined.get("npc_role", ""),
    )
    return combined


# def should_remove_map_only_row(row):
    # """
    # Remove unwanted map-only utility/seasonal/task labels.
    # Only applies when source_status = Map only.
    # """
    # if row.get("source_status") != "Map only":
        # return False

    # map_label = str(row.get("map_label", "") or "").lower()
    # removal_text = [
        # "quests)",
        # "tasks)",
        # "task_master",
        # "mercenary_liaison",
        # "mercenary)",
        # "augs)",
        # "night_of_the_dead",
        # "nights_of_the_dead",
        # "(heroic_adventures)",
    # ]
    # return any(text.lower() in map_label for text in removal_text)


def combine_npc_and_map_data(npc_rows, map_rows):
    maps_by_zone = {}

    # Exclude Fabled NPC rows before matching or output.
    npc_rows = [row for row in npc_rows if not should_exclude_npc_row(row)]

    for map_row in map_rows:
        zone = (map_row.get("map_zone_shortname") or "").strip().lower()
        maps_by_zone.setdefault(zone, []).append(map_row)

    combined_rows = []
    matched_map_object_ids = set()

    for npc_row in npc_rows:
        zone = (npc_row.get("zone_short_name") or "").strip().lower()
        map_candidates = maps_by_zone.get(zone, [])
        best_map_row, match_status = choose_best_map_match(npc_row, map_candidates)

        combined = {
            "source_status": "NPC + Map" if best_map_row else "NPC only",
            "npc_matched": match_status,
        }
        combined.update(npc_columns_from_row(npc_row))

        if best_map_row:
            matched_map_object_ids.add(id(best_map_row))
            distance = distance_between_npc_and_map(npc_row, best_map_row)
            combined.update(map_columns_from_row(best_map_row))
            combined.update({
                "match_distance": format_float(distance["distance"]),
                "match_distance_x": format_float(distance["distance_x"]),
                "match_distance_y": format_float(distance["distance_y"]),
                "match_distance_z": format_float(distance["distance_z"]),
            })
        else:
            combined.update(blank_map_columns())

        combined = apply_matched_row_adjustments(combined)
        combined_rows.append(combined)

    for map_row in map_rows:
        if id(map_row) in matched_map_object_ids:
            continue

        combined = {
            "source_status": "Map only",
            "npc_matched": "No",
            "zone_shortname": map_row.get("map_zone_shortname", ""),
        }

        npc_blanks = blank_npc_columns()
        npc_blanks.pop("zone_shortname", None)
        combined.update(npc_blanks)
        combined.update(map_columns_from_row(map_row))
        combined.update({
            "match_distance": "",
            "match_distance_x": "",
            "match_distance_y": "",
            "match_distance_z": "",
        })
        combined_rows.append(combined)

    # Keep all Map only rows in the workbook.
    # The apply_combined_map_data.py script handles deleting cleanup labels from the map .txt files,
    # which preserves an audit trail in the apply summary CSV.
    return combined_rows


def write_xlsx(output_xlsx: Path, combined_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Map Data"

    fieldnames = [
        "source_status",
        "npc_matched",
        "zone_shortname",
        "npc_name",
        "npc_role",
        "npc_map_label",
        "spawn_x",
        "spawn_y",
        "spawn_z",
        "npc_map_x",
        "npc_map_y",
        "npc_map_z",
        "is_merchant",
        "min_expansion_number",
        "min_expansion",
        "max_expansion_number",
        "max_expansion",
        "scripted_npc",
        "spawned_npc",
        "map_label",
        "map_match_label",
        "map_x",
        "map_y",
        "map_z",
        "map_db_x",
        "map_db_y",
        "map_db_z",
        "map_r",
        "map_g",
        "map_b",
        "map_size",
        "map_source_file",
        "match_distance",
        "match_distance_x",
        "match_distance_y",
        "match_distance_z",
    ]

    ws.append(fieldnames)
    for row in combined_rows:
        ws.append([row.get(field, "") for field in fieldnames])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_fills = {
        "Yes": PatternFill("solid", fgColor="C6EFCE"),
        "Possible": PatternFill("solid", fgColor="FFEB9C"),
        "Coordinate Match": PatternFill("solid", fgColor="D9EAD3"),
        "No": PatternFill("solid", fgColor="FFC7CE"),
    }

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")
        match_cell = row[1]
        fill = status_fills.get(str(match_cell.value), None)
        if fill:
            match_cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    width_by_name = {
        "source_status": 16,
        "npc_matched": 18,
        "zone_shortname": 18,
        "npc_name": 28,
        "npc_role": 24,
        "npc_map_label": 36,
        "map_label": 36,
        "map_match_label": 28,
        "map_source_file": 24,
        "map_r": 10,
        "map_g": 10,
        "map_b": 10,
        "map_size": 10,
        "npc_map_x": 14,
        "npc_map_y": 14,
        "npc_map_z": 14,
    }

    for idx, field in enumerate(fieldnames, start=1):
        width = width_by_name.get(field, 16)
        ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(output_xlsx)


def main(input_folder: Path, output_folder: Path):
    npc_csv_path = output_folder / NPC_LIST_FILENAME
    dragged_folder_name = make_safe_filename(input_folder.name)
    output_xlsx = output_folder / f"{dragged_folder_name} Combined Map Data.xlsx"

    npc_rows, npc_columns = load_csv_rows(npc_csv_path)
    map_rows = extract_map_points_from_folder(input_folder)
    combined_rows = combine_npc_and_map_data(npc_rows, map_rows)
    write_xlsx(output_xlsx, combined_rows)

    yes_count = sum(1 for row in combined_rows if row.get("npc_matched") == "Yes")
    possible_count = sum(1 for row in combined_rows if row.get("npc_matched") == "Possible")
    coord_count = sum(1 for row in combined_rows if row.get("npc_matched") == "Coordinate Match")
    no_count = sum(1 for row in combined_rows if row.get("npc_matched") == "No")
    npc_only_count = sum(1 for row in combined_rows if row.get("source_status") == "NPC only")
    map_only_count = sum(1 for row in combined_rows if row.get("source_status") == "Map only")

    print()
    print("Done.")
    print(f"Input map folder: {input_folder}")
    print(f"NPC list used: {npc_csv_path}")
    print(f"Output workbook: {output_xlsx}")
    print()
    print(f"Combined rows: {len(combined_rows)}")
    print(f"Exact matches: {yes_count}")
    print(f"Possible matches: {possible_count}")
    print(f"Coordinate matches: {coord_count}")
    print(f"No matches: {no_count}")
    print(f"NPC-only rows: {npc_only_count}")
    print(f"Map-only rows: {map_only_count}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Drag a folder onto the batch file, or run:")
        print("python extract_eq_map_points.py C:\\Path\\To\\MapFolder")
        input("Press Enter to exit...")
        sys.exit(1)

    input_folder = Path(sys.argv[1])

    if len(sys.argv) >= 3:
        output_folder = Path(sys.argv[2])
    else:
        output_folder = Path(__file__).parent

    if not input_folder.exists() or not input_folder.is_dir():
        print(f"Invalid folder: {input_folder}")
        input("Press Enter to exit...")
        sys.exit(1)

    if not output_folder.exists() or not output_folder.is_dir():
        print(f"Invalid output folder: {output_folder}")
        input("Press Enter to exit...")
        sys.exit(1)

    main(input_folder, output_folder)

    print()
    input("Press Enter to exit...")
