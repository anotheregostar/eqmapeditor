import csv
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: This script requires openpyxl.")
    print("Install it with: pip install openpyxl")
    input("Press Enter to exit...")
    sys.exit(1)


VALID_MATCH_STATUSES = {"Yes", "Coordinate Match"}
VALID_SOURCE_STATUS = "NPC + Map"
MAP_ONLY_SOURCE_STATUS = "Map only"


# Rows whose map_label contains any of these strings are deleted.
# This applies to both Map only rows and matched NPC + Map rows.
# Matching is case-insensitive.
MAP_ONLY_DELETE_TERMS = [
    # Earlier cleanup terms
    "quests)",
    "Tasks)",
    "Mercenary_Liaison",
    "Mercenary)",
    "Augs)",
    "Night_of_the_Dead",
    "Nights_of_the_Dead",
    "(Heroic_Adventures)",

    # Expanded cleanup terms
    "(Task_Master)",
    "(Adventure_Point_Merchant)",
    "(Wayfarer_Camp_Port)",
    "(Augmentation_Distillers)",
    "(Adventure_Merchant)",
    "(Group_Adventures)",
    "(Trade_Skill_Quests)",
    "(Task_Master,Roam)",
    "(roam,_Temp-Halloween)",
    "(Melee_Augs)",
    "(Mercenary)",
    "(Tradeskill_Quests)",
    "(Tasks)",
    "(Casino)",
    "(Quests)",
    "(Crown_Currency_Merchant)",
    "(Items_50+)",
    "(Items_5-49)",
    "(Hot_Zone_Quests)",
    "(Armor_Quests)",
    "(Wedding_Coordinator)",
    "(Tradeskill_Kits)",
    "(Wedding_Corrdinator)",
    "(Living_Legacy)",
    "(Fellowship_Vendor)",
    "(Hot_Zones)",
    "(Special_Mercenary_Liaison)",
    "(Legendary_Liaison)",
    "(Spirit_Shrouds)",
    "(Missions)",
    "(Classic_Missions)",
    "(Fellowship_Registrar)",
    "(Guild_Banners)",
    "(Mercenary_Liason)",
    "(Mercenary_Liason,Roam)",
    "(Hero`s_Forge)",
    "(Mercenary_Roster_Quest)",
    "(Tutorials)",
    "(Tutorial)",
    "(Overseer_Tetradrachms)",
    "(Group_Armor)",
    "(The_Tamrel_Trials)",
    "(Mine_Quest)",
    "(Tradeskill_Books)",
]


EXPANSION_OPTIONS = [
    (0, "Classic"),
    (1, "The Ruins of Kunark"),
    (2, "The Scars of Velious"),
    (3, "The Shadows of Luclin"),
    (4, "The Planes of Power"),
    (5, "The Legacy of Ykesha"),
    (6, "Lost Dungeons of Norrath"),
    (7, "Gates of Discord"),
    (8, "Omens of War"),
    (9, "Dragons of Norrath"),
    (10, "Depths of Darkhollow"),
    (11, "Prophecy of Ro"),
    (12, "The Serpent's Spine"),
    (13, "The Buried Sea"),
    (14, "Secrets of Faydwer"),
    (15, "Seeds of Destruction"),
    (16, "Underfoot"),
    (17, "House of Thule"),
    (18, "Veil of Alaris"),
    (19, "Rain of Fear"),
    (20, "Call of the Forsaken"),
    (21, "The Darkened Sea"),
    (22, "The Broken Mirror"),
    (23, "Empires of Kunark"),
    (24, "Ring of Scale"),
    (25, "The Burning Lands"),
    (26, "Torment of Velious"),
    (27, "Claws of Veeshan"),
    (28, "Terror of Luclin"),
    (29, "Night of Shadows"),
    (30, "Laurion's Song"),
    (31, "The Outer Brood"),
]


@dataclass
class MapAction:
    row_number: int

    source_status: str
    npc_matched: str
    zone_shortname: str
    npc_name: str

    source_file: str
    map_x: float | None
    map_y: float | None
    map_z: float | None
    map_label: str

    new_x: float | None
    new_y: float | None
    new_z: float | None
    new_label: str

    current_expansion: int
    min_expansion_number: int | None
    max_expansion_number: int | None
    match_distance: float | None

    expansion_ok: bool
    action: str
    reason: str

    matched_line_number: int | None = None
    result: str = "Pending"


def make_safe_filename(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "_", name)


def clean_text(value) -> str:
    if value is None:
        return ""

    text = str(value).strip()

    if text.lower() == "none":
        return ""

    return text


def parse_float(value):
    if value is None or value == "":
        return None

    try:
        return float(value)
    except Exception:
        return None


def parse_int(value):
    if value is None or value == "":
        return None

    try:
        return int(float(value))
    except Exception:
        return None


def fmt_coord(value) -> str:
    parsed = parse_float(value)

    if parsed is None:
        return ""

    return f"{parsed:.4f}"


def almost_equal(a, b, tolerance=0.01) -> bool:
    if a is None or b is None:
        return False

    return abs(float(a) - float(b)) <= tolerance


def expansion_is_ok(current_expansion: int, min_expansion, max_expansion) -> tuple[bool, str]:
    """
    Returns whether an NPC should exist for the requested expansion.

    Rules:
      - If min_expansion_number is blank or -1, it is treated as no minimum.
      - If max_expansion_number is blank or -1, it is treated as no maximum.
      - Delete if min_expansion_number > current expansion.
      - Delete if max_expansion_number < current expansion.
    """
    min_exp = parse_int(min_expansion)
    max_exp = parse_int(max_expansion)

    if min_exp is not None and min_exp != -1 and min_exp > current_expansion:
        return False, f"min_expansion_number {min_exp} is higher than current expansion {current_expansion}"

    if max_exp is not None and max_exp != -1 and max_exp < current_expansion:
        return False, f"max_expansion_number {max_exp} is lower than current expansion {current_expansion}"

    return True, "Expansion acceptable"


def parse_p_line(line: str):
    """
    Parses EQ map point lines:
      P x, y, z, r, g, b, size, label

    Returns dict or None.
    """
    raw = line.rstrip("\n\r")
    stripped = raw.strip()

    if not stripped.startswith("P "):
        return None

    data = stripped[2:].strip()
    parts = data.split(", ")

    if len(parts) < 8:
        return None

    return {
        "x": parse_float(parts[0]),
        "y": parse_float(parts[1]),
        "z": parse_float(parts[2]),
        "r": parts[3].strip(),
        "g": parts[4].strip(),
        "b": parts[5].strip(),
        "size": parts[6].strip(),
        "label": parts[-1].strip(),
    }


def build_p_line(x, y, z, r, g, b, size, label, newline="\n") -> str:
    return f"P {fmt_coord(x)}, {fmt_coord(y)}, {fmt_coord(z)}, {r}, {g}, {b}, {size}, {label}{newline}"


def get_header_map(ws) -> dict[str, int]:
    headers = {}

    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column

    return headers


def get_cell(row, headers, name):
    col = headers.get(name)

    if col is None:
        return None

    return row[col - 1].value


def first_present(row, headers, names):
    for name in names:
        value = get_cell(row, headers, name)

        if value is not None and str(value).strip() != "":
            return value

    return None


def get_required_columns(headers: dict[str, int]) -> list[str]:
    required = [
        "source_status",
        "npc_matched",
        "map_source_file",
        "map_x",
        "map_y",
        "map_z",
        "map_label",
        "npc_map_label",
        "min_expansion_number",
        "max_expansion_number",
    ]

    missing = [name for name in required if name not in headers]

    coord_groups = [
        ("npc_map_x", "npc_map_y", "npc_map_z"),
        ("npc_x", "npc_y", "npc_z"),
    ]

    if not any(all(name in headers for name in group) for group in coord_groups):
        missing.extend(["npc_map_x/npc_map_y/npc_map_z or npc_x/npc_y/npc_z"])

    return missing


def map_label_matches_cleanup_terms(map_label: str) -> bool:
    """
    Return True when map_label contains any configured cleanup term.

    This applies to both Map only rows and NPC + Map rows, so labels such as
    Sheley_Courilan_(Task_Master) are deleted instead of treated as valid
    matched NPC labels.
    """
    label_lower = clean_text(map_label).lower()

    if not label_lower:
        return False

    for term in MAP_ONLY_DELETE_TERMS:
        term_lower = clean_text(term).lower()
        if not term_lower:
            continue

        term_without_outer_parens = term_lower
        if term_without_outer_parens.startswith("(") and term_without_outer_parens.endswith(")"):
            term_without_outer_parens = term_without_outer_parens[1:-1]

        if term_lower in label_lower:
            return True

        if term_without_outer_parens and term_without_outer_parens in label_lower:
            return True

    return False


def make_npc_action_from_row(row, headers, excel_row_number: int, current_expansion: int) -> MapAction | None:
    source_status = clean_text(get_cell(row, headers, "source_status"))
    npc_matched = clean_text(get_cell(row, headers, "npc_matched"))

    if source_status != VALID_SOURCE_STATUS:
        return None

    if npc_matched not in VALID_MATCH_STATUSES:
        return None

    zone_shortname = clean_text(get_cell(row, headers, "zone_shortname"))
    npc_name = clean_text(get_cell(row, headers, "npc_name"))
    source_file = clean_text(get_cell(row, headers, "map_source_file"))
    map_label = clean_text(get_cell(row, headers, "map_label"))
    new_label = clean_text(get_cell(row, headers, "npc_map_label"))

    map_x = parse_float(get_cell(row, headers, "map_x"))
    map_y = parse_float(get_cell(row, headers, "map_y"))
    map_z = parse_float(get_cell(row, headers, "map_z"))

    min_exp = parse_int(get_cell(row, headers, "min_expansion_number"))
    max_exp = parse_int(get_cell(row, headers, "max_expansion_number"))
    match_distance = parse_float(get_cell(row, headers, "match_distance"))

    if "_fabled_" in npc_name.lower():
        return MapAction(
            row_number=excel_row_number,
            source_status=source_status,
            npc_matched=npc_matched,
            zone_shortname=zone_shortname,
            npc_name=npc_name,
            source_file=source_file,
            map_x=map_x,
            map_y=map_y,
            map_z=map_z,
            map_label=map_label,
            new_x=None,
            new_y=None,
            new_z=None,
            new_label="",
            current_expansion=current_expansion,
            min_expansion_number=min_exp,
            max_expansion_number=max_exp,
            match_distance=match_distance,
            expansion_ok=False,
            action="skip",
            reason="Skipped because npc_name contains _Fabled_",
            result="Skipped - Fabled NPC",
        )

    if not source_file or not map_label:
        return None

    new_x = first_present(row, headers, ["npc_map_x", "npc_x"])
    new_y = first_present(row, headers, ["npc_map_y", "npc_y"])
    new_z = first_present(row, headers, ["npc_map_z", "npc_z"])

    if not new_label:
        new_label = npc_name

    ok, reason = expansion_is_ok(current_expansion, min_exp, max_exp)

    return MapAction(
        row_number=excel_row_number,
        source_status=source_status,
        npc_matched=npc_matched,
        zone_shortname=zone_shortname,
        npc_name=npc_name,
        source_file=source_file,
        map_x=map_x,
        map_y=map_y,
        map_z=map_z,
        map_label=map_label,
        new_x=parse_float(new_x),
        new_y=parse_float(new_y),
        new_z=parse_float(new_z),
        new_label=new_label,
        current_expansion=current_expansion,
        min_expansion_number=min_exp,
        max_expansion_number=max_exp,
        match_distance=match_distance,
        expansion_ok=ok,
        action="replace" if ok else "delete",
        reason=reason,
    )


def make_cleanup_delete_action_from_row(row, headers, excel_row_number: int, current_expansion: int) -> MapAction | None:
    """
    Build a cleanup-delete action for any row whose map_label contains a cleanup term.

    This runs before normal NPC replacement/deletion logic so utility labels like
    Sheley_Courilan_(Task_Master) are deleted even when the row is source_status
    NPC + Map and npc_matched Yes.
    """
    map_label = clean_text(get_cell(row, headers, "map_label"))

    if not map_label_matches_cleanup_terms(map_label):
        return None

    source_file = clean_text(get_cell(row, headers, "map_source_file"))

    if not source_file:
        return None

    return MapAction(
        row_number=excel_row_number,
        source_status=clean_text(get_cell(row, headers, "source_status")),
        npc_matched=clean_text(get_cell(row, headers, "npc_matched")),
        zone_shortname=clean_text(get_cell(row, headers, "zone_shortname")),
        npc_name=clean_text(get_cell(row, headers, "npc_name")),
        source_file=source_file,
        map_x=parse_float(get_cell(row, headers, "map_x")),
        map_y=parse_float(get_cell(row, headers, "map_y")),
        map_z=parse_float(get_cell(row, headers, "map_z")),
        map_label=map_label,
        new_x=None,
        new_y=None,
        new_z=None,
        new_label="",
        current_expansion=current_expansion,
        min_expansion_number=parse_int(get_cell(row, headers, "min_expansion_number")),
        max_expansion_number=parse_int(get_cell(row, headers, "max_expansion_number")),
        match_distance=parse_float(get_cell(row, headers, "match_distance")),
        expansion_ok=False,
        action="cleanup_delete",
        reason="Cleanup label matched delete term",
    )


def keep_best_duplicate_npc_actions(actions: list[MapAction]) -> list[MapAction]:
    """
    If multiple actionable NPC rows have the same:
      zone_shortname + npc_name + map_x + map_y + map_z

    keep only the row with the smallest numeric match_distance.

    This prevents duplicate NPC rows from competing for the same exact map label,
    while still allowing the same npc_name to appear multiple times in a zone
    if the map point coordinates are different.

    This intentionally does NOT apply to cleanup delete actions.
    """
    grouped: dict[tuple[str, str, str, str, str], list[MapAction]] = {}
    passthrough: list[MapAction] = []

    for action in actions:
        if action.action not in {"replace", "delete"}:
            passthrough.append(action)
            continue

        if action.source_status != VALID_SOURCE_STATUS:
            passthrough.append(action)
            continue

        zone_key = action.zone_shortname.strip().lower()
        npc_key = action.npc_name.strip().lower()

        map_x_key = "" if action.map_x is None else f"{action.map_x:.4f}"
        map_y_key = "" if action.map_y is None else f"{action.map_y:.4f}"
        map_z_key = "" if action.map_z is None else f"{action.map_z:.4f}"

        if not zone_key or not npc_key or not map_x_key or not map_y_key or not map_z_key:
            passthrough.append(action)
            continue

        key = (
            zone_key,
            npc_key,
            map_x_key,
            map_y_key,
            map_z_key,
        )

        grouped.setdefault(key, []).append(action)

    selected: list[MapAction] = []

    for key, group in grouped.items():
        if len(group) == 1:
            selected.append(group[0])
            continue

        def sort_key(action: MapAction):
            distance = action.match_distance

            if distance is None:
                distance = float("inf")

            return distance, action.row_number

        best = min(group, key=sort_key)
        selected.append(best)

        for action in group:
            if action is best:
                continue

            best_distance = "blank" if best.match_distance is None else best.match_distance
            this_distance = "blank" if action.match_distance is None else action.match_distance

            action.action = "duplicate_skip"
            action.result = "Skipped - duplicate NPC/map point with larger match_distance"
            action.reason = (
                f"Duplicate zone+npc_name+map_x+map_y+map_z; row {best.row_number} "
                f"was used because match_distance {best_distance} is smaller than "
                f"{this_distance}"
            )
            passthrough.append(action)

    return passthrough + selected


def load_actions_from_workbook(workbook_path: Path, current_expansion: int) -> list[MapAction]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb.active
    headers = get_header_map(ws)

    missing = get_required_columns(headers)

    if missing:
        raise ValueError("Workbook is missing required columns: " + ", ".join(missing))

    actions: list[MapAction] = []

    for excel_row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Cleanup has priority over normal NPC replacement. This ensures
        # NPC + Map rows like Sheley_Courilan_(Task_Master) are deleted rather
        # than replaced/kept just because they matched a real NPC.
        cleanup_action = make_cleanup_delete_action_from_row(row, headers, excel_row_number, current_expansion)

        if cleanup_action is not None:
            actions.append(cleanup_action)
            continue

        npc_action = make_npc_action_from_row(row, headers, excel_row_number, current_expansion)

        if npc_action is not None:
            actions.append(npc_action)

    return keep_best_duplicate_npc_actions(actions)


def action_matches_line(action: MapAction, parsed_line) -> bool:
    if parsed_line is None:
        return False

    if parsed_line["label"] != action.map_label:
        return False

    return (
        almost_equal(parsed_line["x"], action.map_x)
        and almost_equal(parsed_line["y"], action.map_y)
        and almost_equal(parsed_line["z"], action.map_z)
    )


def process_map_file(map_file: Path, actions: list[MapAction], backup_folder: Path) -> tuple[int, int, int, int, int]:
    """
    Returns:
      replaced_count,
      expansion_deleted_count,
      map_only_cleanup_deleted_count,
      not_found_count,
      skipped_count
    """
    if not map_file.exists():
        for action in actions:
            action.result = "Map source file not found"

        return 0, 0, 0, len(actions), 0

    relative_backup_path = backup_folder / map_file.name
    shutil.copy2(map_file, relative_backup_path)

    original_lines = map_file.read_text(encoding="utf-8", errors="replace").splitlines(keepends=True)
    new_lines = []

    unused_actions = [action for action in actions if action.action not in {"skip", "duplicate_skip"}]

    replaced_count = 0
    expansion_deleted_count = 0
    map_only_cleanup_deleted_count = 0
    skipped_count = 0

    for line_number, line in enumerate(original_lines, start=1):
        parsed = parse_p_line(line)

        match_index = None

        for idx, action in enumerate(unused_actions):
            if action_matches_line(action, parsed):
                match_index = idx
                break

        if match_index is None:
            new_lines.append(line)
            continue

        action = unused_actions.pop(match_index)
        action.matched_line_number = line_number

        if action.action == "delete":
            action.result = "Deleted"
            expansion_deleted_count += 1
            continue

        if action.action == "cleanup_delete":
            action.result = "Deleted - cleanup term"
            map_only_cleanup_deleted_count += 1
            continue

        if action.action == "replace":
            if action.new_x is None or action.new_y is None or action.new_z is None or not action.new_label:
                action.result = "Skipped - missing replacement coordinate or label"
                skipped_count += 1
                new_lines.append(line)
                continue

            newline = "\r\n" if line.endswith("\r\n") else "\n"
            new_line = build_p_line(
                action.new_x,
                action.new_y,
                action.new_z,
                parsed["r"],
                parsed["g"],
                parsed["b"],
                parsed["size"],
                action.new_label,
                newline=newline,
            )
            new_lines.append(new_line)
            action.result = "Replaced"
            replaced_count += 1
            continue

        action.result = f"Skipped - unsupported action {action.action}"
        skipped_count += 1
        new_lines.append(line)

    for action in unused_actions:
        action.result = "Matching map point line not found"

    map_file.write_text("".join(new_lines), encoding="utf-8", errors="replace")

    return (
        replaced_count,
        expansion_deleted_count,
        map_only_cleanup_deleted_count,
        len(unused_actions),
        skipped_count,
    )


def write_summary(summary_path: Path, actions: list[MapAction]):
    fieldnames = [
        "excel_row_number",
        "source_status",
        "npc_matched",
        "zone_shortname",
        "npc_name",
        "map_source_file",
        "matched_line_number",
        "action",
        "result",
        "reason",
        "current_expansion",
        "min_expansion_number",
        "max_expansion_number",
        "match_distance",
        "map_label",
        "new_label",
        "old_map_x",
        "old_map_y",
        "old_map_z",
        "new_x",
        "new_y",
        "new_z",
    ]

    with summary_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for action in actions:
            writer.writerow({
                "excel_row_number": action.row_number,
                "source_status": action.source_status,
                "npc_matched": action.npc_matched,
                "zone_shortname": action.zone_shortname,
                "npc_name": action.npc_name,
                "map_source_file": action.source_file,
                "matched_line_number": action.matched_line_number or "",
                "action": action.action,
                "result": action.result,
                "reason": action.reason,
                "current_expansion": action.current_expansion,
                "min_expansion_number": action.min_expansion_number if action.min_expansion_number is not None else "",
                "max_expansion_number": action.max_expansion_number if action.max_expansion_number is not None else "",
                "match_distance": action.match_distance if action.match_distance is not None else "",
                "map_label": action.map_label,
                "new_label": action.new_label,
                "old_map_x": fmt_coord(action.map_x),
                "old_map_y": fmt_coord(action.map_y),
                "old_map_z": fmt_coord(action.map_z),
                "new_x": fmt_coord(action.new_x),
                "new_y": fmt_coord(action.new_y),
                "new_z": fmt_coord(action.new_z),
            })


def print_expansion_options():
    print("Expansion options:")

    for number, name in EXPANSION_OPTIONS:
        print(f"  {number:>2} = {name}")


def get_current_expansion_from_user_or_args(argv) -> int:
    if len(argv) >= 4:
        value = argv[3]
    else:
        print()
        print("Enter your current expansion number.")
        print_expansion_options()
        value = input("Current expansion number: ").strip()

    try:
        return int(value)
    except Exception:
        raise ValueError(f"Invalid current expansion number: {value}")


def main():
    if len(sys.argv) < 3:
        print("Usage:")
        print("python apply_combined_map_data.py <map_folder> <script_folder> [current_expansion]")
        print()
        print("Usually you should run this by dragging your map folder onto the batch file.")
        input("Press Enter to exit...")
        sys.exit(1)

    map_folder = Path(sys.argv[1])
    script_folder = Path(sys.argv[2])
    current_expansion = get_current_expansion_from_user_or_args(sys.argv)

    if not map_folder.exists() or not map_folder.is_dir():
        raise FileNotFoundError(f"Invalid map folder: {map_folder}")

    if not script_folder.exists() or not script_folder.is_dir():
        raise FileNotFoundError(f"Invalid script/output folder: {script_folder}")

    dragged_folder_name = make_safe_filename(map_folder.name)
    workbook_path = script_folder / f"{dragged_folder_name} Combined Map Data.xlsx"

    if not workbook_path.exists():
        raise FileNotFoundError(
            f"Could not find expected workbook:\n{workbook_path}\n\n"
            "The workbook name must match the dragged folder name exactly: "
            f"{dragged_folder_name} Combined Map Data.xlsx"
        )

    actions = load_actions_from_workbook(workbook_path, current_expansion)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_folder = script_folder / f"{dragged_folder_name} map backup before expansion apply {timestamp}"
    backup_folder.mkdir(parents=True, exist_ok=True)

    actions_by_file: dict[str, list[MapAction]] = {}

    for action in actions:
        if action.action in {"skip", "duplicate_skip"}:
            continue

        actions_by_file.setdefault(action.source_file, []).append(action)

    total_replaced = 0
    total_expansion_deleted = 0
    total_map_only_cleanup_deleted = 0
    total_not_found = 0
    total_skipped_missing_data = 0

    for source_file, file_actions in sorted(actions_by_file.items()):
        map_file = map_folder / source_file

        replaced, expansion_deleted, map_only_cleanup_deleted, not_found, skipped_missing_data = process_map_file(
            map_file,
            file_actions,
            backup_folder,
        )

        total_replaced += replaced
        total_expansion_deleted += expansion_deleted
        total_map_only_cleanup_deleted += map_only_cleanup_deleted
        total_not_found += not_found
        total_skipped_missing_data += skipped_missing_data

    summary_path = script_folder / f"{dragged_folder_name} Expansion Apply Summary.csv"
    write_summary(summary_path, actions)

    total_actions = len(actions)

    skipped_fabled = sum(
        1 for action in actions
        if action.result == "Skipped - Fabled NPC"
    )

    skipped_duplicates = sum(
        1 for action in actions
        if action.result == "Skipped - duplicate NPC/map point with larger match_distance"
    )

    missing_files = sum(
        1 for action in actions
        if action.result == "Map source file not found"
    )

    total_deleted = total_expansion_deleted + total_map_only_cleanup_deleted

    print()
    print("Done.")
    print(f"Map folder: {map_folder}")
    print(f"Workbook used: {workbook_path}")
    print(f"Current expansion: {current_expansion}")
    print(f"Backup folder: {backup_folder}")
    print(f"Summary CSV: {summary_path}")
    print()
    print(f"Workbook rows considered: {total_actions}")
    print(f"Lines replaced with NPC coordinates/labels: {total_replaced}")
    print(f"Lines deleted due to expansion rules: {total_expansion_deleted}")
    print(f"Lines deleted due to cleanup terms: {total_map_only_cleanup_deleted}")
    print(f"Total deleted lines: {total_deleted}")
    print(f"Matching lines not found: {total_not_found}")
    print(f"Skipped due to missing replacement data: {total_skipped_missing_data}")
    print(f"Duplicate NPC/map-point rows skipped by larger match_distance: {skipped_duplicates}")
    print(f"Fabled NPC rows skipped: {skipped_fabled}")
    print(f"Rows with missing map source file: {missing_files}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print()
        print("ERROR:")
        print(exc)
        print()
        input("Press Enter to exit...")
        sys.exit(1)

    print()
    input("Press Enter to exit...")